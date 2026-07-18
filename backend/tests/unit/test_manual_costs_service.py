from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from types import SimpleNamespace
from unittest.mock import AsyncMock

import pytest
from starlette.datastructures import UploadFile

from app.core.time import utcnow
from app.models.manual_costs import ManualCost
from app.repositories.manual_costs import ManualCostRepository
from app.services.manual_costs import ManualCostService


class _FakeExecuteResult:
    def scalars(self):
        return []


class _CaptureExecuteSession:
    def __init__(self) -> None:
        self.statement = None

    async def execute(self, stmt):
        self.statement = stmt
        return _FakeExecuteResult()


class _RowsResult:
    def __init__(self, rows=None, scalars=None):
        self._rows = rows or []
        self._scalars = scalars or []

    def all(self):
        return self._rows

    def scalars(self):
        return self._scalars


def test_parse_csv_cost_file() -> None:
    service = ManualCostService()
    payload = (
        "vendorCode,unitCost,nmId,barcode,techSize,currency,validFrom,validTo,comment\n"
        "SKU-1,123.45,1001,460000000001,42,RUB,2026-05-01,2026-06-01,test row\n"
    ).encode("utf-8")
    file = UploadFile(filename="costs.csv", file=io.BytesIO(payload))

    rows = service._parse_csv(payload)

    assert len(rows) == 1
    assert rows[0]["vendor_code"] == "SKU-1"
    assert rows[0]["unit_cost"] == Decimal("123.45")
    assert rows[0]["nm_id"] == 1001
    assert rows[0]["currency"] == "RUB"


def test_parse_csv_extended_cost_columns() -> None:
    service = ManualCostService()
    payload = (
        "vendorCode,costPrice,packagingCost,inboundLogisticsCost,nmId,barcode,techSize,supplier,currency,validFrom,validTo,comment\n"
        "SKU-2,200,5.5,7.25,2002,460000000002,43,Best Supplier,USD,,,extended row\n"
    ).encode("utf-8")
    file = UploadFile(filename="costs.csv", file=io.BytesIO(payload))

    rows = service._parse_csv(payload)

    assert file.filename == "costs.csv"
    assert rows[0]["cost_price"] == Decimal("200")
    assert rows[0]["unit_cost"] == Decimal("200")
    assert rows[0]["seller_other_expense"] == Decimal("12.75")
    assert rows[0]["packaging_cost"] == Decimal("5.5")
    assert rows[0]["inbound_logistics_cost"] == Decimal("7.25")
    assert rows[0]["supplier"] == "Best Supplier"
    assert rows[0]["valid_from"] == utcnow().date()
    assert rows[0]["valid_to"] is None


def test_parse_csv_new_seller_other_expense_format() -> None:
    service = ManualCostService()
    payload = (
        "vendorCode,nmId,barcode,techSize,cost_price,seller_other_expense,supplier,valid_from,comment\n"
        "SKU-3,3003,460000000003,44,320,18.4,Seller Three,2026-05-10,new format row\n"
    ).encode("utf-8")

    rows = service._parse_csv(payload)

    assert len(rows) == 1
    assert rows[0]["vendor_code"] == "SKU-3"
    assert rows[0]["cost_price"] == Decimal("320")
    assert rows[0]["unit_cost"] == Decimal("320")
    assert rows[0]["seller_other_expense"] == Decimal("18.4")
    assert rows[0]["packaging_cost"] == Decimal("0")
    assert rows[0]["inbound_logistics_cost"] == Decimal("0")
    assert rows[0]["valid_from"] == date(2026, 5, 10)


def test_operator_baseline_supplier_does_not_count_as_business_trusted() -> None:
    is_placeholder, is_business_trusted, cost_source = ManualCostService._derive_cost_metadata(
        {"supplier": "OPERATOR_TRUSTED_COST"}
    )

    assert is_placeholder is False
    assert is_business_trusted is False
    assert cost_source == "operator_trusted_manual"


def test_generic_uploaded_cost_defaults_to_operator_trusted_manual() -> None:
    is_placeholder, is_business_trusted, cost_source = ManualCostService._derive_cost_metadata(
        {"supplier": "Regular Supplier"}
    )

    assert is_placeholder is False
    assert is_business_trusted is True
    assert cost_source == "operator_trusted_manual"


def test_apply_supplier_confirmation_flags_marks_cost_as_supplier_confirmed() -> None:
    cost = ManualCost(
        account_id=1,
        dedupe_key="cost-1",
        vendor_code="SKU-1",
        unit_cost=Decimal("100"),
        cost_price=Decimal("100"),
    )

    ManualCostService._apply_supplier_confirmation_flags(
        cost=cost,
        is_supplier_confirmed=True,
        user_id=7,
    )

    assert cost.is_supplier_confirmed is True
    assert cost.cost_source == "supplier_confirmed"
    assert cost.supplier_confirmed_by_user_id == 7
    assert cost.supplier_confirmed_at is not None


@pytest.mark.asyncio
async def test_mark_supplier_confirmed_updates_existing_manual_cost() -> None:
    service = ManualCostService()
    cost = ManualCost(
        id=5,
        account_id=1,
        dedupe_key="cost-5",
        vendor_code="SKU-5",
        unit_cost=Decimal("200"),
        cost_price=Decimal("200"),
        cost_source="operator_trusted_manual",
        is_business_trusted=True,
    )
    session = SimpleNamespace(
        get=AsyncMock(return_value=cost),
        flush=AsyncMock(),
    )

    result = await service.mark_supplier_confirmed(
        session,
        cost_id=5,
        user_id=11,
        comment="Supplier invoice received",
    )

    assert result.is_supplier_confirmed is True
    assert result.cost_source == "supplier_confirmed"
    assert result.supplier_confirmed_by_user_id == 11
    assert "Supplier invoice received" in str(result.comment)


@pytest.mark.asyncio
async def test_build_template_csv_uses_new_primary_headers() -> None:
    service = ManualCostService()
    service._template_rows = AsyncMock(
        return_value=[
            {
                "vendor_code": "SKU-10",
                "nm_id": 1010,
                "barcode": "460000000010",
                "tech_size": "45",
                "product_title": "Template row",
                "current_cost_price": Decimal("100"),
                "current_seller_other_expense": Decimal("12"),
            }
        ]
    )

    csv_text = await service.build_template_csv(SimpleNamespace(), account_id=1)
    lines = csv_text.strip().splitlines()

    assert lines[0] == (
        "vendorCode,nmId,barcode,techSize,productTitle,current_cost_price,"
        "current_seller_other_expense,cost_price,seller_other_expense,supplier,valid_from,comment"
    )
    assert lines[1].startswith("SKU-10,1010,460000000010,45,Template row,100,12,,,,")


@pytest.mark.asyncio
async def test_build_template_xlsx_includes_instruction_sheet() -> None:
    from openpyxl import load_workbook

    service = ManualCostService()
    service._template_rows = AsyncMock(
        return_value=[
            {
                "vendor_code": "SKU-10",
                "nm_id": 1010,
                "barcode": "460000000010",
                "tech_size": "45",
                "product_title": "Template row",
                "current_cost_price": "",
                "current_seller_other_expense": "",
            }
        ]
    )

    content = await service.build_template_xlsx(SimpleNamespace(), account_id=1, mode="missing")
    workbook = load_workbook(io.BytesIO(content))

    assert "Инструкция" in workbook.sheetnames
    assert workbook["Себестоимость"].freeze_panes == "A2"
    assert workbook["Себестоимость"]["A1"].value == "vendorCode"


@pytest.mark.asyncio
async def test_list_missing_costs_filters_to_revenue_skus_and_summarizes_coverage() -> None:
    service = ManualCostService()
    service._current_costs_by_sku = AsyncMock(return_value={2: SimpleNamespace(id=20)})
    missing_sku = SimpleNamespace(
        id=1,
        nm_id=1001,
        vendor_code="SKU-1",
        barcode="BC-1",
        tech_size="42",
        title="Missing cost product",
    )
    covered_sku = SimpleNamespace(
        id=2,
        nm_id=1002,
        vendor_code="SKU-2",
        barcode="BC-2",
        tech_size="44",
        title="Covered product",
    )
    no_revenue_sku = SimpleNamespace(
        id=3,
        nm_id=1003,
        vendor_code="SKU-3",
        barcode="BC-3",
        tech_size="46",
        title="No revenue product",
    )
    session = SimpleNamespace(
        execute=AsyncMock(
            side_effect=[
                _RowsResult(rows=[(1, Decimal("25"))]),
                _RowsResult(rows=[(1, Decimal("25")), (2, Decimal("75"))]),
                _RowsResult(scalars=[missing_sku, covered_sku, no_revenue_sku]),
            ]
        )
    )

    result = await service.list_missing_costs(
        session,
        account_id=1,
        date_from=date(2026, 6, 1),
        date_to=date(2026, 6, 30),
        only_revenue=True,
    )

    assert result["total"] == 1
    assert result["summary"]["missing_sku_count"] == 1
    assert result["summary"]["affected_revenue"] == 25.0
    assert result["summary"]["revenue_cost_coverage_percent"] == 75.0
    assert result["items"][0]["sku_id"] == 1
    assert result["items"][0]["recommended_action"] == "Заполнить себестоимость"


@pytest.mark.asyncio
async def test_list_unresolved_costs_page_uses_repository_level_pagination() -> None:
    service = ManualCostService()
    row = ManualCost(
        id=9,
        account_id=1,
        dedupe_key="cost-9",
        vendor_code="SKU-9",
        unit_cost=Decimal("50"),
        cost_price=Decimal("50"),
    )
    expected_page = SimpleNamespace(total=1, limit=50, offset=0, items=[row])
    service.costs.list_unresolved_page = AsyncMock(return_value=expected_page)

    page = await service.list_unresolved_costs_page(
        SimpleNamespace(),
        account_id=1,
        limit=50,
        offset=0,
    )

    assert page is expected_page
    service.costs.list_unresolved_page.assert_awaited_once()


@pytest.mark.asyncio
async def test_list_unresolved_costs_for_product_uses_exact_repository_lookup() -> None:
    service = ManualCostService()
    row = ManualCost(
        id=10,
        account_id=1,
        dedupe_key="cost-10",
        vendor_code="SKU-10",
        nm_id=1010,
        barcode="BC-10",
        unit_cost=Decimal("50"),
        cost_price=Decimal("50"),
    )
    service.costs.list_unresolved_for_product = AsyncMock(return_value=[row])
    session = SimpleNamespace()

    rows = await service.list_unresolved_costs_for_product(
        session,
        account_id=1,
        nm_id=1010,
        sku_id=501,
        vendor_code="SKU-10",
        barcode="BC-10",
        limit=20,
    )

    assert rows == [row]
    service.costs.list_unresolved_for_product.assert_awaited_once_with(
        session,
        account_id=1,
        nm_id=1010,
        sku_id=501,
        vendor_code="SKU-10",
        barcode="BC-10",
        limit=20,
    )


@pytest.mark.asyncio
async def test_unresolved_cost_repository_exact_lookup_filters_product_identifiers() -> None:
    repo = ManualCostRepository()
    session = _CaptureExecuteSession()

    rows = await repo.list_unresolved_for_product(
        session,
        account_id=1,
        nm_id=1010,
        sku_id=501,
        vendor_code="SKU-10",
        barcode="BC-10",
        limit=20,
    )

    sql = str(session.statement.compile(compile_kwargs={"literal_binds": True}))
    assert rows == []
    assert "manual_costs.account_id = 1" in sql
    assert "manual_costs.nm_id = 1010" in sql
    assert "manual_costs.sku_id = 501" in sql
    assert "manual_costs.vendor_code = 'SKU-10'" in sql
    assert "manual_costs.barcode = 'BC-10'" in sql
    assert "manual_costs.sku_id IS NULL" in sql
    assert "manual_costs.is_ambiguous IS true" in sql


@pytest.mark.asyncio
async def test_import_costs_summary_tracks_legacy_manual_cost_fields_warning() -> None:
    service = ManualCostService()
    payload = (
        "vendorCode,costPrice,packagingCost,inboundLogisticsCost,nmId,barcode,techSize,supplier,validFrom,comment\n"
        "SKU-LEGACY,150,5,7,4001,460000000401,41,Supplier A,2026-05-01,legacy row\n"
    ).encode("utf-8")
    file = UploadFile(filename="costs.csv", file=io.BytesIO(payload))
    sku_rows = [
        SimpleNamespace(id=1, nm_id=4001, vendor_code="SKU-LEGACY", barcode="460000000401", tech_size="41"),
    ]
    session = SimpleNamespace(
        add=lambda _obj: None,
        flush=AsyncMock(),
        execute=AsyncMock(
            side_effect=[
                SimpleNamespace(scalar_one_or_none=lambda: {"require_seller_other_expense": True}),
                SimpleNamespace(scalars=lambda: sku_rows),
            ]
        ),
    )
    service._commit_valid_rows = AsyncMock(return_value=1)

    upload, preview = await service.import_costs(
        session,
        account_id=1,
        created_by_user_id=5,
        file=file,
        commit_rows=True,
    )

    assert upload.summary["legacyFieldMappedRows"] == 1
    assert upload.summary["sellerOtherExpenseMissingRows"] == 0
    assert "manual_cost_old_fields_used" in upload.summary["warningCodes"]
    assert preview[0]["used_legacy_cost_fields"] is True


@pytest.mark.asyncio
async def test_import_costs_summary_tracks_missing_seller_other_expense_warning() -> None:
    service = ManualCostService()
    payload = (
        "vendorCode,nmId,barcode,techSize,cost_price,supplier,valid_from,comment\n"
        "SKU-MISSING,4002,460000000402,42,180,Supplier B,2026-05-01,missing seller other\n"
    ).encode("utf-8")
    file = UploadFile(filename="costs.csv", file=io.BytesIO(payload))
    sku_rows = [
        SimpleNamespace(id=2, nm_id=4002, vendor_code="SKU-MISSING", barcode="460000000402", tech_size="42"),
    ]
    session = SimpleNamespace(
        add=lambda _obj: None,
        flush=AsyncMock(),
        execute=AsyncMock(
            side_effect=[
                SimpleNamespace(scalar_one_or_none=lambda: {"require_seller_other_expense": True}),
                SimpleNamespace(scalars=lambda: sku_rows),
            ]
        ),
    )
    service._commit_valid_rows = AsyncMock(return_value=1)

    upload, preview = await service.import_costs(
        session,
        account_id=1,
        created_by_user_id=5,
        file=file,
        commit_rows=True,
    )

    assert upload.summary["legacyFieldMappedRows"] == 0
    assert upload.summary["sellerOtherExpenseMissingRows"] == 1
    assert "seller_other_expense_missing" in upload.summary["warningCodes"]
    assert preview[0]["seller_other_expense_missing"] is True
