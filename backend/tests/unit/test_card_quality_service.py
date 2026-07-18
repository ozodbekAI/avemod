from __future__ import annotations

from dataclasses import replace
from datetime import datetime, timezone
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import AsyncMock, Mock

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import BigInteger, create_engine, select
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.ext.compiler import compiles
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool
from sqlalchemy.schema import CreateTable

from app.models.accounts import WBAccount
from app.models.auth import AuthUser
from app.models.card_quality import CardQualityAnalysisRun, CardQualityFixedFileEntry, CardQualityIssue, CardQualitySnapshot
from app.models.operator import ResultEvent
from app.models.product_cards import WBProductCard
from app.schemas.card_quality import CardQualityFixedFileEntryMutation, CardQualityIssueRead, CardQualityIssueStatusUpdate
from app.services.card_quality import (
    CardQualityAnalysisService,
    CardQualityNormalizationService,
    CardQualityRuleEngine,
    NormalizedCard,
    _sanitize_wb_update_snapshot,
)
from app.services.checker_core.wb_validator import get_catalog, validate_card_characteristics


@compiles(JSONB, "sqlite")
def _compile_jsonb_sqlite(_type, _compiler, **_kw) -> str:
    return "JSON"


@compiles(BigInteger, "sqlite")
def _compile_bigint_sqlite(_type, _compiler, **_kw) -> str:
    return "INTEGER"


VALID_SOURCE_DESCRIPTION = (
    "Костюм женский жакет и брюки повседневный подходит для офиса, поездок и спокойных городских сценариев. "
    "Комплект состоит из жакета и брюк, поэтому его можно носить вместе или разделять с базовыми вещами гардероба. "
    "Описание опирается на характеристики карточки и не добавляет неподтвержденные свойства товара.\n\n"
    "Конструкция комплекта рассчитана на аккуратную посадку без лишнего объема. Жакет формирует верх комплекта, "
    "брюки отвечают за нижнюю часть и помогают собрать цельный образ. Силуэт, крой, тип верха и тип низа важно "
    "проверять по характеристикам и фотографиям перед публикацией изменений.\n\n"
    "Материал указан в характеристиках как хлопок, поэтому в тексте используется только подтвержденный состав без "
    "добавления процентов, если они не переданы в карточке. Ткань и состав лучше сверять с поставщиком, fixed-file "
    "и карточкой WB, чтобы описание не расходилось с фильтрами и фактическими данными.\n\n"
    "Назначение комплекта повседневное: он подходит для работы, прогулок, поездок и ситуаций, где нужен собранный "
    "комплект без сложной стилизации. Покупателю важно видеть понятное применение, особенности модели, посадку, "
    "комплектацию и ограничения ухода, если такие данные подтверждены.\n\n"
    "Перед отправкой на WB нужно сверить название, описание, характеристики, фото и видео. Если меняются материал, "
    "комплектация, модель брюк или тип верха, описание следует перечитать заново, чтобы текст оставался связанным "
    "с карточкой и не превращался в набор общих обещаний."
)


def test_card_quality_subject_keywords_are_loaded_from_wb_catalog(monkeypatch) -> None:
    class Catalog:
        def get_keywords_for_subject(self, subject: str) -> list[str]:
            assert subject == "Костюмы"
            return ["костюм", "жакет"]

    monkeypatch.setattr("app.services.card_quality.get_catalog", lambda: Catalog())

    service = CardQualityAnalysisService()

    assert service._subject_keywords({"subjectName": "Костюмы"}) == ["костюм", "жакет"]


def test_card_quality_no_touch_characteristics_are_not_reported_by_wb_catalog() -> None:
    rows = validate_card_characteristics(
        {
            "subjectID": 215,
            "subjectName": "Костюмы",
            "characteristics": [{"id": 12, "name": "Цвет", "value": ["черный"]}],
        }
    )

    names = {str(row.get("name") or "") for row in rows}

    assert "ИКПУ" not in names
    assert "Артикул OZON" not in names
    assert "Код упаковки" not in names
    assert "Любимые герои" not in names


def test_card_quality_generation_dictionary_values_are_exposed() -> None:
    values = get_catalog().get_allowed_values("Модель костюма") or []

    assert "брючный" in values


def test_card_quality_conditional_fill_characteristic_is_checked_when_active() -> None:
    card = {"characteristics": [{"name": "Тип низа", "value": ["брюки"]}]}

    assert get_catalog().is_no_touch_characteristic("Модель брюк", card=card) is False


def test_card_quality_conditional_fill_characteristic_is_skipped_when_inactive() -> None:
    card = {"characteristics": [{"name": "Тип низа", "value": ["юбка"]}]}

    assert get_catalog().is_no_touch_characteristic("Модель брюк", card=card) is True


def test_card_quality_fixed_file_ignores_no_touch_characteristics() -> None:
    service = CardQualityAnalysisService()
    card = _card(
        characteristics=[
            {"name": "ИКПУ", "value": "old"},
            {"name": "Артикул OZON", "value": "old"},
        ]
    )

    issues = service._apply_fixed_file_priority(card, [], {"ИКПУ": "new", "Артикул OZON": "new"})

    assert issues == []


def test_card_quality_fixed_file_only_characteristics_are_not_ai_candidates() -> None:
    service = CardQualityAnalysisService()
    card = _card(
        characteristics=[
            {"name": "Цвет", "value": ["черный"]},
            {"name": "Состав", "value": ""},
        ]
    )

    issues = service.rules._characteristic_rules(card)

    assert all("Состав" not in issue.field_name for issue in issues)
    assert service._skip_ai_audit_issue({"name": "Состав", "message": "Проверить состав"}) is True


def test_card_quality_ai_audit_skips_no_touch_characteristics() -> None:
    service = CardQualityAnalysisService()

    assert service._skip_ai_audit_issue({"name": "ИКПУ", "message": "Заполнить ИКПУ"}) is True
    assert service._skip_ai_audit_issue({"name": "Артикул OZON", "message": "Нет значения"}) is True


class _SQLiteAsyncSessionAdapter:
    def __init__(self, sync_session: Session):
        self._session = sync_session

    async def execute(self, statement):
        return self._session.execute(statement)

    async def get(self, model, ident):
        return self._session.get(model, ident)

    def add(self, instance) -> None:
        self._session.add(instance)

    async def flush(self) -> None:
        self._session.flush()

    async def commit(self) -> None:
        self._session.commit()

    async def rollback(self) -> None:
        self._session.rollback()

    async def refresh(self, instance) -> None:
        self._session.refresh(instance)


def _card_quality_recheck_session() -> tuple[Session, _SQLiteAsyncSessionAdapter]:
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    with engine.begin() as conn:
        for table in (
            WBAccount.__table__,
            AuthUser.__table__,
            CardQualityAnalysisRun.__table__,
            CardQualitySnapshot.__table__,
            CardQualityIssue.__table__,
            CardQualityFixedFileEntry.__table__,
            ResultEvent.__table__,
        ):
            conn.execute(CreateTable(table, if_not_exists=True))
    sync_session = Session(engine)
    sync_session.add(WBAccount(id=1, name="Checker test account", timezone="Europe/Moscow", is_active=True))
    sync_session.flush()
    return sync_session, _SQLiteAsyncSessionAdapter(sync_session)


def _card(**overrides) -> NormalizedCard:
    values = {
        "account_id": 1,
        "nm_id": 245405620,
        "source_card_id": 10,
        "title": "Костюм женский жакет и брюки повседневный",
        "description": VALID_SOURCE_DESCRIPTION,
        "brand": "Avemod",
        "subject_name": "Костюмы",
        "vendor_code": "AV-1",
        "characteristics": [{"name": "Цвет", "value": ["черный"]}, {"name": "Состав", "value": ["хлопок"]}],
        "photos": [
            "https://example.test/1.jpg",
            "https://example.test/2.jpg",
            "https://example.test/3.jpg",
            "https://example.test/4.jpg",
            "https://example.test/5.jpg",
            "https://example.test/6.jpg",
        ],
        "videos": ["https://example.test/video.mp4"],
        "sizes": [],
        "source_revision": "rev",
        "source_updated_at": None,
    }
    values.update(overrides)
    return NormalizedCard(**values)


def test_card_quality_clean_card_keeps_high_score() -> None:
    issues, summary = CardQualityRuleEngine().analyze(_card())

    non_info = [issue for issue in issues if issue.severity != "info"]

    assert non_info == []
    assert summary["score"] == 100
    assert summary["status"] == "clean"
    assert summary["category_scores"]["photos"] == 100


@pytest.mark.asyncio
async def test_ai_compound_characteristic_exposes_candidate_and_allowed_values() -> None:
    service = CardQualityAnalysisService()
    service.ai_fixer = SimpleNamespace(
        is_enabled=True,
        audit_card=AsyncMock(
            return_value=[
                {
                    "name": "Вид застежки",
                    "category": "characteristics",
                    "severity": "medium",
                    "message": "На фото видна пуговица, а указана только молния.",
                    "fix_action": "compound",
                    "errors": [
                        {
                            "type": "vision_mismatch",
                            "message": "Значение только 'молния' неполное.",
                        }
                    ],
                    "compound_fixes": [
                        {
                            "name": "Вид застежки",
                            "value": ["пуговицы", "молния"],
                            "action": "replace",
                        }
                    ],
                }
            ]
        ),
    )

    issues = await service._apply_ai_audit(
        _card(characteristics=[{"name": "Вид застежки", "value": ["молния"]}]),
        [],
        {},
    )

    issue = next(item for item in issues if item.field_name == "characteristics.Вид застежки")
    assert issue.alternatives == ["пуговицы, молния"]
    assert issue.ai_alternatives == ["пуговицы, молния"]
    assert issue.expected_value_json["candidate_values"] == ["пуговицы, молния"]
    assert "пуговицы" in issue.allowed_values
    assert "молния" in issue.allowed_values


def test_card_quality_missing_core_fields_is_critical_and_bounded() -> None:
    issues, summary = CardQualityRuleEngine().analyze(
        _card(title="", description="", characteristics=[], photos=[], videos=[], brand="", subject_name="")
    )

    assert {issue.issue_code for issue in issues} >= {
        "no_title",
        "no_description",
        "no_photos",
        "no_video",
    }
    assert summary["status"] == "critical"
    assert 0 <= summary["score"] <= 100
    assert summary["critical_count"] >= 2


def test_card_quality_fingerprint_is_stable() -> None:
    engine = CardQualityRuleEngine()
    card = _card(title="")
    issues, _summary = engine.analyze(card)
    issue = next(item for item in issues if item.issue_code == "no_title")

    assert engine.fingerprint(card, issue) == engine.fingerprint(card, issue)


def test_card_quality_video_absence_matches_source_warning() -> None:
    issues, summary = CardQualityRuleEngine().analyze(_card(videos=[]))

    video_issue = next(issue for issue in issues if issue.issue_code == "no_video")

    assert video_issue.severity == "medium"
    assert summary["status"] == "warning"


def test_card_quality_logical_photo_count_uses_photo_objects() -> None:
    service = CardQualityNormalizationService()

    photos = service._extract_photos(
        [
            {
                "big": "https://cdn.example.test/1-big.webp",
                "c516x688": "https://cdn.example.test/1-c516.webp",
                "tm": "https://cdn.example.test/1-tm.webp",
            }
        ],
        {},
    )

    assert len(photos) == 1
    assert photos[0] == {
        "canonical_url": "https://cdn.example.test/1-big.webp",
        "variants": {
            "big": "https://cdn.example.test/1-big.webp",
            "c516x688": "https://cdn.example.test/1-c516.webp",
            "tm": "https://cdn.example.test/1-tm.webp",
        },
    }


def test_card_quality_few_photos_matches_source_rule() -> None:
    issues, _summary = CardQualityRuleEngine().analyze(_card(photos=[{"canonical_url": "https://example.test/1.jpg", "variants": {}}]))

    assert "few_photos" in {issue.issue_code for issue in issues}


def test_card_quality_action_preserves_actual_issue_severity() -> None:
    issue = CardQualityIssue(
        id=55,
        account_id=1,
        nm_id=245405620,
        issue_code="media_no_images",
        category="media",
        severity="critical",
        title="No images",
        business_explanation="Images are required.",
        recommended_fix="Add product photos.",
        status="new",
        fingerprint="fp",
        first_seen_at=datetime.now(timezone.utc),
        last_seen_at=datetime.now(timezone.utc),
    )

    action = CardQualityAnalysisService()._action_from_issue(account_id=1, issue=issue)

    assert action.severity == "critical"
    assert action.can_execute is False


def test_card_quality_action_contract_uses_issue_status_payload_and_estimated_opportunity() -> None:
    issue = CardQualityIssue(
        id=56,
        account_id=1,
        nm_id=245405620,
        issue_code="title_too_short",
        category="title",
        severity="high",
        title="Title too short",
        business_explanation="Short titles reduce discoverability.",
        recommended_fix="Add searchable product details.",
        field_name="title",
        current_value_json="Top",
        expected_value_json={"min_length": 20},
        confidence=0.9,
        status="in_progress",
        fingerprint="fp-title",
        first_seen_at=datetime.now(timezone.utc),
        last_seen_at=datetime.now(timezone.utc),
    )

    action = CardQualityAnalysisService()._action_from_issue(account_id=1, issue=issue)

    assert action.source_module == "checker"
    assert action.source_id == "56"
    assert action.action_type == "CARD_QUALITY_FIX"
    assert action.status == "in_progress"
    assert action.reason == "Short titles reduce discoverability."
    assert action.next_step == "Add searchable product details."
    assert action.expected_impact_amount is None
    assert action.priority_score == 15.0
    assert action.payload["issue_code"] == "title_too_short"
    assert action.payload["field_name"] == "title"
    assert action.payload["current_value_json"] == "Top"
    assert action.payload["expected_value_json"] == {"min_length": 20}
    assert action.payload["impact_kind"] == "estimated_opportunity"
    assert action.payload["checker_problem_bridge"] is True
    assert action.payload["content_quality_signal"] is True
    assert action.payload["problem_ux_contract"] is True
    assert action.payload["trust_state"] == "opportunity"
    assert action.payload["impact_type"] == "opportunity"
    assert action.payload["financial_loss_confirmed"] is False
    assert action.trust_state == "opportunity"
    assert action.impact_type == "opportunity"
    assert action.evidence_ledger is not None
    assert action.evidence_ledger.impact_type == "opportunity"
    assert action.money_trust is not None
    assert action.money_trust.show_as_confirmed_money is False
    assert action.solve_map is not None
    assert action.solve_map.primary_action_code == "run_checker"
    assert action.solve_map.steps[1].target_href == "/checker/245405620"


def test_card_quality_content_issue_does_not_become_confirmed_financial_loss() -> None:
    issue = CardQualityIssue(
        id=57,
        account_id=1,
        nm_id=245405620,
        issue_code="no_photos",
        category="media",
        severity="critical",
        title="Нет фотографий",
        business_explanation="Без фотографий покупателю сложно оценить товар.",
        recommended_fix="Добавьте фотографии товара.",
        status="new",
        fingerprint="fp-photo",
        first_seen_at=datetime.now(timezone.utc),
        last_seen_at=datetime.now(timezone.utc),
    )

    action = CardQualityAnalysisService()._action_from_issue(account_id=1, issue=issue)

    assert action.source_module == "checker"
    assert action.payload["bridge_kind"] == "content_quality"
    assert action.trust_state == "opportunity"
    assert action.impact_type == "opportunity"
    assert action.evidence_ledger is not None
    assert action.evidence_ledger.impact_type == "opportunity"
    assert action.money_trust is not None
    assert action.money_trust.show_as_confirmed_money is False
    assert action.payload["financial_loss_confirmed"] is False


@pytest.mark.asyncio
async def test_checker_product_recheck_improves_score_and_records_result_event() -> None:
    sync_session, session = _card_quality_recheck_session()
    now = datetime(2026, 7, 12, tzinfo=timezone.utc)
    snapshot = CardQualitySnapshot(
        id=1,
        account_id=1,
        nm_id=1001,
        source_card_id=10,
        source_revision="rev-before",
        title="Short",
        description="Old description",
        brand="Avemod",
        subject_name="Костюмы",
        vendor_code="VC-1",
        characteristics_json=[],
        media_json={},
        photos_count=0,
        video_count=0,
        analyzed_at=now,
        score=60,
        status="warning",
        summary_json={"score": 60, "status": "warning"},
    )
    issue = CardQualityIssue(
        id=10,
        account_id=1,
        nm_id=1001,
        snapshot_id=1,
        issue_code="no_title",
        category="title",
        severity="critical",
        title="No title",
        business_explanation="Title is missing.",
        recommended_fix="Add title.",
        status="new",
        fingerprint="fp-before",
        first_seen_at=now,
        last_seen_at=now,
    )
    sync_session.add_all([snapshot, issue])
    sync_session.flush()

    service = CardQualityAnalysisService()

    async def _fake_analyze_product(session_arg, *, account_id, nm_id, force, requested_by_user_id=None, run=None):
        issue.status = "resolved"
        issue.resolved_at = now
        sync_session.add(
            CardQualitySnapshot(
                id=2,
                account_id=account_id,
                nm_id=nm_id,
                source_card_id=10,
                source_revision="rev-after",
                title="Clean title",
                description="Clean description",
                brand="Avemod",
                subject_name="Костюмы",
                vendor_code="VC-1",
                characteristics_json=[],
                media_json={},
                photos_count=6,
                video_count=1,
                analyzed_at=now.replace(hour=1),
                score=95,
                status="clean",
                summary_json={"score": 95, "status": "clean"},
            )
        )
        sync_session.add(
            CardQualityAnalysisRun(
                id=99,
                account_id=account_id,
                run_type="single_product",
                status="completed",
                requested_by_user_id=requested_by_user_id,
                started_at=now.replace(hour=1),
                finished_at=now.replace(hour=1),
                cards_total=1,
                cards_processed=1,
                cards_analyzed=1,
                issues_resolved=1,
                last_processed_key=str(nm_id),
            )
        )
        await session_arg.flush()
        return SimpleNamespace(message="card quality analyzed locally")

    service.analyze_product = _fake_analyze_product  # type: ignore[method-assign]

    response = await service.recheck_product(session, account_id=1, nm_id=1001, requested_by_user_id=42)

    assert response.status == "completed"
    assert response.run_id == 99
    assert response.previous_score == 60
    assert response.new_score == 95
    assert response.previous_open_issue_count == 1
    assert response.new_open_issue_count == 0
    assert response.resolved_issue_ids == [10]
    assert response.result_status == "improved"
    assert response.action_center_updates[0]["result_badge"] == "resolved_after_recheck"
    event = sync_session.get(ResultEvent, response.result_event_id)
    assert event is not None
    assert event.event_type == "recheck_result"
    assert event.source_module == "checker"
    assert event.payload_json["saved_money_claimed"] is False
    assert event.payload_json["comparison"]["outcome"] == "improved"


@pytest.mark.asyncio
async def test_checker_product_recheck_pending_when_no_after_data() -> None:
    sync_session, session = _card_quality_recheck_session()
    service = CardQualityAnalysisService()

    async def _fake_analyze_product(session_arg, *, account_id, nm_id, force, requested_by_user_id=None, run=None):
        return SimpleNamespace(message="product card is not present in finance product cards")

    service.analyze_product = _fake_analyze_product  # type: ignore[method-assign]

    response = await service.recheck_product(session, account_id=1, nm_id=100404, requested_by_user_id=42)

    assert response.status == "completed"
    assert response.result_status == "pending_data"
    assert response.new_score is None
    assert response.new_open_issue_count is None
    event = sync_session.get(ResultEvent, response.result_event_id)
    assert event is not None
    assert event.source_module == "checker"
    assert event.payload_json["after_snapshot"] == {}


def test_card_quality_issue_contract_content_issue_is_opportunity_not_confirmed_loss() -> None:
    issue = CardQualityIssueRead.model_validate(
        CardQualityIssue(
            id=157,
            account_id=1,
            nm_id=245405620,
            issue_code="title_too_short",
            category="title",
            severity="high",
            title="Title too short",
            business_explanation="Short titles reduce discoverability.",
            recommended_fix="Add searchable product details.",
            field_name="title",
            current_value_json="Top",
            suggested_value="Костюм женский жакет и брюки повседневный",
            status="new",
            fingerprint="fp-title-contract",
            first_seen_at=datetime.now(timezone.utc),
            last_seen_at=datetime.now(timezone.utc),
        ),
        from_attributes=True,
    )

    assert issue.impact_type == "opportunity"
    assert issue.impact_type != "confirmed_loss"
    assert issue.trust_state == "opportunity"
    assert issue.issue_group == "title"
    assert issue.score_band == "critical"
    assert issue.opportunity_score is not None
    assert issue.expected_opportunity_count == 1
    assert issue.can_fix_locally is True
    assert issue.issue_id == 157
    assert issue.action_id == "card_quality:157"
    assert issue.problem_instance_id is None
    assert issue.current_value == "Top"
    assert issue.can_accept_local is True
    assert issue.accept_local_disabled_reason is None
    assert issue.can_mark_fixed is True
    assert issue.mark_fixed_disabled_reason is None
    assert issue.can_save_draft is True
    assert issue.save_draft_disabled_reason is None
    assert issue.can_preview_wb is True
    assert issue.preview_wb_disabled_reason is None
    assert issue.can_apply_to_wb is True
    assert issue.result_status == "pending_data"
    assert issue.apply_wb_disabled_reason is None
    assert issue.can_recheck is True
    assert issue.recheck_disabled_reason is None


def test_card_quality_missing_source_data_becomes_data_blocker() -> None:
    issue = CardQualityIssue(
        id=58,
        account_id=1,
        nm_id=245405620,
        issue_code="source_data_missing",
        category="data",
        severity="critical",
        title="Не хватает данных карточки",
        business_explanation="Checker не может проверить карточку без исходных данных.",
        recommended_fix="Запустите синхронизацию карточек.",
        status="blocked",
        fingerprint="fp-data",
        first_seen_at=datetime.now(timezone.utc),
        last_seen_at=datetime.now(timezone.utc),
    )

    action = CardQualityAnalysisService()._action_from_issue(account_id=1, issue=issue)

    assert action.payload["bridge_kind"] == "content_data_blocker"
    assert action.trust_state == "blocked"
    assert action.impact_type == "data_blocker"
    assert action.evidence_ledger is not None
    assert action.evidence_ledger.missing_data
    assert action.money_trust is not None
    assert action.money_trust.show_as_confirmed_money is False


def test_card_quality_issue_contract_data_blocker_has_blocked_state_and_missing_data() -> None:
    issue = CardQualityIssueRead.model_validate(
        CardQualityIssue(
            id=158,
            account_id=1,
            nm_id=245405620,
            issue_code="source_data_missing",
            category="data",
            severity="critical",
            title="Не хватает данных карточки",
            business_explanation="Checker не может проверить карточку без исходных данных.",
            recommended_fix="Запустите синхронизацию карточек.",
            status="blocked",
            fingerprint="fp-data-contract",
            first_seen_at=datetime.now(timezone.utc),
            last_seen_at=datetime.now(timezone.utc),
        ),
        from_attributes=True,
    )

    assert issue.impact_type == "data_blocker"
    assert issue.trust_state == "blocked"
    assert issue.issue_group == "data_blocker"
    assert issue.can_fix_locally is False
    assert issue.can_apply_to_wb is False
    assert issue.apply_disabled_reason == "data_blocker_requires_source_fix"
    assert issue.can_accept_local is False
    assert issue.accept_local_disabled_reason == "data_blocker_requires_source_fix"
    assert issue.can_mark_fixed is False
    assert issue.mark_fixed_disabled_reason == "data_blocker_requires_source_fix"
    assert issue.can_save_draft is False
    assert issue.save_draft_disabled_reason == "data_blocker_requires_source_fix"
    assert issue.can_preview_wb is False
    assert issue.preview_wb_disabled_reason == "data_blocker_requires_source_fix"
    assert issue.apply_wb_disabled_reason == "data_blocker_requires_source_fix"
    assert issue.missing_data == ["source_data_missing"]
    assert issue.evidence_ledger is not None
    assert issue.evidence_ledger.missing_data == ["source_data_missing"]


def test_card_quality_issue_contract_human_check_cannot_auto_apply() -> None:
    issue = CardQualityIssueRead.model_validate(
        CardQualityIssue(
            id=159,
            account_id=1,
            nm_id=245405620,
            issue_code="ai_visual_mismatch",
            category="characteristics",
            severity="medium",
            title="Material needs review",
            field_name="characteristics.Фактура материала",
            suggested_value="габардин",
            requires_human_check=True,
            status="new",
            fingerprint="fp-human-contract",
            first_seen_at=datetime.now(timezone.utc),
            last_seen_at=datetime.now(timezone.utc),
        ),
        from_attributes=True,
    )

    assert issue.issue_group == "characteristics"
    assert issue.impact_type == "opportunity"
    assert issue.can_accept_local is False
    assert issue.accept_local_disabled_reason == "human_check_requires_manual_review"
    assert issue.can_preview_wb is True
    assert issue.can_apply_to_wb is False
    assert issue.apply_disabled_reason == "human_check_required"
    assert issue.apply_wb_disabled_reason == "human_check_required"


def test_card_quality_issue_contract_unsupported_wb_apply_reason_is_explicit() -> None:
    issue = CardQualityIssueRead.model_validate(
        CardQualityIssue(
            id=160,
            account_id=1,
            nm_id=245405620,
            issue_code="media_no_images",
            category="media",
            severity="critical",
            title="No images",
            field_name="photos",
            suggested_value="https://example.test/photo.webp",
            status="new",
            fingerprint="fp-media-contract",
            first_seen_at=datetime.now(timezone.utc),
            last_seen_at=datetime.now(timezone.utc),
        ),
        from_attributes=True,
    )

    assert issue.issue_group == "media"
    assert issue.can_accept_local is False
    assert issue.accept_local_disabled_reason == "media_requires_dedicated_media_flow"
    assert issue.can_mark_fixed is True
    assert issue.can_save_draft is False
    assert issue.save_draft_disabled_reason == "media_requires_dedicated_media_flow"
    assert issue.can_preview_wb is False
    assert issue.preview_wb_disabled_reason == "unsupported_wb_apply_field:photos"
    assert issue.can_apply_to_wb is False
    assert issue.apply_disabled_reason == "unsupported_wb_apply_field:photos"
    assert issue.apply_wb_disabled_reason == "unsupported_wb_apply_field:photos"


@pytest.mark.parametrize(
    ("code", "category", "field_name", "expected_group"),
    [
        ("title_too_short", "title", "title", "title"),
        ("description_too_short", "description", "description", "description"),
        ("wb_allowed_values", "characteristics", "characteristics.Состав", "characteristics"),
        ("media_no_video_info", "media", "videos", "media"),
        ("subject_mismatch", "identity", "subject_name", "category"),
        ("no_description", "description", "description", "description"),
    ],
)
def test_card_quality_issue_group_maps_from_code_category_and_field(code: str, category: str, field_name: str, expected_group: str) -> None:
    issue = CardQualityIssueRead.model_validate(
        CardQualityIssue(
            id=161,
            account_id=1,
            nm_id=245405620,
            issue_code=code,
            category=category,
            severity="medium",
            title=code,
            field_name=field_name,
            status="new",
            fingerprint=f"fp-{code}",
            first_seen_at=datetime.now(timezone.utc),
            last_seen_at=datetime.now(timezone.utc),
        ),
        from_attributes=True,
    )

    assert issue.issue_group == expected_group


def test_card_quality_issue_with_business_metrics_is_operational_opportunity_until_financial_confirmation() -> None:
    issue = CardQualityIssue(
        id=59,
        account_id=1,
        nm_id=245405620,
        issue_code="description_weak_conversion",
        category="description",
        severity="high",
        title="Описание можно усилить",
        business_explanation="Описание связано с низкой конверсией по карточке.",
        recommended_fix="Уточните преимущества товара в описании.",
        ai_evidence_json={"business_metrics": {"conversion_rate": 0.018, "orders": 12}},
        status="new",
        fingerprint="fp-business",
        first_seen_at=datetime.now(timezone.utc),
        last_seen_at=datetime.now(timezone.utc),
    )

    action = CardQualityAnalysisService()._action_from_issue(account_id=1, issue=issue)

    assert action.source_module == "checker"
    assert action.payload["checker_problem_bridge"] is True
    assert action.payload["bridge_kind"] == "business_metrics"
    assert action.payload["business_metric_evidence"] is True
    assert action.trust_state == "opportunity"
    assert action.impact_type == "opportunity"
    assert action.payload["financial_loss_confirmed"] is False
    assert "recheck" in action.allowed_actions


def test_card_quality_confirmed_financial_evidence_can_be_confirmed_loss() -> None:
    issue = CardQualityIssue(
        id=60,
        account_id=1,
        nm_id=245405620,
        issue_code="description_confirmed_loss",
        category="description",
        severity="high",
        title="Описание связано с измеренной потерей",
        business_explanation="После проверки есть подтверждённая финансовая метрика.",
        recommended_fix="Уточните описание и перепроверьте карточку.",
        ai_evidence_json={
            "confirmed_financial_evidence": True,
            "financial_final": True,
            "confirmed_loss_amount": 1200,
        },
        status="new",
        fingerprint="fp-confirmed-finance",
        first_seen_at=datetime.now(timezone.utc),
        last_seen_at=datetime.now(timezone.utc),
    )

    action = CardQualityAnalysisService()._action_from_issue(account_id=1, issue=issue)

    assert action.payload["bridge_kind"] == "business_metrics"
    assert action.payload["business_metric_evidence"] is True
    assert action.payload["financial_loss_confirmed"] is True
    assert action.trust_state == "confirmed"
    assert action.impact_type == "confirmed_loss"
    assert action.money_trust is not None
    assert action.money_trust.show_as_confirmed_money is True
    assert action.evidence_ledger is not None
    assert action.evidence_ledger.impact_type == "confirmed_loss"


def test_card_quality_queue_bucket_matches_source_semantics() -> None:
    service = CardQualityAnalysisService()
    media = CardQualityIssue(
        id=70,
        account_id=1,
        nm_id=100,
        issue_code="media_no_images",
        category="media",
        severity="critical",
        title="No images",
        field_name="photos",
        status="new",
        fingerprint="fp-media",
        first_seen_at=datetime.now(timezone.utc),
        last_seen_at=datetime.now(timezone.utc),
    )
    human = CardQualityIssue(
        id=71,
        account_id=1,
        nm_id=100,
        issue_code="visual_check",
        category="characteristics",
        severity="medium",
        title="Check material",
        field_name="characteristics.Фактура материала",
        status="new",
        requires_human_check=True,
        fingerprint="fp-human",
        first_seen_at=datetime.now(timezone.utc),
        last_seen_at=datetime.now(timezone.utc),
    )

    assert service.issue_belongs_to_bucket(media, "media") is True
    assert service.issue_belongs_to_bucket(media, "actionable") is False
    assert service.issue_belongs_to_bucket(human, "human_check") is True
    assert service.issue_belongs_to_bucket(human, "actionable") is True


def test_card_quality_fixed_file_parser_accepts_source_template_headers() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Артикул WB", "Бренд", "Предмет", "Характеристика", "Эталонное значение"])
    sheet.append([245405620, "Avemod", "Костюмы", "Состав", "хлопок 80%, полиэстер 20%"])
    payload = BytesIO()
    workbook.save(payload)

    entries = CardQualityAnalysisService()._parse_fixed_file_excel(payload.getvalue())

    assert entries == [
        {
            "nm_id": 245405620,
            "brand": "Avemod",
            "subject_name": "Костюмы",
            "char_name": "Состав",
            "fixed_value": "хлопок 80%, полиэстер 20%",
        }
    ]


def test_card_quality_fixed_file_parser_accepts_wide_source_workbook() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист1"
    sheet.append(["nmID", "brand", "subjectName", "Пол", "Состав", "Ставка НДС"])
    sheet.append([245405620, "Avemod", "Костюмы", "Женский", "хлопок 80%;полиэстер 20%", 5])
    sheet.append([245405621, "Avemod", "Костюмы", "", "полиэстер", None])
    payload = BytesIO()
    workbook.save(payload)

    entries = CardQualityAnalysisService()._parse_fixed_file_excel(payload.getvalue())

    assert entries == [
        {
            "nm_id": 245405620,
            "brand": "Avemod",
            "subject_name": "Костюмы",
            "char_name": "Пол",
            "fixed_value": "Женский",
        },
        {
            "nm_id": 245405620,
            "brand": "Avemod",
            "subject_name": "Костюмы",
            "char_name": "Состав",
            "fixed_value": "хлопок 80%;полиэстер 20%",
        },
        {
            "nm_id": 245405620,
            "brand": "Avemod",
            "subject_name": "Костюмы",
            "char_name": "Ставка НДС",
            "fixed_value": "5",
        },
        {
            "nm_id": 245405621,
            "brand": "Avemod",
            "subject_name": "Костюмы",
            "char_name": "Состав",
            "fixed_value": "полиэстер",
        },
    ]


@pytest.mark.asyncio
async def test_card_quality_fixed_file_table_flow_filters_updates_and_exports() -> None:
    sync_session, session = _card_quality_recheck_session()
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["nmID", "brand", "subjectName", "Пол", "Состав", "Модель брюк"])
        sheet.append([245405620, "Avemod", "Костюмы", "Женский", "хлопок", "классические"])
        sheet.append([245405621, "Avemod", "Брюки", "Женский", "полиэстер", "палаццо"])
        payload = BytesIO()
        workbook.save(payload)

        service = CardQualityAnalysisService()
        upload = await service.upload_fixed_file(
            session,
            account_id=1,
            content=payload.getvalue(),
            filename="fixed.xlsx",
            replace_all=True,
        )

        assert upload.upserted == 6
        assert upload.total == 6
        status = await service.fixed_file_status(session, account_id=1)
        assert status.total_cards == 2
        assert status.total_characteristics == 3

        page = await service.list_fixed_file_entries(
            session,
            account_id=1,
            search="Состав",
            sort_by="fixed_value",
            sort_dir="desc",
            limit=10,
            offset=0,
        )
        assert page.total == 2
        assert [item.fixed_value for item in page.items] == ["хлопок", "полиэстер"]

        updated = await service.update_fixed_file_entry(
            session,
            account_id=1,
            entry_id=page.items[0].id,
            payload=CardQualityFixedFileEntryMutation(fixed_value="хлопок 80%; полиэстер 20%"),
        )
        assert updated.fixed_value == "хлопок 80%; полиэстер 20%"

        exported = await service.export_fixed_file_entries(
            session,
            account_id=1,
            char_name="Состав",
            sort_by="nm_id",
            sort_dir="asc",
        )
        loaded = load_workbook(BytesIO(exported))
        rows = list(loaded.active.iter_rows(values_only=True))
        assert rows[0] == ("nmID", "brand", "subjectName", "Характеристика", "Эталонное значение")
        assert rows[1][3:] == ("Состав", "хлопок 80%; полиэстер 20%")
        assert rows[2][3:] == ("Состав", "полиэстер")
    finally:
        sync_session.close()


def test_card_quality_ai_audit_dedupe_key_accepts_list_values() -> None:
    service = CardQualityAnalysisService()

    assert service._hashable_issue_value(["палаццо"]) == '["палаццо"]'


def test_card_quality_ai_category_mapping_prefers_source_field_path() -> None:
    service = CardQualityAnalysisService()

    assert service._map_ai_issue_category("text", "characteristics.Тип посадки") == "characteristics"
    assert service._map_ai_issue_category("qualification", "characteristics.Состав") == "characteristics"
    assert service._map_ai_issue_category("text", "description") == "description"


def test_card_quality_no_safe_ai_fix_requires_human_check() -> None:
    service = CardQualityAnalysisService()
    issue = service.rules._issue(
        "ai_photo",
        "characteristics",
        "medium",
        "Material cannot be verified",
        "AI cannot verify this from safe evidence.",
        "Review manually.",
        "characteristics.Фактура материала",
        "кожаный",
    )

    merged = service._merge_ai_fix(
        normalized=_card(),
        issue=issue,
        fix={"recommended_value": None, "suggestion_kind": "no_safe_fix", "reason": "Недостаточно visual evidence"},
        recommended=None,
        invalid_reason=None,
    )

    assert merged.requires_human_check is True
    assert merged.suggested_value is None


def test_card_quality_product_quality_snapshot_payload_contains_card_fields() -> None:
    service = CardQualityAnalysisService()
    snapshot = SimpleNamespace(
        nm_id=245405620,
        source_card_id=10,
        title="Брюки палаццо кожаные",
        description="Описание товара",
        brand="Avemod",
        subject_name="Брюки",
        vendor_code="AV-1",
        characteristics_json=[{"name": "Цвет", "value": ["кэмел"]}],
        media_json={
            "photos": [{"canonical_url": "https://cdn.example.test/1.webp"}],
            "videos": [],
            "sizes": [{"techSize": "42"}],
        },
        photos_count=1,
        video_count=0,
        source_revision="rev",
        source_updated_at=None,
    )

    payload = service._snapshot_card_payload(snapshot)

    assert payload["title"] == "Брюки палаццо кожаные"
    assert payload["description"] == "Описание товара"
    assert payload["characteristics"] == [{"name": "Цвет", "value": ["кэмел"]}]
    assert payload["photos"][0]["canonical_url"] == "https://cdn.example.test/1.webp"
    assert payload["primary_photo"] == "https://cdn.example.test/1.webp"


class _FakeIssueSession:
    def __init__(self, issue: CardQualityIssue) -> None:
        self.issue = issue
        self.added: list[object] = []
        self.committed = False

    async def get(self, model, issue_id: int):
        assert model is CardQualityIssue
        return self.issue if issue_id == self.issue.id else None

    def add(self, item: object) -> None:
        self.added.append(item)

    async def commit(self) -> None:
        self.committed = True


class _FakeScalarResult:
    def __init__(self, *, scalar_value=None, rows=None):
        self.scalar_value = scalar_value
        self.rows = rows or []

    def scalar(self):
        return self.scalar_value

    def scalars(self):
        return self

    def __iter__(self):
        return iter(self.rows)


class _FakeListIssuesSession:
    def __init__(self, issue: CardQualityIssue | list[CardQualityIssue]):
        self.issues = issue if isinstance(issue, list) else [issue]
        self.calls = 0

    async def execute(self, _stmt):
        self.calls += 1
        if self.calls == 1:
            return _FakeScalarResult(scalar_value=len(self.issues))
        if self.calls == 2:
            return _FakeScalarResult(rows=self.issues)
        return _FakeScalarResult(scalar_value=0)


class _FakeAnalyzeFailureSession:
    def __init__(self) -> None:
        self.added: list[object] = []
        self.rolled_back = False
        self.committed = False
        self.next_id = 100

    def add(self, item: object) -> None:
        self.added.append(item)

    async def flush(self) -> None:
        for item in self.added:
            if isinstance(item, CardQualityAnalysisRun) and getattr(item, "id", None) is None:
                item.id = self.next_id
                self.next_id += 1

    async def rollback(self) -> None:
        self.rolled_back = True

    async def get(self, _model, _key):
        return None

    async def commit(self) -> None:
        assert self.rolled_back is True
        self.committed = True


class _FakeBuildPayloadSession:
    async def execute(self, _stmt):
        return _FakeScalarResult(rows=[])


def _issue(status: str = "new") -> CardQualityIssue:
    now = datetime.now(timezone.utc)
    return CardQualityIssue(
        id=77,
        account_id=1,
        nm_id=245405620,
        issue_code="title_missing",
        category="title",
        severity="critical",
        title="Title missing",
        status=status,
        fingerprint="fp",
        first_seen_at=now,
        last_seen_at=now,
    )


def _quality_issue(
    *,
    issue_id: int,
    code: str,
    category: str,
    severity: str,
    field_name: str | None,
    status: str = "new",
    source: str | None = "code",
) -> CardQualityIssue:
    now = datetime.now(timezone.utc)
    return CardQualityIssue(
        id=issue_id,
        account_id=1,
        nm_id=245405620,
        issue_code=code,
        category=category,
        severity=severity,
        title=code,
        field_name=field_name,
        status=status,
        source=source,
        fingerprint=f"fp-{issue_id}",
        first_seen_at=now,
        last_seen_at=now,
    )


@pytest.mark.asyncio
async def test_card_quality_analyze_records_failed_run_after_rollback() -> None:
    service = CardQualityAnalysisService()
    service.normalizer.normalize_product = AsyncMock(return_value=_card())
    service.latest_snapshot = AsyncMock(return_value=None)
    service.rules.analyze = Mock(side_effect=RuntimeError("classifier exploded"))
    session = _FakeAnalyzeFailureSession()

    with pytest.raises(RuntimeError, match="classifier exploded"):
        await service.analyze_product(session, account_id=1, nm_id=245405620, force=True, requested_by_user_id=7)

    failed_runs = [item for item in session.added if isinstance(item, CardQualityAnalysisRun) and item.status == "failed"]
    assert session.rolled_back is True
    assert session.committed is True
    assert failed_runs
    assert failed_runs[-1].cards_failed == 1
    assert failed_runs[-1].error_summary == "RuntimeError"


@pytest.mark.asyncio
async def test_card_quality_status_reopen_clears_resolved_at() -> None:
    resolved_at = datetime.now(timezone.utc)
    issue = CardQualityIssue(
        id=77,
        account_id=1,
        nm_id=245405620,
        issue_code="title_missing",
        category="title",
        severity="critical",
        title="Title missing",
        status="done",
        fingerprint="fp",
        first_seen_at=resolved_at,
        last_seen_at=resolved_at,
        resolved_at=resolved_at,
    )
    session = _FakeIssueSession(issue)

    updated = await CardQualityAnalysisService().update_issue_status(
        session, account_id=1, issue_id=77, status="new", changed_by_user_id=10, reason="recheck"
    )

    assert updated.status == "new"
    assert updated.resolved_at is None
    assert session.committed is True


@pytest.mark.asyncio
async def test_card_quality_status_rejects_illegal_transition() -> None:
    now = datetime.now(timezone.utc)
    issue = CardQualityIssue(
        id=78,
        account_id=1,
        nm_id=245405620,
        issue_code="title_missing",
        category="title",
        severity="critical",
        title="Title missing",
        status="new",
        fingerprint="fp",
        first_seen_at=now,
        last_seen_at=now,
    )

    with pytest.raises(ValueError, match="illegal_status_transition"):
        await CardQualityAnalysisService().update_issue_status(
            _FakeIssueSession(issue), account_id=1, issue_id=78, status="resolved", changed_by_user_id=10, reason=None
        )


@pytest.mark.asyncio
async def test_card_quality_action_center_in_progress_persists() -> None:
    issue = _issue()
    session = _FakeIssueSession(issue)

    updated = await CardQualityAnalysisService().update_issue_status(
        session,
        account_id=1,
        issue_id=77,
        status="in_progress",
        changed_by_user_id=10,
        reason="owner started",
    )

    assert updated.status == "in_progress"
    assert issue.status == "in_progress"
    assert issue.status_reason == "owner started"


@pytest.mark.asyncio
async def test_card_quality_done_persists_fixed_metadata_and_history() -> None:
    issue = _issue("in_progress")
    session = _FakeIssueSession(issue)

    updated = await CardQualityAnalysisService().update_issue_status(
        session,
        account_id=1,
        issue_id=77,
        status="done",
        changed_by_user_id=10,
        reason="fixed in WB",
        fixed_value="Новое название",
    )

    assert updated.status == "done"
    assert updated.fixed_value == "Новое название"
    assert updated.fixed_at is not None
    assert updated.fixed_by_user_id == 10
    assert session.added[-1].old_status == "in_progress"
    assert session.added[-1].new_status == "done"


@pytest.mark.asyncio
async def test_card_quality_human_check_issue_rejects_quick_accept_marker() -> None:
    issue = _issue()
    issue.requires_human_check = True

    with pytest.raises(ValueError, match="human_check_issue_requires_manual_review"):
        await CardQualityAnalysisService().update_issue_status(
            _FakeIssueSession(issue),
            account_id=1,
            issue_id=77,
            status="done",
            changed_by_user_id=10,
            reason="accepted_recommendation_status_only",
            fixed_value="Небезопасный кандидат",
        )

    assert issue.status == "new"
    assert issue.fixed_value is None


def test_card_quality_status_update_has_no_wb_apply_flag_by_default() -> None:
    payload = CardQualityIssueStatusUpdate(status="done", fixed_value="Новое значение")
    dumped = payload.model_dump()

    assert dumped["fixed_value"] == "Новое значение"
    assert "apply_to_wb" not in dumped


@pytest.mark.asyncio
async def test_card_quality_default_fix_does_not_write_to_wb(monkeypatch: pytest.MonkeyPatch) -> None:
    service = CardQualityAnalysisService()
    issue = _quality_issue(
        issue_id=77,
        code="wb_allowed_values",
        category="characteristics",
        severity="medium",
        field_name="characteristics.Состав",
    )
    issue.suggested_value = "хлопок"
    issue.ai_suggested_value = "хлопок"
    submit = AsyncMock()
    monkeypatch.setattr(service, "_apply_issue_fix_to_wb", submit)

    response = await service.fix_issue(
        _FakeIssueSession(issue),
        account_id=1,
        issue_id=77,
        fixed_value=None,
        changed_by_user_id=10,
        apply_to_wb=False,
        confirm=False,
    )

    assert response.status == "fixed_local"
    assert response.wb_write_status == "not_requested"
    assert response.message == "Issue marked fixed locally. WB was not changed."
    assert issue.status == "done"
    assert service._action_from_issue(account_id=1, issue=issue).status == "done"
    submit.assert_not_called()


@pytest.mark.asyncio
async def test_card_quality_accept_local_wrapper_does_not_write_to_wb(monkeypatch: pytest.MonkeyPatch) -> None:
    service = CardQualityAnalysisService()
    issue = _quality_issue(
        issue_id=77,
        code="title_too_short",
        category="title",
        severity="high",
        field_name="title",
    )
    issue.suggested_value = "Костюм женский жакет и брюки повседневный"
    submit = AsyncMock()
    monkeypatch.setattr(service, "_apply_issue_fix_to_wb", submit)

    response = await service.accept_issue_local(
        _FakeIssueSession(issue),
        account_id=1,
        issue_id=77,
        fixed_value=None,
        changed_by_user_id=10,
    )

    assert response.status == "fixed_local"
    assert response.wb_write_status == "not_requested"
    assert response.issue.status == "done"
    assert response.issue.status_reason == "accepted_locally_no_wb_write"
    assert response.issue.can_apply_to_wb is False
    assert response.issue.apply_wb_disabled_reason == "issue_status_done"
    submit.assert_not_called()


@pytest.mark.asyncio
async def test_card_quality_mark_fixed_does_not_write_to_wb(monkeypatch: pytest.MonkeyPatch) -> None:
    service = CardQualityAnalysisService()
    issue = _quality_issue(
        issue_id=77,
        code="no_photos",
        category="media",
        severity="critical",
        field_name="photos",
    )
    submit = AsyncMock()
    monkeypatch.setattr(service, "_apply_issue_fix_to_wb", submit)

    updated = await service.mark_issue_fixed(
        _FakeIssueSession(issue),
        account_id=1,
        issue_id=77,
        fixed_value=None,
        changed_by_user_id=10,
    )

    assert updated.status == "done"
    assert updated.status_reason == "marked_fixed_no_wb_write"
    assert service._action_from_issue(account_id=1, issue=issue).status == "done"
    submit.assert_not_called()


@pytest.mark.asyncio
async def test_card_quality_save_draft_keeps_issue_status_in_sync_and_no_wb_write(monkeypatch: pytest.MonkeyPatch) -> None:
    service = CardQualityAnalysisService()
    issue = _quality_issue(
        issue_id=77,
        code="description_too_short",
        category="description",
        severity="medium",
        field_name="description",
    )
    submit = AsyncMock()
    monkeypatch.setattr(service, "_apply_issue_fix_to_wb", submit)

    updated = await service.save_issue_draft(
        _FakeIssueSession(issue),
        account_id=1,
        issue_id=77,
        fixed_value="Черновик описания карточки",
        changed_by_user_id=10,
    )

    assert updated.status == "in_progress"
    assert updated.fixed_value == "Черновик описания карточки"
    assert updated.status_reason == "draft_saved_no_wb_write"
    assert service._action_from_issue(account_id=1, issue=issue).status == "in_progress"
    submit.assert_not_called()


@pytest.mark.asyncio
async def test_card_quality_submit_without_confirm_returns_confirmation_required(monkeypatch: pytest.MonkeyPatch) -> None:
    service = CardQualityAnalysisService()
    issue = _quality_issue(
        issue_id=77,
        code="wb_allowed_values",
        category="characteristics",
        severity="medium",
        field_name="characteristics.Состав",
    )
    submit = AsyncMock()
    monkeypatch.setattr(service, "_apply_issue_fix_to_wb", submit)

    response = await service.fix_issue(
        _FakeIssueSession(issue),
        account_id=1,
        issue_id=77,
        fixed_value="хлопок",
        changed_by_user_id=10,
        apply_to_wb=True,
        confirm=False,
    )

    assert response.status == "confirmation_required"
    assert response.wb_write_status == "confirmation_required"
    assert response.preview is not None
    assert response.preview.audit["preview_diff_required"] is True
    assert response.preview.audit["content_token_permission_required"] is True
    assert issue.status == "new"
    submit.assert_not_called()


@pytest.mark.asyncio
async def test_card_quality_human_check_fix_requires_manual_value() -> None:
    issue = _quality_issue(
        issue_id=77,
        code="ai_visual_mismatch",
        category="characteristics",
        severity="medium",
        field_name="characteristics.Фактура материала",
    )
    issue.requires_human_check = True
    issue.suggested_value = "габардин"

    with pytest.raises(ValueError, match="human_check_issue_requires_manual_review"):
        await CardQualityAnalysisService().fix_issue(
            _FakeIssueSession(issue),
            account_id=1,
            issue_id=77,
            fixed_value=None,
            changed_by_user_id=10,
            apply_to_wb=False,
            confirm=False,
        )

    assert issue.status == "new"


@pytest.mark.asyncio
async def test_card_quality_human_check_cannot_apply_to_wb_even_with_confirm(monkeypatch: pytest.MonkeyPatch) -> None:
    service = CardQualityAnalysisService()
    issue = _quality_issue(
        issue_id=77,
        code="ai_visual_mismatch",
        category="characteristics",
        severity="medium",
        field_name="characteristics.Фактура материала",
    )
    issue.requires_human_check = True
    submit = AsyncMock()
    monkeypatch.setattr(service, "_apply_issue_fix_to_wb", submit)

    response = await service.fix_issue(
        _FakeIssueSession(issue),
        account_id=1,
        issue_id=77,
        fixed_value="габардин",
        changed_by_user_id=10,
        apply_to_wb=True,
        confirm=True,
    )

    assert response.status == "blocked"
    assert response.wb_write_status == "blocked"
    assert response.preview is not None
    assert response.preview.can_apply_to_wb is False
    assert response.preview.blocked_reason == "human_check_required"
    assert issue.status == "new"
    submit.assert_not_called()


@pytest.mark.asyncio
async def test_card_quality_unsupported_field_cannot_apply_to_wb(monkeypatch: pytest.MonkeyPatch) -> None:
    service = CardQualityAnalysisService()
    issue = _quality_issue(
        issue_id=77,
        code="media_no_images",
        category="media",
        severity="critical",
        field_name="photos",
    )
    submit = AsyncMock()
    monkeypatch.setattr(service, "_apply_issue_fix_to_wb", submit)

    response = await service.fix_issue(
        _FakeIssueSession(issue),
        account_id=1,
        issue_id=77,
        fixed_value="https://example.test/photo.webp",
        changed_by_user_id=10,
        apply_to_wb=True,
        confirm=True,
    )

    assert response.status == "blocked"
    assert response.wb_write_status == "blocked"
    assert response.preview is not None
    assert response.preview.can_apply_to_wb is False
    assert response.preview.blocked_reason == "unsupported_wb_apply_field:photos"
    assert response.preview.apply_disabled_reason == "unsupported_wb_apply_field:photos"
    assert issue.status == "new"
    submit.assert_not_called()


@pytest.mark.asyncio
async def test_card_quality_wb_submit_waits_validation_and_keeps_status_in_sync(monkeypatch: pytest.MonkeyPatch) -> None:
    service = CardQualityAnalysisService()
    issue = _quality_issue(
        issue_id=77,
        code="wb_allowed_values",
        category="characteristics",
        severity="medium",
        field_name="characteristics.Состав",
    )
    apply_result = {
        "submitted": True,
        "external_status": "waiting_wb_validation",
        "audit": {
            "event_type": "card_quality_wb_submit_attempted",
            "confirm_required": True,
            "confirmed": True,
            "content_token_permission_checked": True,
            "follow_up_status": "waiting_wb_validation",
        },
    }
    submit = AsyncMock(return_value=apply_result)
    monkeypatch.setattr(service, "_apply_issue_fix_to_wb", submit)

    response = await service.fix_issue(
        _FakeIssueSession(issue),
        account_id=1,
        issue_id=77,
        fixed_value="хлопок",
        changed_by_user_id=10,
        apply_to_wb=True,
        confirm=True,
    )

    assert response.status == "submitted_to_wb"
    assert response.wb_write_status == "submitted_waiting_validation"
    assert response.apply_result == apply_result
    assert issue.status == "in_progress"
    assert issue.status_reason == "wb_submit_attempted_waiting_validation"
    assert service._action_from_issue(account_id=1, issue=issue).status == "in_progress"
    submit.assert_awaited_once()


@pytest.mark.asyncio
async def test_card_quality_recheck_refreshes_issue_status_and_result(monkeypatch: pytest.MonkeyPatch) -> None:
    service = CardQualityAnalysisService()
    issue = _quality_issue(
        issue_id=77,
        code="title_too_short",
        category="title",
        severity="high",
        field_name="title",
        status="done",
    )
    issue.resolved_at = datetime.now(timezone.utc)
    calls: list[dict[str, object]] = []

    async def _fake_analyze_product(session, **kwargs):
        calls.append(kwargs)
        issue.status = "resolved"
        issue.status_reason = "auto_resolve_absent_from_latest_analysis"
        return SimpleNamespace(status="ok", raw={"issues_count": 0})

    monkeypatch.setattr(service, "analyze_product", _fake_analyze_product)

    updated = await service.recheck_issue(
        _FakeIssueSession(issue),
        account_id=1,
        issue_id=77,
        requested_by_user_id=10,
    )

    assert calls == [{"account_id": 1, "nm_id": 245405620, "force": True, "requested_by_user_id": 10}]
    assert updated.status == "resolved"
    assert updated.status_reason == "auto_resolve_absent_from_latest_analysis"
    assert updated.can_recheck is True
    assert service._action_from_issue(account_id=1, issue=issue).status == "done"


@pytest.mark.asyncio
async def test_card_quality_missing_content_token_blocks_wb_submit(monkeypatch: pytest.MonkeyPatch) -> None:
    service = CardQualityAnalysisService()
    issue = _quality_issue(
        issue_id=77,
        code="wb_allowed_values",
        category="characteristics",
        severity="medium",
        field_name="characteristics.Состав",
    )
    monkeypatch.setattr(
        service,
        "_apply_issue_fix_to_wb",
        AsyncMock(side_effect=ValueError("content_token_permission_required:WB token for category 'content' is not configured")),
    )

    response = await service.fix_issue(
        _FakeIssueSession(issue),
        account_id=1,
        issue_id=77,
        fixed_value="хлопок",
        changed_by_user_id=10,
        apply_to_wb=True,
        confirm=True,
    )

    assert response.status == "blocked"
    assert response.wb_write_status == "blocked"
    assert "content_token_permission_required" in str(response.message)
    assert issue.status == "new"


def test_wb_update_snapshot_uses_official_update_fields_only() -> None:
    payload = _sanitize_wb_update_snapshot(
        {
            "nmID": 245405620,
            "vendorCode": "AV-1",
            "kizMarked": True,
            "brand": "Avemod",
            "title": "Новое название",
            "description": "Описание",
            "dimensions": {"length": 35, "width": 40, "height": 15, "weightBrutto": 3},
            "characteristics": [{"id": 14177450, "name": "Состав", "value": ["хлопок"]}],
            "sizes": [{"chrtID": 123, "techSize": "ONE SIZE", "skus": ["1234567890123"]}],
            "subjectID": 777,
            "subjectName": "Костюмы",
            "imtID": 987,
            "photos": ["https://example.test/photo.webp"],
            "video": "https://example.test/video.mp4",
            "tags": [{"id": 1}],
            "updatedAt": "2026-07-09T00:00:00Z",
        }
    )

    assert set(payload) == {
        "nmID",
        "vendorCode",
        "kizMarked",
        "brand",
        "title",
        "description",
        "dimensions",
        "characteristics",
        "sizes",
    }
    assert payload["nmID"] == 245405620
    assert payload["kizMarked"] is True
    assert "photos" not in payload
    assert "video" not in payload
    assert "tags" not in payload


@pytest.mark.asyncio
async def test_wb_update_payload_preserves_existing_kiz_marked_flag() -> None:
    service = CardQualityAnalysisService()
    card = WBProductCard(
        id=10,
        account_id=1,
        nm_id=245405620,
        vendor_code="AV-1",
        title="Старое название",
        description="Описание",
        brand="Avemod",
        kiz_marked=True,
        dimensions={"length": 35, "width": 40, "height": 15, "weightBrutto": 3},
        payload={
            "nmID": 245405620,
            "vendorCode": "AV-1",
            "title": "Старое название",
            "description": "Описание",
            "sizes": [{"chrtID": 123, "techSize": "ONE SIZE", "skus": ["1234567890123"]}],
        },
    )
    issue = _issue()
    issue.field_name = "title"

    payload, target = await service._build_wb_update_payload(
        _FakeBuildPayloadSession(),
        account_id=1,
        card=card,
        issue=issue,
        fixed_value="Новое название",
    )

    assert payload["title"] == "Новое название"
    assert payload["kizMarked"] is True
    assert target["kizMarked"] is True


@pytest.mark.asyncio
async def test_card_quality_issue_pages_use_source_pipeline_order() -> None:
    service = CardQualityAnalysisService()
    wb_issue = _quality_issue(
        issue_id=90,
        code="wb_allowed_values",
        category="characteristics",
        severity="critical",
        field_name="characteristics.Тип верха",
    )
    title_issue = _quality_issue(
        issue_id=91,
        code="title_too_short",
        category="title",
        severity="medium",
        field_name="title",
    )

    page = await service.list_issues(
        _FakeListIssuesSession([wb_issue, title_issue]),
        account_id=1,
        category=None,
        status=None,
        include_info=True,
        limit=20,
        offset=0,
    )

    assert [item.issue_code for item in page.items] == ["title_too_short", "wb_allowed_values"]
    assert service._issue_payload(title_issue)["source_order"] < service._issue_payload(wb_issue)["source_order"]


def test_card_quality_queue_keeps_status_then_source_pipeline_order() -> None:
    service = CardQualityAnalysisService()
    wb_issue = _quality_issue(
        issue_id=92,
        code="wb_allowed_values",
        category="characteristics",
        severity="critical",
        field_name="characteristics.Тип верха",
    )
    title_issue = _quality_issue(
        issue_id=93,
        code="title_too_short",
        category="title",
        severity="medium",
        field_name="title",
    )
    postponed_title = _quality_issue(
        issue_id=94,
        code="title_too_long",
        category="title",
        severity="medium",
        field_name="title",
        status="postponed",
    )

    ordered = sorted([wb_issue, postponed_title, title_issue], key=service._queue_sort_key)

    assert [issue.issue_code for issue in ordered] == ["title_too_short", "wb_allowed_values", "title_too_long"]


def test_card_quality_final_filters_collapse_compound_overlaps() -> None:
    service = CardQualityAnalysisService()
    compound = replace(
        service.rules._issue(
            "ai_mixed",
            "characteristics",
            "medium",
            "Compound fix",
            "Move conflicting values.",
            "Review compound fix.",
            "characteristics.Тип низа",
            "брюки",
        ),
        source="ai",
        requires_human_check=True,
        error_details=[
            {
                "type": "compound",
                "fixes": [
                    {"name": "Модель брюк", "value": "палаццо"},
                    {"field_path": "characteristics.Модель юбки", "value": "__CLEAR__"},
                ],
            }
        ],
    )
    covered = service.rules._issue(
        "wb_allowed_values",
        "characteristics",
        "medium",
        "Covered field",
        "Same field is covered by compound fix.",
        "Set value.",
        "characteristics.Модель брюк",
        "классические",
        "палаццо",
    )

    collapsed = service._collapse_compound_overlaps([compound, covered])

    assert collapsed == [compound]


def test_card_quality_final_filters_merge_same_field_competitors_into_human_check() -> None:
    service = CardQualityAnalysisService()
    wb_issue = service.rules._issue(
        "wb_allowed_values",
        "characteristics",
        "critical",
        "WB value",
        "WB suggests one value.",
        "Set WB value.",
        "characteristics.Тип верха",
        "жакет",
        "жакет",
    )
    ai_issue = replace(
        service.rules._issue(
            "ai_photo",
            "characteristics",
            "medium",
            "AI value",
            "AI suggests another value.",
            "Set AI value.",
            "characteristics.Тип верха",
            "жакет",
            "жилет",
        ),
        source="ai",
        ai_suggested_value="жилет",
    )

    collapsed = service._collapse_same_field_competitors([ai_issue, wb_issue])

    assert len(collapsed) == 1
    assert collapsed[0].requires_human_check is True
    assert collapsed[0].suggested_value is None
    assert collapsed[0].ai_alternatives == ["жилет", "жакет"]
    assert collapsed[0].ai_evidence["merged_issue_codes"] == ["ai_photo", "wb_allowed_values"]


def test_card_quality_final_filters_drop_description_refresh_when_stronger_description_issue_exists() -> None:
    service = CardQualityAnalysisService()
    refresh = service.rules._issue(
        "description_refresh_needed",
        "description",
        "medium",
        "Refresh description",
        "Characteristics changed.",
        "Generate description.",
        "description",
        "old",
    )
    too_short = service.rules._issue(
        "description_too_short",
        "description",
        "medium",
        "Too short",
        "Description is short.",
        "Generate full description.",
        "description",
        "short",
    )

    collapsed = service._collapse_description_refresh_overlaps([refresh, too_short])

    assert [issue.issue_code for issue in collapsed] == ["description_too_short"]


def test_card_quality_final_filters_drop_destructive_ai_issue_without_safe_value() -> None:
    service = CardQualityAnalysisService()
    issue = replace(
        service.rules._issue(
            "ai_qualification",
            "characteristics",
            "medium",
            "AI proposed unsafe clear",
            "AI proposed a destructive action.",
            "Review manually.",
            "characteristics.Коллекция",
            "лето",
        ),
        source="ai",
        suggested_value=None,
        ai_suggested_value=None,
        requires_human_check=True,
        error_details=[{"type": "clear", "fix_action": "clear"}],
    )

    finalized = service._finalize_rule_issues(_card(), [issue], {})

    assert finalized == []


def test_card_quality_final_filters_drop_multivalue_reorder_only_issue() -> None:
    service = CardQualityAnalysisService()
    issue = replace(
        service.rules._issue(
            "ai_photo",
            "characteristics",
            "medium",
            "AI reordered values",
            "Only order changed.",
            "No-op.",
            "characteristics.Назначение",
            "офис, повседневная, классический",
            "офис, классический, повседневная",
        ),
        source="ai",
        suggested_value="офис, классический, повседневная",
        ai_suggested_value="офис, классический, повседневная",
    )

    finalized = service._finalize_rule_issues(_card(), [issue], {})

    assert finalized == []


def test_card_quality_visual_risky_allowed_value_is_not_fuzzy_corrected() -> None:
    service = CardQualityAnalysisService()

    valid, corrected, reason = service._validate_allowed_value(
        "габардиин",
        ["мех", "твид", "габардин"],
        [{"type": "allowed_values", "invalidValues": ["костюмная"]}],
        field_name="characteristics.Фактура материала",
    )

    assert valid is False
    assert corrected is None
    assert reason == "value_not_allowed:габардиин"


@pytest.mark.asyncio
async def test_card_quality_ignored_persists_reason() -> None:
    issue = _issue()

    updated = await CardQualityAnalysisService().update_issue_status(
        _FakeIssueSession(issue),
        account_id=1,
        issue_id=77,
        status="ignored",
        changed_by_user_id=10,
        reason="not relevant",
    )

    assert updated.status == "ignored"
    assert updated.status_reason == "not relevant"
    assert updated.resolved_at is not None


@pytest.mark.asyncio
async def test_card_quality_postponed_persists_reason_and_until() -> None:
    issue = _issue()
    postponed_until = datetime(2026, 7, 10, tzinfo=timezone.utc)

    updated = await CardQualityAnalysisService().update_issue_status(
        _FakeIssueSession(issue),
        account_id=1,
        issue_id=77,
        status="postponed",
        changed_by_user_id=10,
        reason="wait for photos",
        postponed_until=postponed_until,
    )

    assert updated.status == "postponed"
    assert updated.status_reason == "wait for photos"
    assert updated.postponed_until == postponed_until
    assert updated.resolved_at is None


@pytest.mark.asyncio
async def test_card_quality_blocked_is_deterministic_and_visible() -> None:
    issue = _issue()

    updated = await CardQualityAnalysisService().update_issue_status(
        _FakeIssueSession(issue),
        account_id=1,
        issue_id=77,
        status="blocked",
        changed_by_user_id=10,
        reason="needs supplier data",
    )

    payload = CardQualityAnalysisService()._issue_payload(issue)
    action = CardQualityAnalysisService()._action_from_issue(account_id=1, issue=issue)

    assert updated.status == "blocked"
    assert updated.status_reason == "needs supplier data"
    assert payload["status"] == "blocked"
    assert payload["status_reason"] == "needs supplier data"
    assert action.status == "blocked"


@pytest.mark.asyncio
async def test_card_quality_actions_and_issues_page_show_same_status() -> None:
    issue = _issue("blocked")
    issue.status_reason = "needs supplier data"
    service = CardQualityAnalysisService()

    action = service._action_from_issue(account_id=1, issue=issue)
    page = await service.list_issues(
        _FakeListIssuesSession(issue),
        account_id=1,
        category=None,
        status=None,
        include_info=True,
        limit=20,
        offset=0,
    )

    assert action.status == "blocked"
    assert page.items[0].status == "blocked"
    assert page.items[0].status_reason == "needs supplier data"
