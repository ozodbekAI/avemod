from __future__ import annotations

import base64
import csv
from io import BytesIO, StringIO
from typing import Any

from openpyxl import Workbook, load_workbook

from app.core.redaction import scrub_sensitive_payload


REGIONAL_REQUIRED = {"vendor_code", "region"}
HAND_REQUIRED = {"available_qty"}
HEADER_SCAN_ROWS = 20


HEADER_ALIASES = {
    "артикул продавца": "vendor_code",
    "артикул": "vendor_code",
    "vendor_code": "vendor_code",
    "article": "vendor_code",
    "название": "name",
    "предмет": "subject",
    "бренд": "brand",
    "артикул wb": "nm_id",
    "nm_id": "nm_id",
    "баркод": "barcode",
    "barcode": "barcode",
    "размер": "size_name",
    "size": "size_name",
    "регион": "region",
    "region": "region",
    "склад": "warehouse_name",
    "warehouse": "warehouse_name",
    "итого заказов, шт": "orders_qty",
    "заказали, шт": "orders_qty",
    "orders": "orders_qty",
    "остатки склад вб, шт": "stock_qty",
    "остатки склад вб шт": "stock_qty",
    "остатки на текущий день, шт": "stock_qty",
    "остаток": "stock_qty",
    "stock": "stock_qty",
    "наличие": "available_qty",
    "available_qty": "available_qty",
    "количество": "available_qty",
    "источник": "source_name",
}


def hand_stock_template_csv() -> str:
    return "vendor_code,nm_id,barcode,size_name,available_qty,source_name\nART-1,123456789,4600000000000,42,10,main\n"


def parse_table_upload(
    content: bytes, filename: str, *, import_type: str
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    rows = _read_rows(content, filename)
    normalized = [_normalize_row(row) for row in rows]
    if import_type == "regional_supply":
        normalized = [
            row
            for row in normalized
            if row.get("vendor_code") or row.get("nm_id") or row.get("region")
        ]
        warnings = _missing_warnings(normalized, REGIONAL_REQUIRED)
    else:
        normalized = [
            row
            for row in normalized
            if row.get("vendor_code") or row.get("nm_id") or row.get("available_qty")
        ]
        warnings = _missing_warnings(normalized, HAND_REQUIRED)
    metadata = {
        "file_name": filename,
        "sheet_name": "Детальные данные"
        if filename.lower().endswith(".xlsx")
        else None,
        "rows_total": len(normalized),
        "products": len(
            {
                (row.get("nm_id"), row.get("vendor_code"))
                for row in normalized
                if row.get("nm_id") or row.get("vendor_code")
            }
        ),
        "regions": len({row.get("region") for row in normalized if row.get("region")}),
        "sizes": len(
            {row.get("size_name") for row in normalized if row.get("size_name")}
        ),
        "warnings": warnings,
        "sample_rows": [scrub_sensitive_payload(row) for row in normalized[:10]],
    }
    return metadata, [scrub_sensitive_payload(row) for row in normalized]


def build_export_xlsx(
    *,
    summary: dict[str, Any],
    region_rows: list[dict[str, Any]],
    movements: list[dict[str, Any]],
    unmatched: list[dict[str, Any]] | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append(["key", "value"])
    for key, value in sorted(summary.items()):
        ws.append([key, str(value)])
    _append_sheet(wb, "region_rows", region_rows)
    _append_sheet(wb, "movements", movements)
    if unmatched:
        _append_sheet(wb, "unmatched", unmatched)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_base64(content: bytes) -> str:
    return base64.b64encode(content).decode("ascii")


def _read_rows(content: bytes, filename: str) -> list[dict[str, Any]]:
    lowered = filename.lower()
    if lowered.endswith(".csv"):
        text = content.decode("utf-8-sig")
        return list(csv.DictReader(StringIO(text)))
    if not lowered.endswith(".xlsx"):
        raise ValueError("Only .xlsx and .csv files are supported")
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet_name = (
        "Детальные данные" if "Детальные данные" in wb.sheetnames else wb.sheetnames[0]
    )
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header_index = _find_header_index(rows)
    if header_index is None:
        return []
    headers = [str(value or "").strip() for value in rows[header_index]]
    return [
        {
            headers[index]: value
            for index, value in enumerate(row)
            if index < len(headers)
        }
        for row in rows[header_index + 1 :]
    ]


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for key, value in row.items():
        mapped = HEADER_ALIASES.get(str(key or "").strip().casefold())
        if mapped:
            normalized[mapped] = value
    for key in ("nm_id", "orders_qty", "stock_qty", "available_qty"):
        if key in normalized:
            normalized[key] = _int_or_zero(normalized[key])
    for key in (
        "vendor_code",
        "barcode",
        "size_name",
        "region",
        "warehouse_name",
        "source_name",
    ):
        if key in normalized and normalized[key] is not None:
            normalized[key] = str(normalized[key]).strip()
    for key in ("name", "subject", "brand"):
        if key in normalized and normalized[key] is not None:
            normalized[key] = str(normalized[key]).strip()
    return normalized


def _find_header_index(rows: list[tuple[Any, ...]]) -> int | None:
    required = {"vendor_code", "region", "orders_qty", "stock_qty"}
    for index, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        mapped = {
            HEADER_ALIASES.get(str(value or "").strip().casefold())
            for value in row
            if str(value or "").strip()
        }
        if required.issubset({value for value in mapped if value}):
            return index
    return None


def _missing_warnings(rows: list[dict[str, Any]], required: set[str]) -> list[str]:
    if not rows:
        return ["empty_file"]
    warnings: list[str] = []
    for field in sorted(required):
        if not any(row.get(field) not in (None, "") for row in rows):
            warnings.append(f"missing_{field}")
    return warnings


def _append_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["empty"])
        return
    keys = sorted({key for row in rows for key in row})
    ws.append(keys)
    for row in rows:
        ws.append([row.get(key) for key in keys])


def _int_or_zero(value: Any) -> int:
    try:
        return max(int(float(value or 0)), 0)
    except Exception:
        return 0
