from __future__ import annotations

import hashlib
import io
import json
import re
import time
import copy
from dataclasses import dataclass
from dataclasses import replace
from datetime import datetime
from types import SimpleNamespace
from typing import Any

from sqlalchemy import String, cast, delete, func, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.cache import TTLMemoryCache
from app.core.config import get_settings
from app.core.http import WBAPIError
from app.core.time import utcnow
from app.core.wb_sync import DomainSyncBase
from app.models.accounts import WBAPICategory
from app.models.card_quality import (
    CardQualityAnalysisRun,
    CardQualityFixedFileEntry,
    CardQualityIssue,
    CardQualityIssueStatusHistory,
    CardQualitySnapshot,
)
from app.models.operator import PortalIntegration, PortalModuleSyncRun, ResultEvent
from app.models.product_cards import (
    WBProductCard,
    WBProductCardCharacteristic,
    WBProductCardSize,
)
from app.schemas.card_quality import (
    CardQualityAnalyzeResponse,
    CardQualityFixedFileEntriesPage,
    CardQualityFixedFileEntryRead,
    CardQualityFixedFileEntryMutation,
    CardQualityFixedFileStatus,
    CardQualityFixedFileUploadResponse,
    CardQualityIssueApplyPreview,
    CardQualityIssueFixResponse,
    CardQualityIssueRead,
    CardQualityIssuesGrouped,
    CardQualityIssuesPage,
    CardQualityProductListItem,
    CardQualityProductsPage,
    CardQualityProductRecheckResponse,
    CardQualityQueueProgress,
    CardQualityRunRead,
    CardQualityRunsPage,
    checker_action_capabilities,
    checker_contract_fields,
)
from app.schemas.portal import PortalActionRead, PortalProductQualityRead
from app.services.result_tracking import ResultTrackingService
from app.services.accounts import AccountService
from app.services.checker_problem_bridge import build_checker_problem_bridge
from app.services.guided_fixes import GuidedFixMapper
from app.services.checker_core.ai_fixer import CheckerAIFixer
from app.services.checker_core.text_policy import (
    describe_description_failures,
    normalize_generated_description,
    validate_description,
    validate_description_facts,
)
from app.services.checker_core.super_validator import super_validator_service
from app.services.checker_core.title_policy import (
    should_keep_current_title_as_safer,
    validate_title,
)
from app.services.checker_core.vision_service import (
    VisionService,
    build_product_dna_audit,
    is_grounded_product_dna_audit,
    product_dna_to_text,
)
from app.services.checker_core.wb_logic_prompt import build_wb_logic_block
from app.services.checker_core.wb_validator import (
    calculate_card_fcs,
    find_best_match,
    get_catalog,
    is_fixed_file_only_characteristic,
    is_no_touch_characteristic,
    validate_card_characteristics,
)


ACTIVE_ISSUE_STATUSES = {"new", "in_progress", "postponed", "blocked"}
PRESERVED_USER_DECISION_STATUSES = {"postponed", "ignored", "done", "blocked"}
MAX_FIX_RETRIES = 2
CHECKER_STATUS_TO_ACTION_STATUS = {
    "pending": "new",
    "new": "new",
    "fixed": "done",
    "auto_fixed": "done",
    "done": "done",
    "skipped": "ignored",
    "ignored": "ignored",
    "postponed": "postponed",
    "blocked": "blocked",
}
ACTION_STATUS_TO_CHECKER_STATUS = {
    "new": "new",
    "in_progress": "in_progress",
    "done": "done",
    "postponed": "postponed",
    "ignored": "ignored",
    "blocked": "blocked",
}
SEVERITY_WEIGHTS = {"critical": 25, "high": 15, "medium": 7, "low": 3, "info": 0}
CATEGORY_CAPS = {
    "title": 35,
    "description": 30,
    "characteristics": 30,
    "media": 30,
    "identity": 20,
    "completeness": 15,
}
DATE_CONTEXT_WORDS = {
    "дата",
    "сертификат",
    "сертифика",
    "декларац",
    "регистрац",
    "срок",
    "действия",
    "годен",
    "годности",
    "expiry",
    "certificate",
    "declaration",
    "issue date",
    "valid until",
    "validity",
}
DATE_FIELD_HINTS = {"date", "дата", "certificate", "сертификат", "декларац"}
VISUAL_ALWAYS_REVIEW_ONLY_FIELDS = {"фактура материала"}
VISUAL_REQUIRE_GROUNDED_DNA_FIELDS = {
    "комплектация",
    "тип верха",
    "тип низа",
    "модель брюк",
    "модель юбки",
    "модель костюма",
    "тип карманов",
    "особенности модели",
    "декоративные элементы",
    "вид застежки",
}
VISUAL_RISKY_FIELDS = (
    VISUAL_ALWAYS_REVIEW_ONLY_FIELDS | VISUAL_REQUIRE_GROUNDED_DNA_FIELDS
)
NON_UPDATABLE_WB_KEYS = {
    "photos",
    "photo",
    "videos",
    "video",
    "tags",
    "createdAt",
    "updatedAt",
    "needKiz",
    "nmUUID",
    "mediaFiles",
}
WB_CARD_UPDATE_TOP_LEVEL_KEYS = {
    "nmID",
    "vendorCode",
    "kizMarked",
    "brand",
    "title",
    "description",
    "dimensions",
    "characteristics",
    "sizes",
}
SOURCE_BASIC_ISSUE_ORDER = {
    "no_title": 0,
    "title_missing": 0,
    "title_too_short": 1,
    "title_policy_violation": 2,
    "no_photos": 10,
    "media_no_images": 10,
    "no_description": 20,
    "description_missing": 20,
    "few_photos": 30,
    "media_too_few_images": 30,
    "description_too_short": 40,
    "description_too_long": 41,
    "description_policy_violation": 42,
    "title_too_long": 50,
    "title_repeated_words": 51,
    "title_excessive_punctuation_caps": 52,
    "title_equals_vendor_code": 53,
    "no_video": 60,
    "media_no_video_info": 60,
    "add_more_photos": 70,
    "media_duplicate_urls": 71,
    "media_invalid_url": 72,
}
SOURCE_STATUS_ORDER = {
    "new": 0,
    "in_progress": 1,
    "blocked": 2,
    "postponed": 3,
    "ignored": 4,
    "done": 5,
    "resolved": 6,
}
SOURCE_SEVERITY_ORDER = {
    "critical": 0,
    "high": 1,
    "medium": 2,
    "warning": 2,
    "low": 3,
    "info": 4,
}


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(item).strip() for item in value if str(item).strip())
    return str(value).strip()


def _parse_multi_value(value: Any) -> list[str]:
    if value is None:
        return []
    raw = _as_text(value)
    if not raw or raw == "__CLEAR__":
        return []
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list):
            return [str(item).strip() for item in parsed if str(item).strip()]
        if parsed is not None:
            text = str(parsed).strip()
            return [text] if text else []
    except Exception:
        pass
    if ";" in raw:
        return [part.strip() for part in raw.split(";") if part.strip()]
    if "," in raw:
        return [part.strip() for part in raw.split(",") if part.strip()]
    return [raw]


def _canonical_apply_field_path(
    field_path: str | None, category: str | None = None
) -> str:
    raw = str(field_path or "").strip()
    lower = raw.lower()
    if not lower:
        cat = str(category or "").strip().lower()
        if cat == "title":
            return "title"
        if cat in {"description", "seo"}:
            return "description"
        return ""
    if lower in {"subject", "subject_name", "category"}:
        return "subject_name"
    if lower in {"kizmarked", "kiz_marked"}:
        return "kizmarked"
    if lower.startswith("characteristics."):
        return f"characteristics.{raw.split('.', 1)[1].strip().lower()}"
    return lower


def _set_characteristic_value(
    characteristics: list[dict[str, Any]],
    field_path: str,
    value: Any,
    charc_id: int | None = None,
) -> None:
    char_name = str(field_path or "").split(".", 1)[1].strip()
    parsed_value = _parse_multi_value(value)
    for item in characteristics:
        if not isinstance(item, dict):
            continue
        same_name = str(item.get("name") or "").strip().lower() == char_name.lower()
        same_id = bool(charc_id and str(item.get("id")) == str(charc_id))
        if same_name or same_id:
            item["value"] = parsed_value
            if charc_id and not item.get("id"):
                item["id"] = charc_id
            if not item.get("name"):
                item["name"] = char_name
            return
    new_item: dict[str, Any] = {"name": char_name, "value": parsed_value}
    if charc_id:
        new_item["id"] = charc_id
    characteristics.append(new_item)


def _sanitize_wb_update_snapshot(raw_data: dict[str, Any]) -> dict[str, Any]:
    snapshot = {
        key: copy.deepcopy(value)
        for key, value in (raw_data or {}).items()
        if key in WB_CARD_UPDATE_TOP_LEVEL_KEYS
    }
    for key in NON_UPDATABLE_WB_KEYS:
        snapshot.pop(key, None)
    if "nmID" not in snapshot or snapshot.get("nmID") in (None, ""):
        raise ValueError("WB card update requires nmID")
    if not snapshot.get("vendorCode"):
        raise ValueError("WB card update requires vendorCode")
    if not isinstance(snapshot.get("sizes"), list):
        raise ValueError("WB card update requires sizes list")
    snapshot["nmID"] = int(snapshot["nmID"])
    snapshot["vendorCode"] = str(snapshot["vendorCode"]).strip()
    snapshot["kizMarked"] = bool(snapshot.get("kizMarked", False))
    return snapshot


def _embedded_wb_error_details(data: Any) -> str:
    if not isinstance(data, dict):
        return ""
    error_text = str(data.get("errorText") or data.get("message") or "").strip()
    additional = data.get("additionalErrors")
    if data.get("error") is True and error_text:
        return error_text
    if data.get("error") is True and additional:
        try:
            return json.dumps(additional, ensure_ascii=False)
        except TypeError:
            return str(additional)
    if error_text:
        return error_text
    if additional:
        try:
            return json.dumps(additional, ensure_ascii=False)
        except TypeError:
            return str(additional)
    return ""


def _norm_text(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _field_key(field_name: str | None) -> str:
    raw = _norm_text(field_name)
    if raw.startswith("characteristics."):
        raw = raw.split("characteristics.", 1)[1].strip()
    return raw


def _split_issue_values(raw_value: Any) -> list[str]:
    if isinstance(raw_value, list):
        return [str(item).strip() for item in raw_value if str(item).strip()]
    raw = str(raw_value or "").strip()
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list):
            return [str(item).strip() for item in parsed if str(item).strip()]
        if parsed is not None and not isinstance(parsed, (dict, list)):
            raw = str(parsed).strip()
    except Exception:
        pass
    parts = re.split(r"[;,]", raw.strip().strip("[]"))
    return [part.strip().strip("'\"") for part in parts if part.strip().strip("'\"")]


def _issue_values_equivalent(current_value: Any, suggested_value: Any) -> bool:
    current = _norm_text(current_value)
    suggested = _norm_text(suggested_value)
    if not current or not suggested:
        return False
    if current == suggested:
        return True
    current_parts = [_norm_text(item) for item in _split_issue_values(current_value)]
    suggested_parts = [
        _norm_text(item) for item in _split_issue_values(suggested_value)
    ]
    if not current_parts or not suggested_parts:
        return False
    return current_parts == suggested_parts or (
        len(current_parts) > 1 and set(current_parts) == set(suggested_parts)
    )


def _contains_date_context_text(value: Any) -> bool:
    text = _norm_text(value)
    if not text:
        return False
    tokens = set(re.findall(r"[a-zа-яё]+", text))
    exact_markers = {"date", "дата", "expiry", "certificate", "declaration"}
    if tokens & exact_markers:
        return True
    stem_markers = DATE_CONTEXT_WORDS - {"дата", "date"}
    return any(marker in text for marker in stem_markers)


@dataclass(frozen=True)
class NormalizedCard:
    account_id: int
    nm_id: int
    source_card_id: int
    title: str
    description: str
    brand: str
    subject_name: str
    vendor_code: str
    characteristics: list[dict[str, Any]]
    photos: list[dict[str, Any]]
    videos: list[str]
    sizes: list[dict[str, Any]]
    source_revision: str
    source_updated_at: datetime | None
    subject_id: int | None = None
    product_dna_json: dict[str, Any] | None = None
    product_dna_audit: dict[str, Any] | None = None
    product_dna_text: str = ""


@dataclass(frozen=True)
class RuleIssue:
    issue_code: str
    category: str
    severity: str
    title: str
    business_explanation: str
    recommended_fix: str
    field_name: str | None = None
    current_value_json: Any | None = None
    expected_value_json: Any | None = None
    confidence: float = 1.0
    suggested_value: str | None = None
    alternatives: list[Any] | None = None
    charc_id: int | None = None
    allowed_values: list[Any] | None = None
    error_details: list[Any] | None = None
    ai_suggested_value: str | None = None
    ai_reason: str | None = None
    ai_alternatives: list[Any] | None = None
    ai_confidence: float | None = None
    requires_human_check: bool = False
    ai_reason_short: str | None = None
    ai_reason_full: str | None = None
    ai_evidence: dict[str, Any] | None = None
    ai_used_sources: list[Any] | None = None
    photo_evidence: list[Any] | None = None
    source: str | None = "code"
    score_impact: int = 0


class CardQualityNormalizationService:
    async def normalize_product(
        self, session: AsyncSession, *, account_id: int, nm_id: int
    ) -> NormalizedCard | None:
        card = (
            await session.execute(
                select(WBProductCard).where(
                    WBProductCard.account_id == account_id, WBProductCard.nm_id == nm_id
                )
            )
        ).scalar_one_or_none()
        if card is None:
            return None
        characteristics = list(
            (
                await session.execute(
                    select(WBProductCardCharacteristic)
                    .where(
                        WBProductCardCharacteristic.account_id == account_id,
                        WBProductCardCharacteristic.product_card_id == card.id,
                    )
                    .order_by(WBProductCardCharacteristic.id)
                )
            ).scalars()
        )
        sizes = list(
            (
                await session.execute(
                    select(WBProductCardSize)
                    .where(
                        WBProductCardSize.account_id == account_id,
                        WBProductCardSize.product_card_id == card.id,
                    )
                    .order_by(WBProductCardSize.id)
                )
            ).scalars()
        )
        normalized_characteristics = [
            {
                "char_id": item.char_id,
                "name": self._clean_text(item.name),
                "value": self._jsonable(item.value),
            }
            for item in characteristics
        ]
        normalized_sizes = [
            {
                "chrt_id": item.chrt_id,
                "size_id": item.size_id,
                "tech_size": self._clean_text(item.tech_size),
                "skus": self._jsonable(item.skus),
            }
            for item in sizes
        ]
        photos = self._extract_photos(card.photos, card.payload)
        videos = self._extract_videos(card.video, card.payload)
        payload = {
            "checker_engine_version": "finance_checker_core_wb_optimizer_v2",
            "account_id": account_id,
            "nm_id": nm_id,
            "title": self._clean_text(card.title),
            "description": self._clean_text(card.description),
            "brand": self._clean_text(card.brand),
            "subject_name": self._clean_text(card.subject_name),
            "vendor_code": self._clean_text(card.vendor_code),
            "characteristics": normalized_characteristics,
            "photos": photos,
            "videos": videos,
            "sizes": normalized_sizes,
            "source_updated_at": self._iso(card.updated_at_wb or card.updated_at),
        }
        source_revision = hashlib.sha256(
            json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str).encode(
                "utf-8"
            )
        ).hexdigest()
        return NormalizedCard(
            account_id=account_id,
            nm_id=nm_id,
            source_card_id=int(card.id),
            title=payload["title"],
            description=payload["description"],
            brand=payload["brand"],
            subject_id=card.subject_id,
            subject_name=payload["subject_name"],
            vendor_code=payload["vendor_code"],
            characteristics=normalized_characteristics,
            photos=photos,
            videos=videos,
            sizes=normalized_sizes,
            source_revision=source_revision,
            source_updated_at=card.updated_at_wb or card.updated_at,
        )

    def _clean_text(self, value: Any) -> str:
        return re.sub(r"\s+", " ", str(value or "")).strip()

    def _jsonable(self, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, (str, int, float, bool, list, dict)):
            return value
        return str(value)

    def _extract_photos(self, photos: Any, payload: Any) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []
        seen: set[str] = set()
        for source in (
            photos,
            (payload or {}).get("photos") if isinstance(payload, dict) else None,
            (payload or {}).get("media") if isinstance(payload, dict) else None,
        ):
            self._collect_photo_objects(source, result, seen)
        return result

    def _extract_videos(self, video: Any, payload: Any) -> list[str]:
        result: list[str] = []
        if video:
            result.append(str(video))
        if isinstance(payload, dict):
            for key in ("video", "videos"):
                self._collect_urls(payload.get(key), result)
        return list(dict.fromkeys(item for item in result if item))

    def _collect_urls(self, value: Any, result: list[str]) -> None:
        if isinstance(value, str):
            if value.strip():
                result.append(value.strip())
            return
        if isinstance(value, list):
            for item in value:
                self._collect_urls(item, result)
            return
        if isinstance(value, dict):
            for key in ("big", "c516x688", "tm", "url", "link", "photo", "src"):
                if value.get(key):
                    self._collect_urls(value.get(key), result)

    def _collect_photo_objects(
        self, value: Any, result: list[dict[str, Any]], seen: set[str]
    ) -> None:
        if isinstance(value, str):
            canonical_url = value.strip()
            if canonical_url and canonical_url not in seen:
                seen.add(canonical_url)
                result.append({"canonical_url": canonical_url, "variants": {}})
            return
        if isinstance(value, list):
            for item in value:
                self._collect_photo_objects(item, result, seen)
            return
        if not isinstance(value, dict):
            return
        variants: dict[str, str] = {}
        for key in ("big", "c516x688", "tm", "url", "link", "photo", "src"):
            url = value.get(key)
            if isinstance(url, str) and url.strip():
                variants[key] = url.strip()
        if variants:
            canonical_url = (
                variants.get("big")
                or variants.get("url")
                or variants.get("link")
                or next(iter(variants.values()))
            )
            if canonical_url not in seen:
                seen.add(canonical_url)
                result.append({"canonical_url": canonical_url, "variants": variants})
            return
        for nested in value.values():
            self._collect_photo_objects(nested, result, seen)

    def _iso(self, value: Any) -> str | None:
        return value.isoformat() if hasattr(value, "isoformat") else None


class CardQualityRuleEngine:
    def __init__(self) -> None:
        self.settings = get_settings()

    def analyze(self, card: NormalizedCard) -> tuple[list[RuleIssue], dict[str, Any]]:
        issues: list[RuleIssue] = []
        issues.extend(self._source_basic_rules(card))
        issues.extend(self._wb_catalog_rules(card))
        score, category_scores = self.score(issues)
        counts = self._counts(issues)
        summary = {
            "score": score,
            "status": self.status(score, issues),
            "category_scores": category_scores,
            "issues_count": len(issues),
            **counts,
            "checked_categories": [
                "title",
                "description",
                "characteristics",
                "photos",
                "video",
            ],
            "wb_logic_prompt_version": "wb_optimizer_core",
            "checker_pipeline": "source_wb_optimizer_order",
        }
        try:
            summary["fcs"] = calculate_card_fcs(self._wb_card_payload(card))
        except Exception as exc:
            summary["fcs"] = {"error": str(exc)[:200]}
        return issues, summary

    def _source_basic_rules(self, card: NormalizedCard) -> list[RuleIssue]:
        """Port of wb-optimizer CardAnalyzer rule order.

        The source checker intentionally keeps only structural media/text
        checks before WB catalog validation and AI. Extra local heuristics here
        change the downstream prompt inputs, so this list mirrors source order.
        """
        settings = self.settings
        min_title = int(getattr(settings, "MIN_TITLE_LENGTH", 40) or 40)
        max_title = int(getattr(settings, "MAX_TITLE_LENGTH", 60) or 60)
        min_description = int(getattr(settings, "MIN_DESCRIPTION_LENGTH", 1000) or 1000)
        max_description = int(getattr(settings, "MAX_DESCRIPTION_LENGTH", 1800) or 1800)
        min_photos = max(1, int(getattr(settings, "MIN_PHOTOS_COUNT", 3) or 3))
        recommended_photos = max(
            min_photos,
            int(
                getattr(settings, "RECOMMENDED_PHOTOS_COUNT", min_photos) or min_photos
            ),
        )
        title = str(card.title or "").strip()
        description = str(card.description or "").strip()
        photos_count = len(card.photos or [])
        videos_count = len(card.videos or [])
        issues: list[RuleIssue] = []

        if not title:
            issues.append(
                self._issue(
                    "no_title",
                    "title",
                    "critical",
                    "Отсутствует название товара",
                    "Карточка без названия не будет отображаться в поиске",
                    "Сгенерировать название через dedicated title prompt по подтверждённым фактам карточки.",
                    "title",
                    None,
                    score_impact=30,
                )
            )
        elif len(title) < min_title:
            issues.append(
                self._issue(
                    "title_too_short",
                    "title",
                    "critical",
                    "Название слишком короткое",
                    f"Название должно быть минимум {min_title} символов для хорошей индексации",
                    "Сгенерировать новое название через dedicated title prompt.",
                    "title",
                    title,
                    score_impact=20,
                )
            )

        if photos_count == 0:
            issues.append(
                self._issue(
                    "no_photos",
                    "photos",
                    "critical",
                    "Отсутствуют фотографии",
                    "Карточка без фотографий не будет показываться покупателям",
                    "Добавить фотографии товара.",
                    "photos",
                    "0 фото",
                    score_impact=30,
                )
            )

        if not description:
            issues.append(
                self._issue(
                    "no_description",
                    "description",
                    "critical",
                    "Отсутствует описание",
                    "Описание важно для SEO и конверсии",
                    "Сгенерировать описание через dedicated description prompt.",
                    "description",
                    None,
                    score_impact=20,
                )
            )

        if 0 < photos_count < min_photos:
            issues.append(
                self._issue(
                    "few_photos",
                    "photos",
                    "medium",
                    "Недостаточно фотографий",
                    f"Для карточки нужно минимум {min_photos} фото, иначе покупателю не хватает базового визуального контекста",
                    f"Добавьте ещё {min_photos - photos_count} фото, чтобы довести карточку до {min_photos}",
                    "photos",
                    f"{photos_count} фото",
                    f"Добавьте ещё {min_photos - photos_count} фото, чтобы довести карточку до {min_photos}",
                    score_impact=10,
                )
            )

        if description and len(description) < min_description:
            issues.append(
                self._issue(
                    "description_too_short",
                    "description",
                    "medium",
                    "Описание слишком короткое",
                    f"Описание должно быть минимум {min_description} символов",
                    "Сгенерировать полное описание через dedicated description prompt.",
                    "description",
                    description[:500],
                    score_impact=10,
                )
            )

        if description and len(description) > max_description:
            issues.append(
                self._issue(
                    "description_too_long",
                    "description",
                    "medium",
                    "Описание слишком длинное",
                    f"Описание не должно превышать {max_description} символов",
                    "Сгенерировать новое описание в лимите WB checker.",
                    "description",
                    description[:500],
                    score_impact=8,
                )
            )

        if description:
            ctx = self._wb_card_payload(card)
            failures = describe_description_failures(ctx)
            if not failures.get("valid", True):
                issues.append(
                    self._issue(
                        "description_policy_violation",
                        "description",
                        "medium",
                        "Описание не проходит политику качества",
                        "Текущее описание нарушает требования к структуре или содержит запрещенные элементы",
                        "Сгенерировать описание только по подтверждённым фактам.",
                        "description",
                        description[:500],
                        {
                            "policy": "wb_optimizer_description_policy",
                            "details": failures,
                        },
                        alternatives=[],
                        error_details=[failures],
                        score_impact=8,
                    )
                )

        if title and len(title) > max_title:
            shortened = title[:max_title].rsplit(" ", 1)[0] or title[:max_title]
            issues.append(
                self._issue(
                    "title_too_long",
                    "title",
                    "medium",
                    "Название слишком длинное",
                    f"Название обрезается после {max_title} символов",
                    "Сгенерировать короткое название через dedicated title prompt.",
                    "title",
                    title,
                    shortened,
                    score_impact=5,
                )
            )

        if videos_count == 0:
            issues.append(
                self._issue(
                    "no_video",
                    "video",
                    "medium",
                    "Отсутствует видео",
                    "Видео повышает конверсию на 30-40%",
                    "Добавьте видео для повышения конверсии",
                    "videos",
                    "Нет видео",
                    "Добавьте видео для повышения конверсии",
                    score_impact=1,
                )
            )

        if min_photos <= photos_count < recommended_photos:
            issues.append(
                self._issue(
                    "add_more_photos",
                    "photos",
                    "medium",
                    "Добавьте больше фото",
                    "Дополнительные фото улучшают конверсию",
                    f"Рекомендуем {recommended_photos} фото",
                    "photos",
                    f"{photos_count} фото",
                    f"Рекомендуем {recommended_photos} фото",
                    score_impact=5,
                )
            )

        return issues

    def score(self, issues: list[RuleIssue]) -> tuple[int, dict[str, int]]:
        category_penalties: dict[str, int] = {}
        for issue in issues:
            category = issue.category
            category_penalties[category] = min(
                category_penalties.get(category, 0)
                + SEVERITY_WEIGHTS.get(issue.severity, 0),
                CATEGORY_CAPS.get(category, 25),
            )
        total_penalty = min(sum(category_penalties.values()), 100)
        category_scores = {
            category: max(0, 100 - penalty)
            for category, penalty in category_penalties.items()
        }
        for category in (
            "title",
            "description",
            "characteristics",
            "photos",
            "video",
            "media",
            "identity",
            "completeness",
        ):
            category_scores.setdefault(category, 100)
        return max(0, 100 - total_penalty), category_scores

    def status(self, score: int, issues: list[RuleIssue]) -> str:
        severities = {issue.severity for issue in issues}
        if "critical" in severities or score < 50:
            return "critical"
        if severities & {"high", "medium", "low"} or score < 90:
            return "warning"
        return "clean"

    def fingerprint(self, card: NormalizedCard, issue: RuleIssue) -> str:
        context = {
            "account_id": card.account_id,
            "nm_id": card.nm_id,
            "code": issue.issue_code,
            "category": issue.category,
            "field": _field_key(issue.field_name),
            "charc_id": issue.charc_id or 0,
            "current": self._stable_value_context(issue.current_value_json),
            "expected": self._stable_value_context(issue.expected_value_json),
        }
        raw = json.dumps(context, ensure_ascii=False, sort_keys=True, default=str)
        return hashlib.sha256(raw.encode("utf-8")).hexdigest()

    def restore_key(
        self, issue_code: str | None, field_name: str | None, charc_id: int | None
    ) -> tuple[str, str, int]:
        return (
            str(issue_code or "").strip().lower(),
            _field_key(field_name),
            int(charc_id or 0),
        )

    def _stable_value_context(self, value: Any) -> Any:
        if isinstance(value, dict):
            return {
                str(key): self._stable_value_context(value[key])
                for key in sorted(value)
            }
        if isinstance(value, list):
            return [self._stable_value_context(item) for item in value]
        if isinstance(value, str):
            return _norm_text(value)
        return value

    def _title_rules(self, card: NormalizedCard) -> list[RuleIssue]:
        title = card.title
        issues: list[RuleIssue] = []
        card_ctx = self._wb_card_payload(card)
        if not title:
            return [
                self._issue(
                    "title_missing",
                    "title",
                    "critical",
                    "Название карточки отсутствует",
                    "Покупатель и поиск WB не видят нормальное название товара.",
                    "Заполнить понятное название с типом товара, брендом и ключевым отличием.",
                    "title",
                    title,
                )
            ]
        policy_valid, policy_reason = (True, "")
        if card.subject_id is not None:
            policy_valid, policy_reason = validate_title(
                title, card_ctx, strict_content=True
            )
        if not policy_valid:
            severity = (
                "high"
                if "корот" in policy_reason.lower()
                or "категори" in policy_reason.lower()
                else "medium"
            )
            issues.append(
                self._issue(
                    "title_policy_violation",
                    "title",
                    severity,
                    "Название не проходит проверку качества",
                    policy_reason or "Название не соответствует правилам checker.",
                    "Подготовить новое название по фактам из карточки: категория, модельный признак, подтверждённые характеристики.",
                    "title",
                    title,
                    {"policy": "wb_optimizer_title_policy", "reason": policy_reason},
                )
            )
        if len(title) < 12:
            issues.append(
                self._issue(
                    "title_too_short",
                    "title",
                    "high",
                    "Название слишком короткое",
                    "Короткое название обычно хуже объясняет товар и его назначение.",
                    "Расширить название: добавить тип товара, материал/назначение и ключевой признак.",
                    "title",
                    title,
                    {"min_length": 12},
                )
            )
        if len(title) > 120:
            issues.append(
                self._issue(
                    "title_too_long",
                    "title",
                    "medium",
                    "Название слишком длинное",
                    "Слишком длинное название сложно сканировать и может выглядеть как спам.",
                    "Сократить название до основного товара и важных характеристик.",
                    "title",
                    title,
                    {"max_length": 120},
                )
            )
        words = [word.lower() for word in re.findall(r"[A-Za-zА-Яа-яЁё0-9]+", title)]
        if any(words.count(word) >= 3 for word in set(words) if len(word) > 2):
            issues.append(
                self._issue(
                    "title_repeated_words",
                    "title",
                    "medium",
                    "В названии повторяются слова",
                    "Повторы снижают читаемость и могут выглядеть как keyword stuffing.",
                    "Убрать лишние повторы слов.",
                    "title",
                    title,
                )
            )
        if re.search(r"[!?.,;:]{4,}", title) or sum(
            1 for ch in title if ch.isupper()
        ) > max(12, len(title) * 0.7):
            issues.append(
                self._issue(
                    "title_excessive_punctuation_caps",
                    "title",
                    "low",
                    "Название выглядит агрессивно оформленным",
                    "Избыточные знаки или caps lock ухудшают доверие к карточке.",
                    "Использовать обычный регистр и умеренную пунктуацию.",
                    "title",
                    title,
                )
            )
        if (
            card.vendor_code
            and title.strip().lower() == card.vendor_code.strip().lower()
        ):
            issues.append(
                self._issue(
                    "title_equals_vendor_code",
                    "identity",
                    "high",
                    "Название совпадает с артикулом продавца",
                    "Артикул не объясняет покупателю товар.",
                    "Заменить артикул на понятное товарное название.",
                    "title",
                    title,
                )
            )
        return issues

    def _description_rules(self, card: NormalizedCard) -> list[RuleIssue]:
        description = card.description
        card_ctx = self._wb_card_payload(card)
        if not description:
            return [
                self._issue(
                    "description_missing",
                    "description",
                    "high",
                    "Описание отсутствует",
                    "Описание помогает объяснить состав, посадку, применение и ограничения товара.",
                    "Добавить описание с ключевыми свойствами товара.",
                    "description",
                    description,
                )
            ]
        issues: list[RuleIssue] = []
        policy_valid, policy_reason = (True, "")
        if card.subject_id is not None:
            policy_valid, policy_reason = validate_description(
                description, card_ctx, strict_structure=True
            )
        if not policy_valid:
            details = describe_description_failures(card_ctx)
            issues.append(
                self._issue(
                    "description_policy_violation",
                    "description",
                    "medium",
                    "Описание не проходит проверку качества",
                    policy_reason or "Описание не соответствует правилам checker.",
                    "Переписать описание только по подтверждённым фактам: состав, посадка, комплектация, уход и сценарии использования.",
                    "description",
                    description,
                    {
                        "policy": "wb_optimizer_description_policy",
                        "reason": policy_reason,
                        "details": details,
                    },
                )
            )
        if len(description) < 80:
            issues.append(
                self._issue(
                    "description_too_short",
                    "description",
                    "medium",
                    "Описание слишком короткое",
                    "Короткое описание часто не закрывает вопросы покупателя.",
                    "Добавить материал, назначение, комплектацию, уход и сценарии использования.",
                    "description",
                    description,
                    {"min_length": 80},
                )
            )
        if description.strip().lower() == card.title.strip().lower():
            issues.append(
                self._issue(
                    "description_duplicates_title",
                    "description",
                    "medium",
                    "Описание дублирует название",
                    "Дублирование не добавляет полезной информации.",
                    "Написать отдельное описание, а не копию названия.",
                    "description",
                    description,
                )
            )
        if len(set(re.findall(r"[A-Za-zА-Яа-яЁё0-9]+", description.lower()))) < 8:
            issues.append(
                self._issue(
                    "description_no_useful_details",
                    "description",
                    "medium",
                    "В описании мало полезных деталей",
                    "Покупателю не хватает фактов о товаре.",
                    "Добавить конкретные характеристики и преимущества без выдуманных данных.",
                    "description",
                    description,
                )
            )
        return issues

    def _characteristic_rules(self, card: NormalizedCard) -> list[RuleIssue]:
        if not card.characteristics:
            return [
                self._issue(
                    "characteristics_missing",
                    "characteristics",
                    "high",
                    "Характеристики отсутствуют",
                    "Фильтры и сравнение товаров зависят от характеристик.",
                    "Заполнить доступные характеристики товара.",
                    "characteristics",
                    [],
                )
            ]
        issues: list[RuleIssue] = []
        seen: dict[str, Any] = {}
        card_payload = self._wb_card_payload(card)
        for item in card.characteristics:
            name = str(item.get("name") or "").strip().lower()
            value = item.get("value")
            if not name:
                issues.append(
                    self._issue(
                        "characteristic_name_missing",
                        "characteristics",
                        "low",
                        "Есть характеристика без названия",
                        "Такую характеристику сложно использовать в проверке качества.",
                        "Проверить синхронизацию характеристики.",
                        "characteristics",
                        item,
                    )
                )
                continue
            if is_no_touch_characteristic(
                name, card=card_payload
            ) or is_fixed_file_only_characteristic(name):
                continue
            if value in (None, "", [], {}):
                issues.append(
                    self._issue(
                        "characteristic_value_empty",
                        "characteristics",
                        "medium",
                        f"Пустое значение характеристики: {item.get('name')}",
                        "Пустые значения ухудшают полноту карточки.",
                        "Заполнить значение характеристики или удалить пустой атрибут.",
                        str(item.get("name") or name),
                        value,
                    )
                )
            if name in seen and seen[name] != value:
                issues.append(
                    self._issue(
                        "characteristic_conflicting_values",
                        "characteristics",
                        "medium",
                        f"Конфликт значений характеристики: {item.get('name')}",
                        "Разные значения одного поля могут запутать покупателя.",
                        "Оставить одно корректное значение характеристики.",
                        str(item.get("name") or name),
                        [seen[name], value],
                    )
                )
            seen[name] = value
        high_value = {"цвет", "состав", "материал", "размер", "пол", "комплектация"}
        present = set(seen)
        if card.subject_name and not present.intersection(high_value):
            issues.append(
                self._issue(
                    "high_value_characteristics_missing",
                    "characteristics",
                    "low",
                    "Нет базовых характеристик для фильтров",
                    "Цвет, материал, размер или комплектация часто нужны для выбора товара.",
                    "Проверить и заполнить доступные базовые характеристики.",
                    "characteristics",
                    list(present),
                    sorted(high_value),
                )
            )
        return issues

    def _wb_catalog_rules(self, card: NormalizedCard) -> list[RuleIssue]:
        if card.subject_id is None:
            return []
        card_payload = self._wb_card_payload(card)
        try:
            rows = validate_card_characteristics(card_payload)
        except Exception as exc:
            return [
                self._issue(
                    "wb_catalog_unavailable",
                    "characteristics",
                    "info",
                    "Проверка WB-справочника недоступна",
                    "Локальный справочник характеристик WB не смог проверить карточку.",
                    "Проверить подключение справочника checker и повторить анализ карточки.",
                    "characteristics",
                    {"error": str(exc)[:200]},
                )
            ]
        issues: list[RuleIssue] = []
        for row in rows:
            code = (
                str(
                    row.get("category") or row.get("code") or "wb_catalog_issue"
                ).strip()
                or "wb_catalog_issue"
            )
            name = str(row.get("name") or "Характеристика").strip()
            if is_no_touch_characteristic(
                name, card=card_payload
            ) or is_fixed_file_only_characteristic(name):
                continue
            is_fixed = bool(row.get("is_fixed_field"))
            wb_severity = str(row.get("severity") or "warning").strip().lower()
            severity = (
                "critical" if wb_severity == "critical" and not is_fixed else "medium"
            )
            allowed_values = (
                row.get("allowed_values")
                if isinstance(row.get("allowed_values"), list)
                else []
            )
            suggested = self._stringify_suggestion(row.get("suggested_value"))
            message = str(row.get("message") or "Ошибка характеристики").strip()
            explanation = self._format_wb_error_description(row)
            errors = row.get("errors") if isinstance(row.get("errors"), list) else []
            requires_human_check = self._requires_human_check(
                name=name, errors=errors, suggested=suggested
            )
            evidence = {
                "observed": [explanation],
                "constraint": "WB allowed values / limits",
                "wb_logic": build_wb_logic_block(include_output=False),
            }
            issues.append(
                RuleIssue(
                    issue_code=f"wb_fixed_{code}" if is_fixed else f"wb_{code}",
                    category="characteristics",
                    severity=severity,
                    title=message[:500],
                    business_explanation=explanation,
                    recommended_fix=self._wb_recommended_fix(
                        name=name,
                        suggested=suggested,
                        requires_human_check=requires_human_check,
                    ),
                    field_name=f"characteristics.{name}",
                    current_value_json=row.get("value"),
                    expected_value_json={
                        "allowed_values": allowed_values[:50],
                        "errors": errors,
                    },
                    confidence=0.9
                    if suggested and not requires_human_check and not is_fixed
                    else 0.65,
                    suggested_value=None
                    if is_fixed or requires_human_check
                    else suggested,
                    alternatives=self._extract_issue_example_values(errors)
                    or [item for item in allowed_values[:8] if str(item).strip()],
                    charc_id=self._int(row.get("charc_id")),
                    allowed_values=allowed_values,
                    error_details=errors,
                    ai_suggested_value=None
                    if is_fixed or requires_human_check
                    else suggested,
                    ai_reason=message,
                    ai_alternatives=[
                        item for item in allowed_values[:8] if str(item).strip()
                    ],
                    ai_confidence=0.9
                    if suggested and not requires_human_check and not is_fixed
                    else 0.65,
                    requires_human_check=requires_human_check or is_fixed,
                    ai_reason_short=message[:240],
                    ai_reason_full=explanation,
                    ai_evidence=evidence,
                    ai_used_sources=["allowed_values", "card_characteristics"],
                    source="auto_fix"
                    if row.get("auto_fixed") and suggested and not is_fixed
                    else "code",
                    score_impact=0
                    if is_fixed
                    else self._calculate_wb_score_impact(row),
                )
            )
        return issues

    def _extract_issue_example_values(
        self, error_details: list[Any] | None, limit: int = 12
    ) -> list[str]:
        out: list[str] = []
        seen: set[str] = set()
        for item in error_details or []:
            if not isinstance(item, dict):
                continue
            for raw in item.get("exampleValues") or []:
                value = str(raw or "").strip()
                if not value or value in seen:
                    continue
                seen.add(value)
                out.append(value)
                if len(out) >= limit:
                    return out
        return out

    def _format_wb_error_description(self, row: dict[str, Any]) -> str:
        parts: list[str] = []
        for err in row.get("errors") or []:
            if not isinstance(err, dict):
                continue
            err_type = str(err.get("type") or "").strip()
            if err_type == "limit":
                parts.append(
                    f"Лимит: {err.get('min', 0)}-{err.get('max', 999)}, текущее: {err.get('actual')}"
                )
            elif err_type == "allowed_values":
                invalid = (err.get("invalidValues") or [])[:3]
                if invalid:
                    parts.append(
                        f"Недопустимые значения: {', '.join(str(x) for x in invalid)}"
                    )
            elif err_type == "wrong_category":
                parts.append(
                    str(
                        err.get("message")
                        or "Характеристика не входит в список допустимых для данной категории"
                    )
                )
        return "; ".join(parts) if parts else str(row.get("message") or "")

    def _calculate_wb_score_impact(self, row: dict[str, Any]) -> int:
        errors = row.get("errors") or []
        if any(
            isinstance(item, dict) and item.get("type") == "wrong_category"
            for item in errors
        ):
            return 10
        base = 8
        if any(
            isinstance(item, dict) and item.get("type") == "limit" for item in errors
        ):
            base += 4
        if any(
            isinstance(item, dict) and item.get("type") == "allowed_values"
            for item in errors
        ):
            base += 2
        return min(base, 15)

    def _wb_card_payload(self, card: NormalizedCard) -> dict[str, Any]:
        return {
            "nmID": card.nm_id,
            "nm_id": card.nm_id,
            "subjectID": card.subject_id,
            "subject_id": card.subject_id,
            "subjectName": card.subject_name,
            "subject_name": card.subject_name,
            "brand": card.brand,
            "title": card.title,
            "description": card.description,
            "vendorCode": card.vendor_code,
            "vendor_code": card.vendor_code,
            "characteristics": card.characteristics,
            "photos": [
                self._photo_url(item) for item in card.photos if self._photo_url(item)
            ],
            "video": card.videos,
        }

    def _wb_severity(self, value: Any, category: Any) -> str:
        raw = str(value or "").strip().lower()
        cat = str(category or "").strip().lower()
        if raw in {"critical", "high", "medium", "low", "info"}:
            return raw
        if cat in {"required_missing", "missing_required", "wrong_category"}:
            return "high"
        if cat in {"allowed_values", "limit", "fixed_field"}:
            return "medium"
        return "low"

    def _requires_human_check(
        self, *, name: str, errors: list[Any], suggested: str | None
    ) -> bool:
        key = _field_key(name)
        destructive = any(
            isinstance(item, dict)
            and str(item.get("fix_action") or item.get("type") or "").lower()
            in {"clear", "swap"}
            for item in errors
        )
        return (
            key in VISUAL_RISKY_FIELDS
            or destructive
            or not bool(str(suggested or "").strip())
        )

    def _wb_recommended_fix(
        self, *, name: str, suggested: str | None, requires_human_check: bool
    ) -> str:
        if suggested and not requires_human_check:
            return f"Заменить значение «{name}» на «{suggested}» из справочника WB."
        if suggested:
            return f"Проверить «{name}» вручную и подтвердить кандидат «{suggested}» только если он точно соответствует товару."
        return f"Проверить «{name}» вручную: безопасного автоматического значения нет."

    def _stringify_suggestion(self, value: Any) -> str | None:
        if value in (None, "", [], {}):
            return None
        if isinstance(value, list):
            return (
                ", ".join(str(item).strip() for item in value if str(item).strip())
                or None
            )
        return str(value).strip() or None

    def _int(self, value: Any) -> int | None:
        try:
            if value in (None, ""):
                return None
            return int(value)
        except (TypeError, ValueError):
            return None

    def _media_rules(self, card: NormalizedCard) -> list[RuleIssue]:
        issues: list[RuleIssue] = []
        if not card.photos:
            issues.append(
                self._issue(
                    "media_no_images",
                    "media",
                    "critical",
                    "У карточки нет фотографий",
                    "Без фото карточка не готова к нормальной продаже.",
                    "Добавить главное изображение и дополнительные фото товара.",
                    "photos",
                    [],
                )
            )
        elif len(card.photos) < 3:
            issues.append(
                self._issue(
                    "media_too_few_images",
                    "media",
                    "medium",
                    "Мало фотографий товара",
                    "Несколько ракурсов повышают доверие и снижают вопросы.",
                    "Добавить дополнительные фото: общий вид, детали, материал, размер.",
                    "photos",
                    card.photos,
                    {"min_count": 3},
                )
            )
        photo_urls = [self._photo_url(photo) for photo in card.photos]
        duplicates = [
            url for url in set(photo_urls) if url and photo_urls.count(url) > 1
        ]
        if duplicates:
            issues.append(
                self._issue(
                    "media_duplicate_urls",
                    "media",
                    "low",
                    "Повторяются ссылки на фото",
                    "Дубликаты не добавляют покупателю новой информации.",
                    "Заменить дубли на другие ракурсы.",
                    "photos",
                    duplicates,
                )
            )
        invalid = [
            url
            for url in photo_urls
            if url and not url.startswith(("http://", "https://"))
        ]
        if invalid:
            issues.append(
                self._issue(
                    "media_invalid_url",
                    "media",
                    "medium",
                    "Есть некорректные ссылки на изображения",
                    "Невалидные ссылки могут означать проблему синхронизации медиа.",
                    "Пересинхронизировать карточку или проверить источник фото.",
                    "photos",
                    invalid,
                )
            )
        if not card.videos:
            issues.append(
                self._issue(
                    "media_no_video_info",
                    "media",
                    "info",
                    "Видео отсутствует",
                    "Видео может помочь объяснить товар, но не является обязательным.",
                    "Добавить видео, если оно реально помогает показать товар.",
                    "video",
                    [],
                )
            )
        return issues

    def _photo_url(self, photo: Any) -> str:
        if isinstance(photo, dict):
            return str(photo.get("canonical_url") or "").strip()
        return str(photo or "").strip()

    def _identity_rules(self, card: NormalizedCard) -> list[RuleIssue]:
        issues: list[RuleIssue] = []
        if not card.brand:
            issues.append(
                self._issue(
                    "brand_missing",
                    "identity",
                    "low",
                    "Бренд не указан",
                    "Бренд помогает идентифицировать товар.",
                    "Заполнить бренд, если он доступен в источнике.",
                    "brand",
                    card.brand,
                )
            )
        if not card.subject_name:
            issues.append(
                self._issue(
                    "subject_missing",
                    "identity",
                    "medium",
                    "Предмет товара не указан",
                    "Предмет нужен для категории, фильтров и правил карточки.",
                    "Проверить категорию/предмет товара в WB.",
                    "subject_name",
                    card.subject_name,
                )
            )
        return issues

    def _issue(
        self,
        code: str,
        category: str,
        severity: str,
        title: str,
        explanation: str,
        fix: str,
        field: str | None = None,
        current: Any = None,
        expected: Any = None,
        *,
        alternatives: list[Any] | None = None,
        error_details: list[Any] | None = None,
        score_impact: int | None = None,
    ) -> RuleIssue:
        return RuleIssue(
            issue_code=code,
            category=category,
            severity=severity,
            title=title,
            business_explanation=explanation,
            recommended_fix=fix,
            field_name=field,
            current_value_json=current,
            expected_value_json=expected,
            suggested_value=None
            if isinstance(expected, dict)
            else self._stringify_suggestion(expected),
            alternatives=alternatives or [],
            error_details=error_details or [],
            confidence=1.0,
            score_impact=score_impact
            if score_impact is not None
            else SEVERITY_WEIGHTS.get(severity, 0),
        )

    def _counts(self, issues: list[RuleIssue]) -> dict[str, int]:
        return {
            "critical_count": sum(
                1 for issue in issues if issue.severity == "critical"
            ),
            "high_count": sum(1 for issue in issues if issue.severity == "high"),
            "medium_count": sum(1 for issue in issues if issue.severity == "medium"),
            "low_count": sum(1 for issue in issues if issue.severity == "low"),
            "info_count": sum(1 for issue in issues if issue.severity == "info"),
        }


class CardQualityAnalysisService:
    def __init__(self) -> None:
        self.normalizer = CardQualityNormalizationService()
        self.rules = CardQualityRuleEngine()
        self.guided_fixes = GuidedFixMapper()
        self.ai_fixer = CheckerAIFixer()
        self.vision = VisionService()
        self.accounts = AccountService()
        self.product_cards_sync = DomainSyncBase()
        self.product_cards_sync.domain = "product_cards"
        self.product_cards_sync.category = WBAPICategory.CONTENT.value
        self.settings = get_settings()
        self._product_cards_cache: TTLMemoryCache[CardQualityProductsPage] = (
            TTLMemoryCache(default_ttl_seconds=120)
        )
        self._queue_progress_cache: TTLMemoryCache[CardQualityQueueProgress] = (
            TTLMemoryCache(default_ttl_seconds=60)
        )

    def clear_runtime_caches(self) -> None:
        self._product_cards_cache.clear()
        self._queue_progress_cache.clear()

    async def analyze_product(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        nm_id: int,
        force: bool = False,
        requested_by_user_id: int | None = None,
        run: CardQualityAnalysisRun | None = None,
    ) -> PortalProductQualityRead:
        normalized = await self.normalizer.normalize_product(
            session, account_id=account_id, nm_id=nm_id
        )
        if normalized is None:
            return PortalProductQualityRead(
                status="empty",
                module="card_quality",
                source="card_quality",
                mode="local",
                nm_id=nm_id,
                message="product card is not present in finance product cards",
            )
        existing = await self.latest_snapshot(
            session, account_id=account_id, nm_id=nm_id
        )
        if (
            existing is not None
            and existing.source_revision == normalized.source_revision
            and not force
        ):
            if run is not None:
                run.cards_processed += 1
                run.cards_skipped_unchanged += 1
                run.last_processed_key = str(normalized.nm_id)
                run.cursor_json = {"last_nm_id": int(normalized.nm_id)}
                run.heartbeat_at = utcnow()
            return await self.product_quality(
                session, account_id=account_id, nm_id=nm_id
            )
        own_run = run is None
        if run is None:
            run = CardQualityAnalysisRun(
                account_id=account_id,
                run_type="single_product",
                status="running",
                requested_by_user_id=requested_by_user_id,
                started_at=utcnow(),
                cards_total=1,
            )
            session.add(run)
            await session.flush()
        started = time.perf_counter()
        try:
            rule_issues, summary = self.rules.analyze(normalized)
            fixed_fields = await self._load_fixed_field_values(
                session, normalized=normalized
            )
            rule_issues = self._apply_fixed_file_priority(
                normalized, rule_issues, fixed_fields
            )
            normalized = await self._attach_product_dna(normalized)
            rule_issues = await self._apply_ai_audit(
                normalized, rule_issues, fixed_fields
            )
            rule_issues = await self._apply_ai_fixes(normalized, rule_issues)
            rule_issues = await self._apply_text_ai_fixes(normalized, rule_issues)
            rule_issues = self._finalize_rule_issues(
                normalized, rule_issues, fixed_fields
            )
            score, category_scores = self._score_with_super_validator(
                normalized, rule_issues, summary
            )
            counts = self.rules._counts(rule_issues)
            summary = {
                **summary,
                "score": score,
                "status": self.rules.status(score, rule_issues),
                "category_scores": category_scores,
                "issues_count": len(rule_issues),
                **counts,
                "fixed_file_fields": sorted(fixed_fields),
            }
            snapshot = CardQualitySnapshot(
                account_id=account_id,
                nm_id=nm_id,
                source_card_id=normalized.source_card_id,
                source_revision=normalized.source_revision,
                title=normalized.title,
                description=normalized.description,
                brand=normalized.brand,
                subject_name=normalized.subject_name,
                vendor_code=normalized.vendor_code,
                characteristics_json=normalized.characteristics,
                media_json={
                    "photos": normalized.photos,
                    "videos": normalized.videos,
                    "sizes": normalized.sizes,
                },
                photos_count=len(normalized.photos),
                video_count=len(normalized.videos),
                source_updated_at=normalized.source_updated_at,
                analyzed_at=utcnow(),
                score=int(summary["score"]),
                status=str(summary["status"]),
                summary_json={
                    **summary,
                    "product_dna": normalized.product_dna_json or {},
                    "product_dna_audit": normalized.product_dna_audit or {},
                    "product_dna_enabled": bool(self.vision.is_enabled),
                },
            )
            try:
                async with session.begin_nested():
                    session.add(snapshot)
                    await session.flush()
            except IntegrityError:
                existing_snapshot = (
                    await session.execute(
                        select(CardQualitySnapshot).where(
                            CardQualitySnapshot.account_id == account_id,
                            CardQualitySnapshot.nm_id == nm_id,
                            CardQualitySnapshot.source_revision
                            == normalized.source_revision,
                        )
                    )
                ).scalar_one_or_none()
                if not force or existing_snapshot is None:
                    if own_run:
                        return await self.product_quality(
                            session, account_id=account_id, nm_id=nm_id
                        )
                    run.cards_processed += 1
                    run.cards_skipped_unchanged += 1
                    run.last_processed_key = str(normalized.nm_id)
                    run.cursor_json = {"last_nm_id": int(normalized.nm_id)}
                    run.heartbeat_at = utcnow()
                    return await self.product_quality(
                        session, account_id=account_id, nm_id=nm_id
                    )
                try:
                    session.expunge(snapshot)
                except Exception:
                    pass
                existing_snapshot.source_card_id = normalized.source_card_id
                existing_snapshot.title = normalized.title
                existing_snapshot.description = normalized.description
                existing_snapshot.brand = normalized.brand
                existing_snapshot.subject_name = normalized.subject_name
                existing_snapshot.vendor_code = normalized.vendor_code
                existing_snapshot.characteristics_json = normalized.characteristics
                existing_snapshot.media_json = {
                    "photos": normalized.photos,
                    "videos": normalized.videos,
                    "sizes": normalized.sizes,
                }
                existing_snapshot.photos_count = len(normalized.photos)
                existing_snapshot.video_count = len(normalized.videos)
                existing_snapshot.source_updated_at = normalized.source_updated_at
                existing_snapshot.analyzed_at = utcnow()
                existing_snapshot.score = int(summary["score"])
                existing_snapshot.status = str(summary["status"])
                existing_snapshot.summary_json = {
                    **summary,
                    "product_dna": normalized.product_dna_json or {},
                    "product_dna_audit": normalized.product_dna_audit or {},
                    "product_dna_enabled": bool(self.vision.is_enabled),
                    "force_reanalyzed_existing_revision": True,
                }
                snapshot = existing_snapshot
            created, resolved = await self._sync_issues(
                session, snapshot=snapshot, normalized=normalized, issues=rule_issues
            )
            run.cards_processed += 1
            run.cards_analyzed += 1
            run.cards_clean += (
                1
                if not [issue for issue in rule_issues if issue.severity != "info"]
                else 0
            )
            run.cards_with_issues += (
                1 if [issue for issue in rule_issues if issue.severity != "info"] else 0
            )
            run.issues_created += created
            run.issues_resolved += resolved
            run.source_revision = normalized.source_revision
            run.last_processed_key = str(normalized.nm_id)
            run.cursor_json = {"last_nm_id": int(normalized.nm_id)}
            run.heartbeat_at = utcnow()
            if own_run:
                run.status = "completed"
                run.finished_at = utcnow()
            await self._upsert_registry(session, account_id=account_id)
            if own_run:
                await self._record_module_run(session, run=run)
            if own_run:
                await session.commit()
            self.clear_runtime_caches()
            response = await self.product_quality(
                session, account_id=account_id, nm_id=nm_id
            )
            response.raw["analysis_duration_ms"] = int(
                (time.perf_counter() - started) * 1000
            )
            return response
        except Exception as exc:
            if own_run:
                await self._record_failed_single_product_run(
                    session,
                    run=run,
                    account_id=account_id,
                    requested_by_user_id=requested_by_user_id,
                    error_summary=exc.__class__.__name__,
                )
            else:
                run.cards_failed += 1
                run.error_summary = exc.__class__.__name__
            raise

    async def _record_failed_single_product_run(
        self,
        session: AsyncSession,
        *,
        run: CardQualityAnalysisRun,
        account_id: int,
        requested_by_user_id: int | None,
        error_summary: str,
    ) -> None:
        run_id = getattr(run, "id", None)
        started_at = getattr(run, "started_at", None) or utcnow()
        try:
            await session.rollback()
        except Exception:
            pass
        try:
            failed_run = (
                await session.get(CardQualityAnalysisRun, run_id)
                if run_id is not None
                else None
            )
            if failed_run is None:
                failed_run = CardQualityAnalysisRun(
                    account_id=account_id,
                    run_type="single_product",
                    requested_by_user_id=requested_by_user_id,
                    started_at=started_at,
                    cards_total=1,
                )
                session.add(failed_run)
                await session.flush()
            failed_run.cards_failed = (
                int(getattr(failed_run, "cards_failed", 0) or 0) + 1
            )
            failed_run.error_summary = error_summary
            failed_run.status = "failed"
            failed_run.finished_at = utcnow()
            await self._record_module_run(session, run=failed_run)
            await session.commit()
        except Exception:
            try:
                await session.rollback()
            except Exception:
                pass

    async def _attach_product_dna(self, normalized: NormalizedCard) -> NormalizedCard:
        if not self.vision.is_enabled:
            return replace(
                normalized,
                product_dna_json={},
                product_dna_audit={
                    "status": "disabled",
                    "trust_state": "disabled",
                    "reason": "vision_service_disabled",
                },
                product_dna_text="",
            )
        photo_urls = [
            self.rules._photo_url(item)
            for item in normalized.photos
            if self.rules._photo_url(item)
        ]
        if not photo_urls:
            return replace(
                normalized,
                product_dna_json={},
                product_dna_audit=build_product_dna_audit(
                    {}, subject_name=normalized.subject_name, photo_count=0
                ),
                product_dna_text="",
            )
        try:
            result = await self.vision.generate_product_dna_result(
                photo_urls[0],
                normalized.subject_name,
                photo_urls=photo_urls,
            )
        except Exception as exc:
            return replace(
                normalized,
                product_dna_json={},
                product_dna_audit={
                    **build_product_dna_audit(
                        {},
                        subject_name=normalized.subject_name,
                        photo_count=len(photo_urls),
                    ),
                    "error": exc.__class__.__name__,
                },
                product_dna_text="",
            )
        dna = (
            result.get("dna")
            if isinstance(result, dict) and isinstance(result.get("dna"), dict)
            else {}
        )
        audit = (
            result.get("audit")
            if isinstance(result, dict) and isinstance(result.get("audit"), dict)
            else build_product_dna_audit(
                dna, subject_name=normalized.subject_name, photo_count=len(photo_urls)
            )
        )
        return replace(
            normalized,
            product_dna_json=dna,
            product_dna_audit=audit,
            product_dna_text=product_dna_to_text(dna)
            if is_grounded_product_dna_audit(audit)
            else "",
        )

    async def _load_fixed_field_values(
        self, session: AsyncSession, *, normalized: NormalizedCard
    ) -> dict[str, str]:
        result = await session.execute(
            select(CardQualityFixedFileEntry).where(
                CardQualityFixedFileEntry.account_id == normalized.account_id,
                CardQualityFixedFileEntry.nm_id == normalized.nm_id,
            )
        )
        entries = list(result.scalars().all())
        return {
            str(entry.char_name).strip(): str(entry.fixed_value).strip()
            for entry in entries
            if str(entry.char_name or "").strip()
            and str(entry.fixed_value or "").strip()
        }

    async def _apply_ai_audit(
        self,
        normalized: NormalizedCard,
        issues: list[RuleIssue],
        fixed_fields: dict[str, str],
    ) -> list[RuleIssue]:
        if not self.ai_fixer.is_enabled:
            return issues
        payload = self.rules._wb_card_payload(normalized)
        payload["characteristics"] = [
            item
            for item in (payload.get("characteristics") or [])
            if not (
                isinstance(item, dict)
                and (
                    is_no_touch_characteristic(item.get("name"), card=payload)
                    or is_fixed_file_only_characteristic(item.get("name"))
                )
            )
        ]
        payload["_fixed_file_chars"] = list(fixed_fields)
        valid_char_names, valid_char_ids = self._valid_category_characteristics(
            normalized
        )
        if valid_char_names:
            payload["_valid_char_names"] = valid_char_names
        try:
            ai_issues = await self.ai_fixer.audit_card(
                card=payload, product_dna=normalized.product_dna_text or ""
            )
        except Exception:
            return issues
        if not ai_issues:
            return issues

        fixed_keys = {_norm_text(name) for name in fixed_fields}
        existing_keys = {
            (
                issue.issue_code,
                _field_key(issue.field_name),
                self._hashable_issue_value(issue.current_value_json),
            )
            for issue in issues
        }
        out = list(issues)
        for item in ai_issues:
            if not isinstance(item, dict):
                continue
            if self._skip_ai_audit_issue(item, payload):
                continue
            field_path = self._normalize_ai_issue_field_path(item.get("name"))
            if not field_path:
                continue
            field_key = _field_key(field_path)
            if field_key in fixed_keys:
                continue
            category = self._map_ai_issue_category(item.get("category"), field_path)
            current_value = self._current_value_for_field(
                normalized, field_path, fallback=item.get("value")
            )
            severity = self._map_ai_issue_severity(item.get("severity"))
            ai_error_details = (
                item.get("errors") if isinstance(item.get("errors"), list) else []
            )
            fix_action = str(item.get("fix_action") or "replace").strip().lower()
            ai_errors = (
                item.get("errors") if isinstance(item.get("errors"), list) else []
            )
            ai_error_types = {
                str(error.get("type") or "").strip().lower()
                for error in ai_errors
                if isinstance(error, dict)
            }
            is_category_mismatch = bool(
                ai_error_types.intersection({"category_mismatch", "wrong_category"})
                or "не входит в допустимый" in str(item.get("message") or "").lower()
            )
            ai_name_key = _norm_text(item.get("name"))
            ai_charc_id = self.rules._int(item.get("charcId"))
            if is_category_mismatch and fix_action == "clear":
                if (
                    ai_name_key
                    and ai_name_key in {_norm_text(name) for name in valid_char_names}
                ) or (ai_charc_id is not None and ai_charc_id in valid_char_ids):
                    continue
            if fix_action in {"swap", "clear", "compound"}:
                marker: dict[str, Any] = {"type": fix_action, "fix_action": fix_action}
                if fix_action == "swap":
                    marker["swap_to_name"] = item.get("swap_to_name")
                    marker["swap_to_value"] = item.get("swap_to_value")
                if fix_action == "compound":
                    marker["fixes"] = (
                        item.get("compound_fixes")
                        if isinstance(item.get("compound_fixes"), list)
                        else []
                    )
                ai_error_details = [*ai_error_details, marker]
            allowed_values = self._allowed_values_for_ai_issue(field_path, category)
            compound_candidates = self._extract_compound_candidate_values(
                ai_error_details, field_path
            )
            expected_value_json: dict[str, Any] = {"ai_errors": ai_error_details}
            if compound_candidates:
                expected_value_json["candidate_values"] = compound_candidates
            if allowed_values:
                expected_value_json["allowed_values"] = allowed_values[:50]
            code = (
                f"ai_{str(item.get('category') or 'mixed').strip().lower() or 'mixed'}"
            )
            key = (code, field_key, self._hashable_issue_value(current_value))
            if key in existing_keys:
                continue
            existing_keys.add(key)
            message = str(item.get("message") or "AI обнаружил проблему").strip()
            out.append(
                RuleIssue(
                    issue_code=code,
                    category=category,
                    severity=severity,
                    title=message[:255],
                    business_explanation=self._format_ai_audit_explanation(item),
                    recommended_fix=self._ai_audit_recommended_fix(
                        field_path, fix_action
                    ),
                    field_name=field_path,
                    current_value_json=current_value,
                    expected_value_json=expected_value_json,
                    confidence=0.7,
                    charc_id=self.rules._int(item.get("charcId")),
                    alternatives=compound_candidates,
                    allowed_values=allowed_values,
                    error_details=ai_error_details,
                    ai_alternatives=compound_candidates,
                    requires_human_check=category in {"title", "description"}
                    or fix_action in {"clear", "swap", "compound"},
                    ai_reason=message,
                    ai_reason_short=message[:240],
                    ai_reason_full=self._format_ai_audit_explanation(item),
                    ai_evidence={"ai_audit": item},
                    ai_used_sources=self._ai_audit_sources(normalized),
                    source="ai",
                    score_impact=SEVERITY_WEIGHTS.get(severity, 0),
                )
            )
        return out

    def _valid_category_characteristics(
        self, normalized: NormalizedCard
    ) -> tuple[list[str], set[int]]:
        if normalized.subject_id is None:
            return [], set()
        try:
            subject_chars = get_catalog().get_subject_chars(int(normalized.subject_id))
        except Exception:
            return [], set()
        names: list[str] = []
        ids: set[int] = set()
        card_payload = self.rules._wb_card_payload(normalized)
        for item in subject_chars or []:
            name = str(getattr(item, "name", "") or "").strip()
            if (
                name
                and not is_no_touch_characteristic(name, card=card_payload)
                and not is_fixed_file_only_characteristic(name)
            ):
                names.append(name)
            charc_id = getattr(item, "charc_id", None)
            try:
                if charc_id is not None:
                    ids.add(int(charc_id))
            except (TypeError, ValueError):
                pass
        return names, ids

    def _skip_ai_audit_issue(
        self, item: dict[str, Any], card_payload: dict[str, Any] | None = None
    ) -> bool:
        text = json.dumps(item, ensure_ascii=False).lower()
        name = str(item.get("name") or "").strip().lower()
        if not name or name in {"цвет", "color", "основной цвет", "цвет товара"}:
            return True
        if is_no_touch_characteristic(name, card=card_payload):
            return True
        if is_fixed_file_only_characteristic(name):
            return True
        if any(
            marker in text
            for marker in ("vendorcode", "vendor code", "артикул продавца", "артикул")
        ):
            return True
        if self._is_ai_audit_date_sensitive(item):
            return True
        if any(
            marker in text
            for marker in (
                "allowed_values",
                "допустим",
                "справочник wb",
                "список допустим",
            )
        ):
            return True
        return False

    def _is_ai_audit_date_sensitive(self, item: dict[str, Any]) -> bool:
        values: list[Any] = [item.get("name"), item.get("message"), item.get("value")]
        values.extend(item.get("errors") or [])
        return any(_contains_date_context_text(value) for value in values)

    def _normalize_ai_issue_field_path(self, raw_name: Any) -> str:
        name = str(raw_name or "").strip()
        if not name:
            return ""
        lowered = name.lower()
        if any(sep in lowered for sep in ("/", "|")) or "," in name:
            return ""
        if lowered in {"title", "название"}:
            return "title"
        if lowered in {"description", "описание"}:
            return "description"
        if lowered.startswith("characteristics."):
            return f"characteristics.{name.split('.', 1)[1].strip()}"
        return f"characteristics.{name}"

    def _map_ai_issue_category(self, raw_category: Any, field_path: str) -> str:
        field_key = _field_key(field_path)
        if field_key == "title":
            return "title"
        if field_key == "description":
            return "description"
        normalized_path = str(field_path or "").strip().lower()
        if normalized_path.startswith("characteristics."):
            return "characteristics"
        if normalized_path.startswith("photos"):
            return "photos"
        if normalized_path.startswith("videos"):
            return "video"
        raw = str(raw_category or "").strip().lower()
        if raw == "text":
            return "description"
        if raw == "identification":
            return "identity"
        if raw == "qualification":
            return "characteristics"
        return "characteristics"

    def _map_ai_issue_severity(self, raw: Any) -> str:
        value = str(raw or "").strip().lower()
        if value == "critical":
            return "critical"
        if value in {"error", "high"}:
            return "high"
        if value in {"warning", "medium"}:
            return "medium"
        if value in {"info", "low"}:
            return value
        return "medium"

    def _current_value_for_field(
        self, normalized: NormalizedCard, field_path: str, *, fallback: Any = None
    ) -> Any:
        field_key = _field_key(field_path)
        if field_key == "title":
            return normalized.title
        if field_key == "description":
            return normalized.description
        for item in normalized.characteristics:
            if not isinstance(item, dict):
                continue
            if _norm_text(item.get("name")) == field_key:
                return item.get("value", item.get("values"))
        return fallback

    def _format_ai_audit_explanation(self, item: dict[str, Any]) -> str:
        message = str(item.get("message") or "").strip()
        details = []
        for error in item.get("errors") or []:
            if isinstance(error, dict) and str(error.get("message") or "").strip():
                details.append(str(error.get("message")).strip())
        if details:
            return f"{message} " + " ".join(details)
        return message or "AI audit detected a card inconsistency."

    def _ai_audit_recommended_fix(self, field_path: str, fix_action: str) -> str:
        if fix_action in {"clear", "swap", "compound"}:
            return "Проверить проблему вручную: AI предложил небезопасное структурное изменение."
        if _field_key(field_path) in {"title", "description"}:
            return "Сгенерировать новый текст по подтверждённым фактам карточки и Product DNA."
        return "Подобрать безопасное значение через AI fix и проверить его backend-валидаторами."

    def _ai_audit_sources(self, normalized: NormalizedCard) -> list[str]:
        sources = ["card_characteristics"]
        if normalized.product_dna_text:
            sources.append("product_dna")
        elif normalized.photos:
            sources.append("photos")
        return sources

    def _apply_fixed_file_priority(
        self,
        normalized: NormalizedCard,
        issues: list[RuleIssue],
        fixed_fields: dict[str, str],
    ) -> list[RuleIssue]:
        if not fixed_fields:
            return issues
        card_payload = self.rules._wb_card_payload(normalized)
        normalized_fixed = {
            _norm_text(name): str(value).strip()
            for name, value in fixed_fields.items()
            if str(value).strip()
            and not is_no_touch_characteristic(name, card=card_payload)
        }
        if not normalized_fixed:
            return issues

        current_by_name = {
            _norm_text(item.get("name")): item
            for item in normalized.characteristics
            if isinstance(item, dict) and _norm_text(item.get("name"))
        }
        out: list[RuleIssue] = []
        controlled_fields = set(normalized_fixed)
        for issue in issues:
            field_key = _field_key(issue.field_name)
            if field_key in controlled_fields and issue.source in {
                "ai",
                "code",
                "wb_catalog",
                "auto_fix",
            }:
                continue
            out.append(issue)

        for name_key, fixed_value in normalized_fixed.items():
            current = current_by_name.get(name_key, {})
            current_value = current.get("value") if isinstance(current, dict) else None
            if _issue_values_equivalent(current_value, fixed_value):
                continue
            display_name = str(
                (current or {}).get("name")
                or next(
                    (name for name in fixed_fields if _norm_text(name) == name_key),
                    name_key,
                )
            ).strip()
            out.append(
                RuleIssue(
                    issue_code="fixed_file_mismatch",
                    category="characteristics",
                    severity="medium",
                    title=f"Значение «{display_name}» отличается от fixed-file",
                    business_explanation="Fixed-file является приоритетным источником для этого поля.",
                    recommended_fix=f"Установить «{display_name}» = «{fixed_value}» из fixed-file.",
                    field_name=f"characteristics.{display_name}",
                    current_value_json=current_value,
                    expected_value_json={"fixed_file_value": fixed_value},
                    confidence=1.0,
                    suggested_value=fixed_value,
                    ai_suggested_value=fixed_value,
                    requires_human_check=False,
                    ai_reason="fixed-file priority",
                    ai_reason_short="fixed-file priority",
                    ai_reason_full="Fixed-file value overrides AI and catalog suggestions for this field.",
                    ai_evidence={
                        "constraint": "fixed_file",
                        "observed": [f"{display_name}: {fixed_value}"],
                    },
                    ai_used_sources=["fixed_file"],
                    source="fixed_file",
                    score_impact=SEVERITY_WEIGHTS["medium"],
                )
            )
        return out

    def _hashable_issue_value(self, value: Any) -> str:
        stable = self.rules._stable_value_context(value)
        return json.dumps(stable, ensure_ascii=False, sort_keys=True, default=str)

    async def _apply_ai_fixes(
        self, normalized: NormalizedCard, issues: list[RuleIssue]
    ) -> list[RuleIssue]:
        if not self.ai_fixer.is_enabled:
            return issues
        updated = list(issues)
        ai_candidates: list[dict[str, Any]] = []
        index_by_id: dict[str, int] = {}
        card_payload = self.rules._wb_card_payload(normalized)
        for index, issue in enumerate(updated):
            if issue.source == "fixed_file" or self._is_date_issue(issue):
                continue
            if is_no_touch_characteristic(
                _field_key(issue.field_name) or issue.field_name or issue.title,
                card=card_payload,
            ):
                continue
            if is_fixed_file_only_characteristic(
                _field_key(issue.field_name) or issue.field_name or issue.title
            ):
                continue
            if issue.category in {"title", "description"} or _field_key(
                issue.field_name
            ) in {"title", "description"}:
                continue
            if not issue.allowed_values and issue.source != "ai":
                continue
            issue_id = str(index)
            index_by_id[issue_id] = index
            ai_candidates.append(
                {
                    "id": issue_id,
                    "code": issue.issue_code,
                    "name": issue.field_name or issue.title,
                    "title": issue.title,
                    "current_value": issue.current_value_json,
                    "description": issue.business_explanation,
                    "allowed_values": issue.allowed_values or [],
                    "errors": issue.error_details or [],
                }
            )
        if not ai_candidates:
            return updated

        pending = ai_candidates
        for attempt in range(MAX_FIX_RETRIES + 1):
            try:
                fixes = await self.ai_fixer.generate_fixes(
                    card=card_payload,
                    issues=pending,
                    product_dna=normalized.product_dna_text or "",
                )
            except Exception:
                return updated
            retry_candidates: list[dict[str, Any]] = []
            for issue_id, fix in fixes.items():
                index = index_by_id.get(str(issue_id))
                if index is None:
                    continue
                issue = updated[index]
                candidate, invalid_reason = self._validated_ai_candidate(
                    normalized, issue, fix
                )
                if (
                    invalid_reason
                    and attempt < MAX_FIX_RETRIES
                    and not bool(fix.get("requires_human_check"))
                ):
                    refix_value = getattr(self.ai_fixer, "refix_value", None)
                    if callable(refix_value):
                        try:
                            refix = await refix_value(
                                card=card_payload,
                                char_name=issue.field_name or issue.title,
                                current_value=fix.get("recommended_value")
                                or issue.current_value_json,
                                failed_reason=invalid_reason,
                                allowed_values=issue.allowed_values or [],
                                limits=self._effective_limits(
                                    issue.error_details or []
                                ),
                                product_dna=normalized.product_dna_text or "",
                            )
                        except Exception:
                            refix = {}
                        if isinstance(refix, dict) and refix:
                            retry_candidate, retry_invalid_reason = (
                                self._validated_ai_candidate(normalized, issue, refix)
                            )
                            fix = refix
                            candidate = retry_candidate
                            invalid_reason = retry_invalid_reason
                            if retry_invalid_reason:
                                updated[index] = self._merge_ai_fix(
                                    normalized=normalized,
                                    issue=issue,
                                    fix=fix,
                                    recommended=candidate,
                                    invalid_reason=retry_invalid_reason,
                                )
                                continue
                        else:
                            updated[index] = self._merge_ai_fix(
                                normalized=normalized,
                                issue=issue,
                                fix=fix,
                                recommended=candidate,
                                invalid_reason=invalid_reason,
                            )
                            continue
                    else:
                        retry_entry = dict(
                            next(
                                (
                                    item
                                    for item in pending
                                    if str(item.get("id")) == str(issue_id)
                                ),
                                {},
                            )
                        )
                        retry_entry["validation_error"] = invalid_reason
                        retry_candidates.append(retry_entry)
                        continue
                updated[index] = self._merge_ai_fix(
                    normalized=normalized,
                    issue=issue,
                    fix=fix,
                    recommended=candidate,
                    invalid_reason=invalid_reason,
                )
            if not retry_candidates:
                break
            pending = retry_candidates
        return updated

    async def _apply_text_ai_fixes(
        self, normalized: NormalizedCard, issues: list[RuleIssue]
    ) -> list[RuleIssue]:
        if not self.ai_fixer.is_enabled:
            return issues
        updated = list(issues)
        has_characteristic_change = any(
            issue.category == "characteristics"
            and issue.source != "fixed_file"
            and not self._is_date_issue(issue)
            for issue in updated
        )
        has_description_issue = any(
            issue.category == "description"
            or _field_key(issue.field_name) == "description"
            for issue in updated
        )
        if has_characteristic_change and not has_description_issue:
            updated.append(
                RuleIssue(
                    issue_code="description_refresh_needed",
                    category="description",
                    severity="medium",
                    title="Нужно обновить описание после исправления характеристик",
                    business_explanation=(
                        "После изменения характеристик описание может стать неактуальным. "
                        "Source Checker генерирует описание уже после characteristic fixes."
                    ),
                    recommended_fix="Сгенерировать новое описание по подтверждённым характеристикам и Product DNA.",
                    field_name="description",
                    current_value_json=normalized.description,
                    expected_value_json={"reason": "characteristics_changed"},
                    confidence=0.85,
                    requires_human_check=False,
                    source="ai",
                    score_impact=SEVERITY_WEIGHTS["medium"],
                )
            )

        text_indices = [
            index
            for index, issue in enumerate(updated)
            if issue.category in {"title", "description"}
            or _field_key(issue.field_name) in {"title", "description"}
        ]
        if not text_indices:
            return updated

        context = self._card_context_with_confirmed_fixes(normalized, updated)
        for index in text_indices:
            issue = updated[index]
            if issue.source == "fixed_file" or self._is_date_issue(issue):
                continue
            is_title = (
                issue.category == "title" or _field_key(issue.field_name) == "title"
            )
            if is_title:
                updated[index] = await self._generate_title_fix(
                    normalized, issue, context
                )
                if (
                    updated[index].suggested_value
                    and not updated[index].requires_human_check
                ):
                    context["title"] = updated[index].suggested_value
                continue
            updated[index] = await self._generate_description_fix(
                normalized, issue, context
            )
            if (
                updated[index].suggested_value
                and not updated[index].requires_human_check
            ):
                context["description"] = updated[index].suggested_value
        return updated

    async def _generate_title_fix(
        self,
        normalized: NormalizedCard,
        issue: RuleIssue,
        context: dict[str, Any],
    ) -> RuleIssue:
        try:
            fix = await self.ai_fixer.generate_title(
                card=context,
                product_dna=normalized.product_dna_text or "",
                seo_keywords=self._subject_keywords(context),
            )
        except Exception:
            return issue
        candidate = (
            self.rules._stringify_suggestion(fix.get("recommended_value"))
            if isinstance(fix, dict)
            else None
        )
        last_candidate = candidate
        invalid_reason = "AI не вернул предложенный title" if not candidate else ""
        if candidate:
            invalid_reason = self._validate_title_fix_candidate(
                normalized, candidate, context
            )
        retry_count = 0
        while invalid_reason and candidate and retry_count < MAX_FIX_RETRIES:
            retry_count += 1
            refix_method = getattr(self.ai_fixer, "refix_title", None)
            if not callable(refix_method):
                break
            try:
                refix = await refix_method(
                    card=context, current_title=candidate, failed_reason=invalid_reason
                )
            except Exception:
                break
            if not isinstance(refix, dict) or not refix:
                break
            if refix.get("recommended_value"):
                candidate = self.rules._stringify_suggestion(
                    refix.get("recommended_value")
                )
                last_candidate = candidate or last_candidate
            fix = {**fix, **refix}
            invalid_reason = (
                self._validate_title_fix_candidate(normalized, candidate or "", context)
                if candidate
                else "AI не вернул предложенный title"
            )

        if invalid_reason:
            draft = (
                last_candidate
                if last_candidate
                and not _issue_values_equivalent(normalized.title, last_candidate)
                else None
            )
            if draft:
                fix = {
                    **fix,
                    "recommended_value": draft,
                    "requires_human_check": True,
                    "suggestion_kind": "draft_text",
                    "reason": str(fix.get("reason") or invalid_reason),
                }
            return self._merge_ai_fix(
                normalized=normalized,
                issue=issue,
                fix=fix,
                recommended=None,
                invalid_reason=invalid_reason,
            )
        return self._merge_ai_fix(
            normalized=normalized,
            issue=issue,
            fix=fix,
            recommended=candidate,
            invalid_reason=None,
        )

    async def _generate_description_fix(
        self,
        normalized: NormalizedCard,
        issue: RuleIssue,
        context: dict[str, Any],
    ) -> RuleIssue:
        try:
            fix = await self.ai_fixer.generate_description(
                card=context,
                product_dna=normalized.product_dna_text or "",
                seo_keywords=self._subject_keywords(context),
            )
        except Exception:
            return issue
        raw_candidate = (
            self.rules._stringify_suggestion(fix.get("recommended_value"))
            if isinstance(fix, dict)
            else None
        )
        candidate = (
            normalize_generated_description(raw_candidate or "")
            if raw_candidate
            else None
        )
        last_candidate = candidate
        invalid_reason = "AI не вернул предложенное описание" if not candidate else ""
        if candidate:
            invalid_reason = self._validate_description_fix_candidate(
                normalized, candidate, context
            )
        retry_count = 0
        while invalid_reason and candidate and retry_count < MAX_FIX_RETRIES:
            retry_count += 1
            refix_method = getattr(self.ai_fixer, "refix_description", None)
            if not callable(refix_method):
                break
            try:
                refix = await refix_method(
                    card=context,
                    current_description=candidate,
                    failed_reason=invalid_reason,
                )
            except Exception:
                break
            if not isinstance(refix, dict) or not refix:
                break
            if refix.get("recommended_value"):
                candidate = normalize_generated_description(
                    self.rules._stringify_suggestion(refix.get("recommended_value"))
                    or ""
                )
                last_candidate = candidate or last_candidate
            fix = {**fix, **refix, "recommended_value": candidate}
            invalid_reason = (
                self._validate_description_fix_candidate(
                    normalized, candidate or "", context
                )
                if candidate
                else "AI не вернул предложенное описание"
            )

        if invalid_reason:
            draft = (
                last_candidate
                if last_candidate
                and len(last_candidate) >= 500
                and not _issue_values_equivalent(normalized.description, last_candidate)
                else None
            )
            if draft:
                fix = {
                    **fix,
                    "recommended_value": draft,
                    "requires_human_check": True,
                    "suggestion_kind": "draft_text",
                    "reason": str(fix.get("reason") or invalid_reason),
                }
            return self._merge_ai_fix(
                normalized=normalized,
                issue=issue,
                fix=fix,
                recommended=None,
                invalid_reason=invalid_reason,
            )
        return self._merge_ai_fix(
            normalized=normalized,
            issue=issue,
            fix=fix,
            recommended=candidate,
            invalid_reason=None,
        )

    def _subject_keywords(self, card: dict[str, Any]) -> list[str]:
        subject = (
            card.get("subjectName")
            or card.get("subject_name")
            or card.get("object")
            or ""
        )
        if not subject:
            return []
        try:
            return get_catalog().get_keywords_for_subject(str(subject))
        except Exception:
            return []

    def _card_context_with_confirmed_fixes(
        self, normalized: NormalizedCard, issues: list[RuleIssue]
    ) -> dict[str, Any]:
        context = copy.deepcopy(self.rules._wb_card_payload(normalized))
        chars = context.get("characteristics")
        if not isinstance(chars, list):
            chars = []
            context["characteristics"] = chars
        for issue in issues:
            if issue.requires_human_check or not issue.suggested_value:
                continue
            field_key = _field_key(issue.field_name)
            if field_key == "title":
                context["title"] = issue.suggested_value
            elif field_key == "description":
                context["description"] = issue.suggested_value
            elif str(issue.field_name or "").startswith("characteristics."):
                _set_characteristic_value(
                    chars, issue.field_name or "", issue.suggested_value, issue.charc_id
                )
        return context

    def _validate_title_fix_candidate(
        self, normalized: NormalizedCard, candidate: str, context: dict[str, Any]
    ) -> str:
        candidate = str(candidate or "").strip()
        valid, reason = validate_title(candidate, context, strict_content=True)
        if not valid:
            return reason
        keep_current, info = should_keep_current_title_as_safer(
            normalized.title, candidate, context
        )
        if keep_current:
            return f"title_business_guard:{info.get('reason') or 'unsafe_candidate'}"
        return ""

    def _validate_description_fix_candidate(
        self, normalized: NormalizedCard, candidate: str, context: dict[str, Any]
    ) -> str:
        candidate = normalize_generated_description(candidate)
        valid, reason = validate_description(candidate, context, strict_structure=True)
        if not valid:
            return reason
        facts_valid, facts_reason = validate_description_facts(
            candidate,
            context,
            allow_visual_facts=is_grounded_product_dna_audit(
                normalized.product_dna_audit
            ),
        )
        if not facts_valid:
            return f"description_factual_guard:{facts_reason}"
        return ""

    def _score_with_super_validator(
        self,
        normalized: NormalizedCard,
        issues: list[RuleIssue],
        summary: dict[str, Any],
    ) -> tuple[int, dict[str, int]]:
        fallback_score, fallback_categories = self.rules.score(issues)
        try:
            raw_data = self.rules._wb_card_payload(normalized)
            adapter_issues = [
                SimpleNamespace(
                    code=issue.issue_code,
                    issue_code=issue.issue_code,
                    category=issue.category,
                    severity=issue.severity,
                    status="new",
                    title=issue.title,
                    description=issue.business_explanation,
                    field_path=issue.field_name,
                    requires_human_check=issue.requires_human_check,
                    suggested_value=issue.suggested_value,
                    ai_suggested_value=issue.ai_suggested_value,
                    score_impact=issue.score_impact,
                )
                for issue in issues
            ]
            card_adapter = SimpleNamespace(
                title=normalized.title,
                description=normalized.description,
                brand=normalized.brand,
                subject_name=normalized.subject_name,
                category_name=normalized.subject_name,
                characteristics=normalized.characteristics,
                product_dna_audit=normalized.product_dna_audit or {},
            )
            sv = super_validator_service.evaluate(
                card=card_adapter,
                raw_data=raw_data,
                issues=adapter_issues,
                base_breakdown={"total_score": summary.get("score", fallback_score)},
                fcs=summary.get("fcs")
                if isinstance(summary.get("fcs"), dict)
                else calculate_card_fcs(raw_data),
            )
        except Exception:
            return fallback_score, fallback_categories

        score = int(
            sv.get("final_score", sv.get("total_score", fallback_score))
            or fallback_score
        )
        categories = dict(fallback_categories)
        categories["title"] = int(
            max(0, min(100, float(sv.get("title_score", 20)) * 5))
        )
        categories["description"] = int(
            max(0, min(100, float(sv.get("description_score", 20)) * 5))
        )
        categories["characteristics"] = int(
            max(0, min(100, float(sv.get("characteristics_score", 20)) * 5))
        )
        categories["consistency"] = int(
            max(0, min(100, float(sv.get("consistency_score", 100))))
        )
        categories["text"] = int(max(0, min(100, float(sv.get("text_score", 100)))))
        categories["attributes"] = int(
            max(0, min(100, float(sv.get("attributes_score", 100))))
        )
        categories["ready_score"] = int(
            max(0, min(100, float(sv.get("ready_score", score))))
        )
        categories["potential_score"] = int(
            max(0, min(100, float(sv.get("potential_score", score))))
        )
        return score, categories

    def issue_belongs_to_bucket(self, issue: CardQualityIssue, bucket: str) -> bool:
        bucket_key = str(bucket or "").strip().lower()
        status = str(getattr(issue, "status", "") or "").strip().lower()
        if status in {"done", "fixed", "auto_fixed", "ignored", "skipped"}:
            return bucket_key == "all"
        category = str(getattr(issue, "category", "") or "").strip().lower()
        field_key = _field_key(getattr(issue, "field_name", None))
        is_media = category in {"media", "photos", "photo", "video"} or field_key in {
            "photos",
            "photo",
            "video",
            "videos",
        }
        if bucket_key == "all":
            return True
        if bucket_key == "media":
            return is_media
        if bucket_key == "human_check":
            return bool(getattr(issue, "requires_human_check", False)) and not is_media
        if bucket_key == "actionable":
            if is_media:
                return False
            has_value = bool(
                str(
                    getattr(issue, "suggested_value", None)
                    or getattr(issue, "ai_suggested_value", None)
                    or ""
                ).strip()
            )
            return (
                bool(getattr(issue, "requires_human_check", False))
                or has_value
                or category in {"title", "description", "characteristics"}
            )
        return False

    async def _queue_issue_rows(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        bucket: str = "actionable",
        nm_id: int | None = None,
        severity: str | None = None,
        include_done: bool = False,
    ) -> list[CardQualityIssue]:
        statuses = {"new", "in_progress", "postponed", "blocked"}
        if include_done:
            statuses |= {"done", "resolved", "ignored"}
        stmt = select(CardQualityIssue).where(
            CardQualityIssue.account_id == account_id,
            CardQualityIssue.severity != "info",
            CardQualityIssue.status.in_(tuple(statuses)),
        )
        if nm_id is not None:
            stmt = stmt.where(CardQualityIssue.nm_id == nm_id)
        rows = list((await session.execute(stmt)).scalars())
        if severity:
            sev = str(severity).strip().lower()
            if sev == "media":
                rows = [
                    row for row in rows if self.issue_belongs_to_bucket(row, "media")
                ]
            elif sev in {"critical", "high"}:
                rows = [
                    row
                    for row in rows
                    if str(row.severity or "").lower() in {"critical", "high"}
                ]
            elif sev in {"warning", "medium", "improvement"}:
                rows = [
                    row
                    for row in rows
                    if str(row.severity or "").lower() in {"medium", "warning", "low"}
                    and not self.issue_belongs_to_bucket(row, "media")
                ]
        rows = [
            row
            for row in rows
            if self.issue_belongs_to_bucket(
                row, "media" if severity == "media" else bucket
            )
        ]
        return sorted(rows, key=self._queue_sort_key)

    def _queue_sort_key(self, issue: CardQualityIssue) -> tuple[Any, ...]:
        return self._source_issue_sort_key(issue, include_status=True)

    def _queue_issue_count_filters(
        self,
        *,
        account_id: int,
        bucket: str,
        severity: str | None,
        include_done: bool,
    ) -> list[Any]:
        statuses = {"new", "in_progress", "postponed", "blocked"}
        if include_done:
            statuses |= {"done", "resolved", "ignored"}

        category_expr = func.lower(func.coalesce(CardQualityIssue.category, ""))
        field_expr = func.lower(func.coalesce(CardQualityIssue.field_name, ""))
        is_media = or_(
            category_expr.in_(("media", "photos", "photo", "video")),
            field_expr.in_(("photos", "photo", "video", "videos")),
        )
        non_bucket_done = CardQualityIssue.status.in_(
            ("done", "fixed", "auto_fixed", "ignored", "skipped")
        )
        has_suggestion = or_(
            func.length(func.trim(func.coalesce(CardQualityIssue.suggested_value, "")))
            > 0,
            func.length(
                func.trim(func.coalesce(CardQualityIssue.ai_suggested_value, ""))
            )
            > 0,
        )

        conditions: list[Any] = [
            CardQualityIssue.account_id == account_id,
            CardQualityIssue.severity != "info",
            CardQualityIssue.status.in_(tuple(statuses)),
        ]
        normalized_severity = str(severity or "").strip().lower()
        if normalized_severity == "media":
            conditions.append(is_media)
        elif normalized_severity in {"critical", "high"}:
            conditions.append(
                func.lower(CardQualityIssue.severity).in_(("critical", "high"))
            )
        elif normalized_severity in {"warning", "medium", "improvement"}:
            conditions.extend(
                [
                    func.lower(CardQualityIssue.severity).in_(
                        ("medium", "warning", "low")
                    ),
                    ~is_media,
                ]
            )

        bucket_key = (
            "media"
            if normalized_severity == "media"
            else str(bucket or "actionable").strip().lower()
        )
        if bucket_key == "media":
            conditions.extend([~non_bucket_done, is_media])
        elif bucket_key == "human_check":
            conditions.extend(
                [
                    ~non_bucket_done,
                    CardQualityIssue.requires_human_check.is_(True),
                    ~is_media,
                ]
            )
        elif bucket_key == "actionable":
            conditions.extend(
                [
                    ~non_bucket_done,
                    ~is_media,
                    or_(
                        CardQualityIssue.requires_human_check.is_(True),
                        has_suggestion,
                        category_expr.in_(("title", "description", "characteristics")),
                    ),
                ]
            )
        return conditions

    def _sort_issues_source_order(
        self, issues: list[CardQualityIssue]
    ) -> list[CardQualityIssue]:
        return sorted(issues, key=self._source_issue_sort_key)

    def _source_issue_sort_key(
        self, issue: CardQualityIssue, *, include_status: bool = False
    ) -> tuple[Any, ...]:
        source_order = self._source_issue_order(issue)
        status_rank = SOURCE_STATUS_ORDER.get(
            str(getattr(issue, "status", "") or "").strip().lower(), 9
        )
        severity_rank = SOURCE_SEVERITY_ORDER.get(
            str(getattr(issue, "severity", "") or "").strip().lower(), 9
        )
        prefix = (status_rank,) if include_status else ()
        return (
            *prefix,
            source_order,
            severity_rank,
            -int(getattr(issue, "score_impact", 0) or 0),
            _field_key(getattr(issue, "field_name", None)),
            int(getattr(issue, "id", 0) or 0),
        )

    def _source_issue_order(self, issue: CardQualityIssue) -> int:
        code = str(getattr(issue, "issue_code", "") or "").strip().lower()
        source = str(getattr(issue, "source", "") or "").strip().lower()
        category = str(getattr(issue, "category", "") or "").strip().lower()
        field_key = _field_key(getattr(issue, "field_name", None))
        if code in SOURCE_BASIC_ISSUE_ORDER:
            return SOURCE_BASIC_ISSUE_ORDER[code]
        if code.startswith("wb_"):
            return 1000
        if source == "fixed_file" or code == "fixed_file_mismatch":
            return 2000
        if code.startswith("ai_") and category in {"title", "description"}:
            return 3000 if category == "title" else 3100
        if code.startswith("ai_"):
            return 3200
        if code == "description_refresh_needed":
            return 4000
        if field_key == "title" or category == "title":
            return 5000
        if field_key == "description" or category == "description":
            return 5100
        if category in {"photos", "photo", "media", "video"}:
            return 5200
        if category == "characteristics":
            return 5300
        return 9000

    async def _cards_by_nm_id(
        self, session: AsyncSession, *, account_id: int, nm_ids: list[int]
    ) -> dict[int, WBProductCard]:
        ids = sorted({int(item) for item in nm_ids if item is not None})
        if not ids:
            return {}
        rows = list(
            (
                await session.execute(
                    select(WBProductCard).where(
                        WBProductCard.account_id == account_id,
                        WBProductCard.nm_id.in_(ids),
                    )
                )
            ).scalars()
        )
        return {int(row.nm_id): row for row in rows}

    def _issue_with_card_payload(
        self,
        issue: CardQualityIssue,
        *,
        card: WBProductCard | None,
        pending_count: int | None,
    ) -> dict[str, Any]:
        payload = self._issue_payload(issue)
        photos = []
        if card is not None:
            raw_photos = card.photos if isinstance(card.photos, list) else []
            photos = [
                self.rules._photo_url(item)
                for item in raw_photos
                if self.rules._photo_url(item)
            ]
        payload.update(
            {
                "card_id": int(card.id) if card is not None else int(issue.nm_id),
                "card_nm_id": int(issue.nm_id),
                "card_title": card.title if card is not None else None,
                "card_vendor_code": card.vendor_code if card is not None else None,
                "card_photos": photos[:3],
                "card_pending_count": pending_count or 0,
                "requires_fixed_file": bool(
                    issue.source == "fixed_file"
                    or any(
                        (d or {}).get("requires_fixed_file")
                        for d in (issue.error_details_json or [])
                        if isinstance(d, dict)
                    )
                ),
            }
        )
        return payload

    def _parse_fixed_file_excel(self, content: bytes) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except (
            Exception
        ) as exc:  # pragma: no cover - dependency is declared, this is a deploy guard.
            raise ValueError("openpyxl_unavailable") from exc
        try:
            workbook = load_workbook(
                io.BytesIO(content), read_only=True, data_only=True
            )
        except Exception as exc:
            raise ValueError(f"fixed_file_parse_failed:{exc}") from exc

        def clean(value: Any) -> str:
            return str(value or "").strip()

        def header_key(value: Any) -> str:
            return re.sub(r"[^a-zа-я0-9]+", "_", clean(value).lower()).strip("_")

        aliases = {
            "nm_id": {
                "nm_id",
                "nmid",
                "nm",
                "артикул_wb",
                "артикул_вб",
                "артикул",
                "номенклатура",
                "id_карточки",
            },
            "brand": {"brand", "бренд"},
            "subject_name": {
                "subject",
                "subject_name",
                "subjectname",
                "предмет",
                "категория",
                "категория_wb",
            },
            "char_name": {
                "char_name",
                "characteristic",
                "характеристика",
                "название_характеристики",
                "поле",
            },
            "fixed_value": {
                "fixed_value",
                "value",
                "значение",
                "эталон",
                "эталонное_значение",
                "правильное_значение",
            },
        }

        parsed: list[dict[str, Any]] = []

        def parse_nm_id(raw_nm: str) -> int | None:
            if not raw_nm:
                return None
            try:
                return int(float(raw_nm.replace(" ", "").replace(",", ".")))
            except ValueError:
                return None

        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            raw_headers = [header_key(value) for value in rows[0]]
            display_headers = [clean(value) for value in rows[0]]
            header_map: dict[str, int] = {}
            for index, raw in enumerate(raw_headers):
                for target, names in aliases.items():
                    if raw in names and target not in header_map:
                        header_map[target] = index

            has_nm_header = "nm_id" in header_map
            has_long_headers = (
                has_nm_header
                and "char_name" in header_map
                and "fixed_value" in header_map
            )

            if has_long_headers:
                data_rows = rows[1:]
                for row in data_rows:

                    def cell(key: str) -> str:
                        index = header_map.get(key)
                        if index is None or index >= len(row):
                            return ""
                        return clean(row[index])

                    nm_id = parse_nm_id(cell("nm_id"))
                    char_name = cell("char_name")
                    fixed_value = cell("fixed_value")
                    if nm_id is None or not char_name or fixed_value == "":
                        continue
                    parsed.append(
                        {
                            "nm_id": nm_id,
                            "brand": cell("brand") or None,
                            "subject_name": cell("subject_name") or None,
                            "char_name": char_name,
                            "fixed_value": fixed_value,
                        }
                    )
                continue

            if has_nm_header:
                meta_indexes = {
                    index
                    for index, raw in enumerate(raw_headers)
                    if any(
                        raw in names
                        for names in (
                            aliases["nm_id"],
                            aliases["brand"],
                            aliases["subject_name"],
                        )
                    )
                }
                data_rows = rows[1:]
                for row in data_rows:

                    def cell_by_index(index: int | None) -> str:
                        if index is None or index >= len(row):
                            return ""
                        return clean(row[index])

                    nm_id = parse_nm_id(cell_by_index(header_map.get("nm_id")))
                    if nm_id is None:
                        continue
                    brand = cell_by_index(header_map.get("brand")) or None
                    subject_name = cell_by_index(header_map.get("subject_name")) or None
                    for index, char_name in enumerate(display_headers):
                        if index in meta_indexes or not char_name:
                            continue
                        if index >= len(row):
                            continue
                        fixed_value = clean(row[index])
                        if fixed_value == "":
                            continue
                        parsed.append(
                            {
                                "nm_id": nm_id,
                                "brand": brand,
                                "subject_name": subject_name,
                                "char_name": char_name,
                                "fixed_value": fixed_value,
                            }
                        )
                continue

            header_map = {
                "nm_id": 0,
                "brand": 1,
                "subject_name": 2,
                "char_name": 3,
                "fixed_value": 4,
            }
            for row in rows:

                def cell(key: str) -> str:
                    index = header_map.get(key)
                    if index is None or index >= len(row):
                        return ""
                    return clean(row[index])

                nm_id = parse_nm_id(cell("nm_id"))
                char_name = cell("char_name")
                fixed_value = cell("fixed_value")
                if nm_id is None or not char_name or fixed_value == "":
                    continue
                parsed.append(
                    {
                        "nm_id": nm_id,
                        "brand": cell("brand") or None,
                        "subject_name": cell("subject_name") or None,
                        "char_name": char_name,
                        "fixed_value": fixed_value,
                    }
                )
        return parsed

    def _merge_ai_fix(
        self,
        *,
        normalized: NormalizedCard,
        issue: RuleIssue,
        fix: dict[str, Any],
        recommended: str | None,
        invalid_reason: str | None,
    ) -> RuleIssue:
        candidate_values = [
            str(item).strip()
            for item in (fix.get("candidate_values") or [])
            if str(item).strip()
        ]
        suggestion_kind = str(fix.get("suggestion_kind") or "").strip().lower()
        review_only_kind = suggestion_kind in {"candidate", "draft_text", "no_safe_fix"}
        # The original catalog issue can be human-check only because it had no
        # safe value yet. Once AI returns a backend-validated exact value, do
        # not keep that old blocker attached to the merged fix.
        requires_human_check = bool(
            fix.get("requires_human_check")
            or invalid_reason
            or review_only_kind
            or (issue.requires_human_check and not recommended)
        )
        ai_reason = (
            str(
                fix.get("reason") or issue.ai_reason or issue.business_explanation or ""
            ).strip()
            or None
        )
        if invalid_reason:
            ai_reason = f"{ai_reason or 'AI fix rejected by backend validation'}; backend_validation={invalid_reason}"
        confirmed_recommended = (
            recommended if recommended and not requires_human_check else None
        )
        raw_recommended = self.rules._stringify_suggestion(fix.get("recommended_value"))
        if (
            raw_recommended
            and requires_human_check
            and raw_recommended not in candidate_values
        ):
            candidate_values = [raw_recommended, *candidate_values]
        confidence = (
            self._float(fix.get("confidence"))
            or issue.ai_confidence
            or issue.confidence
        )
        evidence = dict(
            fix.get("evidence")
            if isinstance(fix.get("evidence"), dict)
            else (issue.ai_evidence or {})
        )
        if invalid_reason:
            evidence = {
                **evidence,
                "backend_validation": invalid_reason,
                "why_human_check_required": invalid_reason,
            }
        photo_evidence = (
            fix.get("photo_evidence")
            if isinstance(fix.get("photo_evidence"), list)
            else issue.photo_evidence
        )
        used_sources = self._trusted_ai_used_sources(
            normalized,
            fix.get("used_sources") or issue.ai_used_sources or [],
            photo_evidence=photo_evidence,
        )
        evidence["visual_trust"] = self._visual_trust_evidence(normalized)
        if requires_human_check and not evidence.get("why_human_check_required"):
            evidence["why_human_check_required"] = ai_reason or "requires_manual_review"
        return RuleIssue(
            **{
                **issue.__dict__,
                "suggested_value": confirmed_recommended,
                "alternatives": candidate_values or issue.alternatives,
                "ai_suggested_value": confirmed_recommended
                or (issue.ai_suggested_value if not requires_human_check else None),
                "ai_reason": ai_reason,
                "ai_alternatives": candidate_values or issue.ai_alternatives,
                "ai_confidence": confidence,
                "requires_human_check": requires_human_check,
                "ai_reason_short": (ai_reason or "")[:240] or issue.ai_reason_short,
                "ai_reason_full": ai_reason or issue.ai_reason_full,
                "ai_evidence": evidence,
                "ai_used_sources": used_sources,
                "photo_evidence": photo_evidence,
                "source": "ai" if confirmed_recommended else issue.source,
                "confidence": confidence,
                "recommended_fix": self._ai_recommended_fix(
                    issue=issue,
                    recommended=confirmed_recommended,
                    requires_human_check=requires_human_check,
                    candidate_values=candidate_values,
                ),
            }
        )

    def _visual_trust_evidence(self, normalized: NormalizedCard) -> dict[str, Any]:
        audit = normalized.product_dna_audit or {}
        return {
            "trust_state": str(audit.get("trust_state") or audit.get("status") or ""),
            "grounded": bool(audit.get("grounded")),
            "reasons": [
                str(item) for item in (audit.get("reasons") or []) if str(item).strip()
            ],
        }

    def _trusted_ai_used_sources(
        self,
        normalized: NormalizedCard,
        sources: list[Any],
        *,
        photo_evidence: list[Any] | None,
    ) -> list[str]:
        grounded_product_dna = is_grounded_product_dna_audit(
            normalized.product_dna_audit
        )
        has_photo_evidence = any(
            isinstance(item, dict) for item in (photo_evidence or [])
        )
        out: list[str] = []
        seen: set[str] = set()
        for raw in sources or []:
            text = str(raw or "").strip()
            norm = text.lower()
            if not text or norm in seen:
                continue
            if norm == "product_dna" and not grounded_product_dna:
                continue
            if (
                norm in {"photo", "photos", "image", "images"}
                and not has_photo_evidence
            ):
                continue
            seen.add(norm)
            out.append(text)
        return out[:8]

    def _validated_ai_candidate(
        self, normalized: NormalizedCard, issue: RuleIssue, fix: dict[str, Any]
    ) -> tuple[str | None, str | None]:
        recommended = self.rules._stringify_suggestion(fix.get("recommended_value"))
        if not recommended:
            return None, None
        if _issue_values_equivalent(issue.current_value_json, recommended):
            return None, "no_op_suggestion"
        if self._is_date_issue(issue):
            return None, "date_field_requires_manual_source"
        field_key = _field_key(issue.field_name)
        if field_key in VISUAL_ALWAYS_REVIEW_ONLY_FIELDS:
            return recommended, "visual_field_requires_manual_check"
        if field_key in VISUAL_REQUIRE_GROUNDED_DNA_FIELDS:
            used_sources = {
                str(item).strip().lower()
                for item in (fix.get("used_sources") or [])
                if str(item).strip()
            }
            has_visual_evidence = bool(fix.get("photo_evidence")) or (
                is_grounded_product_dna_audit(normalized.product_dna_audit)
                and "product_dna" in used_sources
            )
            if not has_visual_evidence:
                return recommended, "visual_risky_field_without_trusted_evidence"
        if issue.allowed_values:
            valid, corrected, reason = self._validate_allowed_value(
                recommended,
                issue.allowed_values,
                issue.error_details or [],
                field_name=issue.field_name,
            )
            if not valid:
                return None, reason
            recommended = corrected or recommended
        if issue.category == "title" or _field_key(issue.field_name) == "title":
            keep_current, reason = should_keep_current_title_as_safer(
                normalized.title, recommended, self.rules._wb_card_payload(normalized)
            )
            if keep_current:
                return (
                    None,
                    f"title_business_guard:{reason.get('reason') or 'unsafe_candidate'}",
                )
        if (
            issue.category == "description"
            or _field_key(issue.field_name) == "description"
        ):
            valid, reason = validate_description_facts(
                recommended,
                self.rules._wb_card_payload(normalized),
                allow_visual_facts=False,
            )
            if not valid:
                return None, f"description_factual_guard:{reason}"
        return recommended, None

    def _validate_allowed_value(
        self,
        value: str,
        allowed_values: list[Any],
        error_details: list[Any],
        *,
        field_name: str | None,
    ) -> tuple[bool, str | None, str]:
        values = _split_issue_values(value) or [str(value).strip()]
        if allowed_values:
            allowed = [
                str(item).strip() for item in allowed_values if str(item).strip()
            ]
            allowed_norm = {_norm_text(item): item for item in allowed}
            corrected: list[str] = []
            invalid: list[str] = []
            fuzzy_allowed = _field_key(field_name) not in VISUAL_RISKY_FIELDS
            for item in values:
                exact = allowed_norm.get(_norm_text(item))
                if exact:
                    corrected.append(exact)
                    continue
                match = (
                    find_best_match(item, allowed, threshold=0.75)
                    if fuzzy_allowed
                    else None
                )
                if match:
                    corrected.append(match)
                else:
                    invalid.append(item)
            if invalid:
                return False, None, f"value_not_allowed:{invalid[0]}"
            values = corrected
        limits = self._effective_limits(error_details)
        min_count = limits.get("min")
        max_count = limits.get("max")
        if isinstance(min_count, int) and len(values) < min_count:
            return False, None, f"min_limit:{min_count}"
        if isinstance(max_count, int) and len(values) > max_count:
            return False, None, f"max_limit:{max_count}"
        corrected_value = (
            ", ".join(values) if len(values) > 1 else (values[0] if values else None)
        )
        return True, corrected_value, ""

    def _effective_limits(self, error_details: list[Any]) -> dict[str, int]:
        for detail in error_details or []:
            if not isinstance(detail, dict):
                continue
            if str(detail.get("type") or "").lower() != "limit":
                continue
            out: dict[str, int] = {}
            for key in ("min", "max"):
                try:
                    if detail.get(key) is not None:
                        out[key] = int(detail[key])
                except (TypeError, ValueError):
                    pass
            return out
        return {}

    def _is_date_issue(self, issue: RuleIssue) -> bool:
        if _field_key(issue.field_name) in DATE_FIELD_HINTS:
            return True
        texts = [
            issue.field_name,
            issue.title,
            issue.business_explanation,
            issue.recommended_fix,
        ]
        texts.extend(str(item) for item in (issue.error_details or []) if item)
        return any(_contains_date_context_text(item) for item in texts)

    def _finalize_rule_issues(
        self,
        normalized: NormalizedCard,
        issues: list[RuleIssue],
        fixed_fields: dict[str, str],
    ) -> list[RuleIssue]:
        fixed_keys = {_norm_text(name) for name in fixed_fields}
        filtered: list[RuleIssue] = []
        for issue in issues:
            if self._is_date_issue(issue) and issue.source != "fixed_file":
                continue
            if (
                _field_key(issue.field_name) in fixed_keys
                and issue.source != "fixed_file"
            ):
                continue
            filtered.append(issue)

        filtered = self._collapse_compound_overlaps(filtered)
        filtered = self._collapse_same_field_competitors(filtered)
        filtered = self._collapse_description_refresh_overlaps(filtered)

        finalized: list[RuleIssue] = []
        for issue in filtered:
            issue = self._apply_safety_gates(normalized, issue)
            if not self._is_actionable_issue(issue):
                continue
            finalized.append(issue)
        return self._sort_rule_issues_source_order(self._dedupe_rule_issues(finalized))

    def _collapse_compound_overlaps(self, issues: list[RuleIssue]) -> list[RuleIssue]:
        compound_targets: list[tuple[int, set[str], set[int], set[str]]] = []
        for index, issue in enumerate(issues):
            fixes = self._extract_compound_fixes(issue.error_details or [])
            if not fixes:
                continue
            target_paths: set[str] = set()
            target_charc_ids: set[int] = set()
            target_names: set[str] = set()
            for fix in fixes:
                if not isinstance(fix, dict):
                    continue
                path = self._normalize_compound_field_path(
                    fix.get("field_path") or fix.get("name")
                )
                if path:
                    target_paths.add(path.lower())
                charc_id = self.rules._int(fix.get("charc_id", fix.get("charcId")))
                if charc_id is not None:
                    target_charc_ids.add(charc_id)
                name = str(fix.get("name") or "").strip().lower()
                if name:
                    target_names.add(name)
            if target_paths or target_charc_ids or target_names:
                compound_targets.append(
                    (index, target_paths, target_charc_ids, target_names)
                )
        if not compound_targets:
            return issues

        collapsed: list[RuleIssue] = []
        for index, issue in enumerate(issues):
            issue_path = self._normalize_compound_field_path(issue.field_name)
            issue_name = ""
            if issue_path.startswith("characteristics."):
                issue_name = issue_path.split("characteristics.", 1)[1].strip().lower()
            remove = False
            for compound_index, paths, charc_ids, names in compound_targets:
                if index == compound_index or issue.category not in {
                    "characteristics",
                    "title",
                    "description",
                    "identity",
                }:
                    continue
                if issue.charc_id is not None and issue.charc_id in charc_ids:
                    remove = True
                    break
                if issue_path and issue_path in paths:
                    remove = True
                    break
                if issue_name and issue_name in names:
                    remove = True
                    break
            if not remove:
                collapsed.append(issue)
        return collapsed

    def _collapse_same_field_competitors(
        self, issues: list[RuleIssue]
    ) -> list[RuleIssue]:
        groups: dict[tuple[str, int], list[RuleIssue]] = {}
        passthrough: list[RuleIssue] = []
        for issue in issues:
            path = self._normalize_compound_field_path(issue.field_name)
            if not path:
                passthrough.append(issue)
                continue
            charc_id = (
                int(issue.charc_id or 0) if path.startswith("characteristics.") else 0
            )
            groups.setdefault((path, charc_id), []).append(issue)

        out: list[RuleIssue] = []
        for items in groups.values():
            if len(items) == 1:
                out.extend(items)
                continue
            description_short = [
                item
                for item in items
                if self._normalize_compound_field_path(item.field_name) == "description"
                and item.issue_code == "description_too_short"
            ]
            winner = sorted(
                description_short or items, key=self._rule_issue_rank_for_field_collapse
            )[0]
            candidate_values: list[str] = []
            merged_codes: list[str] = []
            for item in items:
                if item.issue_code not in merged_codes:
                    merged_codes.append(item.issue_code)
                values: list[Any] = [item.suggested_value, item.ai_suggested_value]
                values.extend(item.ai_alternatives or [])
                values.extend(item.alternatives or [])
                for value in values:
                    text = str(value or "").strip()
                    if text and text not in candidate_values:
                        candidate_values.append(text)
            evidence = dict(winner.ai_evidence or {})
            evidence["merged_issue_codes"] = merged_codes
            if len(candidate_values) > 1:
                reason = "Конфликтующие подсказки по одному полю объединены в один human-check issue."
                evidence["merged_candidate_values"] = candidate_values[:8]
                evidence["why_human_check_required"] = reason
                winner = replace(
                    winner,
                    requires_human_check=True,
                    suggested_value=None,
                    alternatives=candidate_values[:5],
                    ai_alternatives=candidate_values[:5],
                    ai_reason=winner.ai_reason or reason,
                    ai_reason_full=f"{winner.ai_reason_full or ''} {reason}".strip(),
                    ai_evidence=evidence,
                )
            elif len(merged_codes) > 1:
                winner = replace(winner, ai_evidence=evidence)
            out.append(winner)
        out.extend(passthrough)
        return out

    def _collapse_description_refresh_overlaps(
        self, issues: list[RuleIssue]
    ) -> list[RuleIssue]:
        has_stronger_description_issue = any(
            (
                issue.category == "description"
                or _field_key(issue.field_name) == "description"
            )
            and issue.issue_code != "description_refresh_needed"
            for issue in issues
        )
        if not has_stronger_description_issue:
            return issues
        return [
            issue
            for issue in issues
            if not (
                (
                    issue.category == "description"
                    or _field_key(issue.field_name) == "description"
                )
                and issue.issue_code == "description_refresh_needed"
            )
        ]

    def _extract_compound_fixes(self, error_details: list[Any]) -> list[dict[str, Any]]:
        for item in error_details or []:
            if not isinstance(item, dict):
                continue
            if item.get("type") == "compound" or item.get("fix_action") == "compound":
                fixes = item.get("fixes")
                return fixes if isinstance(fixes, list) else []
        return []

    def _extract_compound_candidate_values(
        self, error_details: list[Any], field_path: str | None
    ) -> list[str]:
        target_path = self._normalize_compound_field_path(field_path)
        out: list[str] = []
        for fix in self._extract_compound_fixes(error_details):
            if not isinstance(fix, dict):
                continue
            fix_path = self._normalize_compound_field_path(
                fix.get("field_path") or fix.get("name")
            )
            if target_path and fix_path and target_path != fix_path:
                continue
            raw_value = (
                fix.get("value")
                if "value" in fix
                else fix.get("suggested_value", fix.get("recommended_value"))
            )
            value = self.rules._stringify_suggestion(raw_value)
            if value and value not in out:
                out.append(value)
        return out

    def _allowed_values_for_ai_issue(
        self, field_path: str | None, category: str | None
    ) -> list[Any]:
        normalized_path = self._normalize_compound_field_path(field_path)
        if not normalized_path.startswith("characteristics."):
            return []
        if category and str(category).strip().lower() != "characteristics":
            return []
        char_name = normalized_path.split("characteristics.", 1)[1].strip()
        if not char_name:
            return []
        try:
            values = get_catalog().get_allowed_values(char_name)
        except Exception:
            return []
        return list(values or [])

    def _normalize_compound_field_path(self, raw: Any) -> str:
        text = str(raw or "").strip()
        if not text:
            return ""
        lower = text.lower()
        if lower in {"title", "название"}:
            return "title"
        if lower in {"description", "описание"}:
            return "description"
        if lower.startswith("characteristics."):
            return f"characteristics.{text.split('.', 1)[1].strip()}".lower()
        if lower in {"photos", "photo", "video", "videos", "brand", "subject_name"}:
            return lower
        return f"characteristics.{text}".lower()

    def _rule_issue_rank_for_field_collapse(
        self, issue: RuleIssue
    ) -> tuple[int, int, int]:
        has_value = bool(
            str(issue.suggested_value or issue.ai_suggested_value or "").strip()
        )
        source_rank = {"auto_fix": 0, "code": 1, "fixed_file": 2, "ai": 3}.get(
            str(issue.source or "").lower(), 9
        )
        severity_rank = SOURCE_SEVERITY_ORDER.get(str(issue.severity or "").lower(), 9)
        return (
            0 if has_value and not issue.requires_human_check else 1,
            source_rank,
            severity_rank,
        )

    def _apply_safety_gates(
        self, normalized: NormalizedCard, issue: RuleIssue
    ) -> RuleIssue:
        reason: str | None = None
        field_key = _field_key(issue.field_name)
        if field_key in VISUAL_ALWAYS_REVIEW_ONLY_FIELDS:
            reason = (
                "Точное визуальное значение для этого поля требует ручной проверки."
            )
        elif field_key in VISUAL_REQUIRE_GROUNDED_DNA_FIELDS:
            grounded = is_grounded_product_dna_audit(normalized.product_dna_audit)
            used_sources = {
                str(item).strip().lower()
                for item in (issue.ai_used_sources or [])
                if str(item).strip()
            }
            if not grounded or (
                "product_dna" not in used_sources and not issue.photo_evidence
            ):
                reason = (
                    "Для визуально рискованного поля нет доверенного visual evidence."
                )
        if issue.category == "description" and issue.suggested_value:
            valid, failure = validate_description_facts(
                issue.suggested_value,
                self.rules._wb_card_payload(normalized),
                allow_visual_facts=False,
            )
            if not valid:
                reason = f"Описание требует ручной проверки: {failure}"
        if issue.category == "title" and issue.suggested_value:
            keep_current, info = should_keep_current_title_as_safer(
                normalized.title,
                issue.suggested_value,
                self.rules._wb_card_payload(normalized),
            )
            if keep_current:
                reason = f"Название требует ручной проверки: {info.get('reason') or 'unsafe_candidate'}"
        if self._is_unsafe_destructive_no_value_issue(issue):
            evidence = dict(issue.ai_evidence or {})
            evidence["non_actionable_reason"] = (
                "AI proposed clear/swap without a safe replacement value."
            )
            return replace(
                issue,
                error_details=self._with_non_actionable_marker(
                    issue.error_details or [],
                    "AI предложил очистку/перенос без безопасного значения.",
                ),
                ai_evidence=evidence,
            )
        if not reason:
            return issue
        candidate = str(issue.suggested_value or issue.ai_suggested_value or "").strip()
        alternatives = list(issue.alternatives or [])
        ai_alternatives = list(issue.ai_alternatives or [])
        if candidate:
            if candidate not in alternatives:
                alternatives = [candidate, *alternatives]
            if candidate not in ai_alternatives:
                ai_alternatives = [candidate, *ai_alternatives]
        evidence = dict(issue.ai_evidence or {})
        evidence["why_human_check_required"] = reason
        evidence["visual_trust"] = self._visual_trust_evidence(normalized)
        return RuleIssue(
            **{
                **issue.__dict__,
                "requires_human_check": True,
                "suggested_value": None,
                "alternatives": alternatives or issue.alternatives,
                "ai_reason": issue.ai_reason or reason,
                "ai_suggested_value": None,
                "ai_alternatives": ai_alternatives or issue.ai_alternatives,
                "ai_reason_short": issue.ai_reason_short or reason[:240],
                "ai_reason_full": f"{issue.ai_reason_full or ''} {reason}".strip(),
                "ai_evidence": evidence,
                "ai_used_sources": self._trusted_ai_used_sources(
                    normalized,
                    issue.ai_used_sources or [],
                    photo_evidence=issue.photo_evidence,
                ),
            }
        )

    def _issue_has_fix_action(self, issue: RuleIssue, actions: set[str]) -> bool:
        normalized_actions = {str(item).strip().lower() for item in actions}
        for item in issue.error_details or []:
            if not isinstance(item, dict):
                continue
            action = (
                str(item.get("fix_action") or item.get("type") or "").strip().lower()
            )
            if action in normalized_actions:
                return True
        return False

    def _has_non_actionable_marker(self, issue: RuleIssue) -> bool:
        return any(
            isinstance(item, dict)
            and str(item.get("type") or "").strip().lower() == "non_actionable"
            for item in (issue.error_details or [])
        )

    def _with_non_actionable_marker(
        self, error_details: list[Any], reason: str
    ) -> list[Any]:
        details = [item for item in (error_details or []) if isinstance(item, dict)]
        if any(
            str(item.get("type") or "").strip().lower() == "non_actionable"
            for item in details
        ):
            return details
        return [*details, {"type": "non_actionable", "reason": reason}]

    def _is_unsafe_destructive_no_value_issue(self, issue: RuleIssue) -> bool:
        if str(issue.source or "").strip().lower() not in {"ai", "code"}:
            return False
        if str(issue.suggested_value or issue.ai_suggested_value or "").strip():
            return False
        return self._issue_has_fix_action(issue, {"clear", "swap"})

    def _is_actionable_issue(self, issue: RuleIssue) -> bool:
        if self._has_non_actionable_marker(issue):
            return False
        if issue.source == "fixed_file":
            return True
        if issue.category == "media":
            return True
        if issue.requires_human_check:
            return True
        if _issue_values_equivalent(
            issue.current_value_json, issue.suggested_value or issue.ai_suggested_value
        ):
            return False
        if str(issue.suggested_value or issue.ai_suggested_value or "").strip():
            return True
        if issue.category in {"title", "description"}:
            return True
        if issue.allowed_values and not issue.requires_human_check:
            return True
        return issue.severity in {"critical", "high"}

    def _dedupe_rule_issues(self, issues: list[RuleIssue]) -> list[RuleIssue]:
        groups: dict[str, RuleIssue] = {}
        for issue in issues:
            signature = json.dumps(
                {
                    "code": issue.issue_code,
                    "category": issue.category,
                    "field": _field_key(issue.field_name),
                    "charc_id": issue.charc_id or 0,
                    "current": self.rules._stable_value_context(
                        issue.current_value_json
                    ),
                    "suggested": self.rules._stable_value_context(
                        issue.suggested_value or issue.ai_suggested_value
                    ),
                    "errors": self.rules._stable_value_context(
                        issue.error_details or []
                    ),
                },
                ensure_ascii=False,
                sort_keys=True,
                default=str,
            )
            existing = groups.get(signature)
            if existing is None or self._issue_rank(issue) < self._issue_rank(existing):
                groups[signature] = issue
        return list(groups.values())

    def _sort_rule_issues_source_order(
        self, issues: list[RuleIssue]
    ) -> list[RuleIssue]:
        adapter = SimpleNamespace
        return sorted(
            issues,
            key=lambda issue: self._source_issue_sort_key(
                adapter(
                    issue_code=issue.issue_code,
                    source=issue.source,
                    category=issue.category,
                    field_name=issue.field_name,
                    status="new",
                    severity=issue.severity,
                    score_impact=issue.score_impact,
                    id=0,
                )
            ),
        )

    def _issue_rank(self, issue: RuleIssue) -> tuple[int, int, int]:
        has_confirmed = (
            bool(str(issue.suggested_value or "").strip())
            and not issue.requires_human_check
        )
        source_rank = {
            "fixed_file": 0,
            "auto_fix": 1,
            "code": 2,
            "wb_catalog": 2,
            "ai": 3,
        }.get(str(issue.source or ""), 9)
        severity_rank = {
            "critical": 0,
            "high": 1,
            "medium": 2,
            "low": 3,
            "info": 4,
        }.get(issue.severity, 9)
        return (0 if has_confirmed else 1, source_rank, severity_rank)

    def _ai_recommended_fix(
        self,
        *,
        issue: RuleIssue,
        recommended: str | None,
        requires_human_check: bool,
        candidate_values: list[str],
    ) -> str:
        field_name = str(issue.field_name or issue.title)
        if recommended and not requires_human_check:
            return f"Заменить «{field_name}» на «{recommended}»."
        if recommended:
            return f"Проверить «{field_name}» вручную и подтвердить кандидат «{recommended}»."
        if candidate_values:
            return f"Проверить «{field_name}» вручную. Кандидаты: {', '.join(candidate_values[:3])}."
        return issue.recommended_fix

    def _float(self, value: Any) -> float | None:
        try:
            if value in (None, ""):
                return None
            return float(value)
        except (TypeError, ValueError):
            return None

    async def analyze_account(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        force: bool = False,
        limit: int = 100,
        requested_by_user_id: int | None = None,
    ) -> CardQualityAnalyzeResponse:
        active_run = (
            await session.execute(
                select(CardQualityAnalysisRun)
                .where(
                    CardQualityAnalysisRun.account_id == account_id,
                    CardQualityAnalysisRun.run_type == "account_batch",
                    CardQualityAnalysisRun.status.in_(("queued", "running")),
                )
                .order_by(CardQualityAnalysisRun.id.desc())
                .limit(1)
            )
        ).scalar_one_or_none()
        if active_run is not None:
            return CardQualityAnalyzeResponse(
                status="queued",
                run_id=int(active_run.id),
                account_id=account_id,
                run_type=active_run.run_type,
                cards_total=active_run.cards_total,
                eligible_total=active_run.eligible_total,
                cards_processed=active_run.cards_processed,
                cards_analyzed=active_run.cards_analyzed,
                cards_skipped_unchanged=active_run.cards_skipped_unchanged,
                cards_failed=active_run.cards_failed,
                cards_clean=active_run.cards_clean,
                cards_with_issues=active_run.cards_with_issues,
                issues_created=active_run.issues_created,
                issues_resolved=active_run.issues_resolved,
                message="card quality analysis is already queued or running",
            )
        eligible_total = int(
            (
                await session.execute(
                    select(func.count(func.distinct(WBProductCard.nm_id))).where(
                        WBProductCard.account_id == account_id
                    )
                )
            ).scalar()
            or 0
        )
        run = CardQualityAnalysisRun(
            account_id=account_id,
            run_type="account_batch",
            status="queued",
            requested_by_user_id=requested_by_user_id,
            cards_total=eligible_total,
            eligible_total=eligible_total,
            cursor_json={"last_nm_id": 0, "batch_limit": limit, "force": bool(force)},
            heartbeat_at=utcnow(),
        )
        session.add(run)
        await session.flush()
        await session.commit()
        return CardQualityAnalyzeResponse(
            status="queued",
            run_id=int(run.id),
            account_id=account_id,
            run_type=run.run_type,
            cards_total=run.cards_total,
            eligible_total=run.eligible_total,
            cards_processed=run.cards_processed,
            cards_analyzed=run.cards_analyzed,
            cards_skipped_unchanged=run.cards_skipped_unchanged,
            cards_failed=run.cards_failed,
            cards_clean=run.cards_clean,
            cards_with_issues=run.cards_with_issues,
            issues_created=run.issues_created,
            issues_resolved=run.issues_resolved,
            message="card quality analysis queued",
        )

    async def process_run_batch(
        self, session: AsyncSession, *, run_id: int, batch_limit: int = 100
    ) -> CardQualityRunRead:
        run = await session.get(CardQualityAnalysisRun, run_id)
        if run is None:
            raise ValueError("run_not_found")
        if run.status not in {"queued", "running"}:
            return CardQualityRunRead.model_validate(run, from_attributes=True)
        cursor = dict(run.cursor_json or {})
        last_nm_id = int(cursor.get("last_nm_id") or 0)
        force = bool(cursor.get("force") or False)
        effective_limit = max(
            1, min(int(cursor.get("batch_limit") or batch_limit), batch_limit, 1000)
        )
        run.status = "running"
        run.started_at = run.started_at or utcnow()
        run.heartbeat_at = utcnow()
        nm_ids = list(
            (
                await session.execute(
                    select(WBProductCard.nm_id)
                    .where(
                        WBProductCard.account_id == run.account_id,
                        WBProductCard.nm_id > last_nm_id,
                    )
                    .order_by(WBProductCard.nm_id.asc())
                    .limit(effective_limit)
                )
            ).scalars()
        )
        for nm_id in nm_ids:
            try:
                await self.analyze_product(
                    session,
                    account_id=int(run.account_id),
                    nm_id=int(nm_id),
                    force=force,
                    requested_by_user_id=run.requested_by_user_id,
                    run=run,
                )
            except Exception as exc:
                run.cards_processed += 1
                run.error_summary = exc.__class__.__name__
                run.last_processed_key = str(nm_id)
                run.cursor_json = {
                    "last_nm_id": int(nm_id),
                    "batch_limit": effective_limit,
                    "force": force,
                }
                run.heartbeat_at = utcnow()
        if not nm_ids or run.cards_processed >= run.eligible_total:
            run.status = "partial" if run.cards_failed else "completed"
            run.finished_at = utcnow()
            await self._upsert_registry(session, account_id=int(run.account_id))
            await self._record_module_run(session, run=run)
        await session.flush()
        await session.refresh(run)
        result = CardQualityRunRead.model_validate(run, from_attributes=True)
        await session.commit()
        self.clear_runtime_caches()
        return result

    async def retry_run(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        run_id: int,
        requested_by_user_id: int | None,
    ) -> CardQualityAnalyzeResponse:
        source_run = await session.get(CardQualityAnalysisRun, run_id)
        if source_run is None or int(source_run.account_id) != int(account_id):
            raise ValueError("run_not_found")
        if source_run.status in {"queued", "running"}:
            return CardQualityAnalyzeResponse(
                status="queued",
                run_id=int(source_run.id),
                account_id=account_id,
                run_type=source_run.run_type,
                cards_total=source_run.cards_total,
                eligible_total=source_run.eligible_total,
                cards_processed=source_run.cards_processed,
                cards_analyzed=source_run.cards_analyzed,
                cards_skipped_unchanged=source_run.cards_skipped_unchanged,
                cards_failed=source_run.cards_failed,
                cards_clean=source_run.cards_clean,
                cards_with_issues=source_run.cards_with_issues,
                issues_created=source_run.issues_created,
                issues_resolved=source_run.issues_resolved,
                message="card quality run is already active",
            )
        retry = CardQualityAnalysisRun(
            account_id=account_id,
            run_type=source_run.run_type,
            status="queued",
            requested_by_user_id=requested_by_user_id,
            cards_total=source_run.eligible_total or source_run.cards_total,
            eligible_total=source_run.eligible_total or source_run.cards_total,
            cursor_json={"last_nm_id": 0, "force": True},
            heartbeat_at=utcnow(),
            attempt=int(source_run.attempt or 1) + 1,
        )
        session.add(retry)
        await session.flush()
        await session.commit()
        return CardQualityAnalyzeResponse(
            status="queued",
            run_id=int(retry.id),
            account_id=account_id,
            run_type=retry.run_type,
            cards_total=retry.cards_total,
            eligible_total=retry.eligible_total,
            message="card quality retry queued",
        )

    async def product_quality(
        self, session: AsyncSession, *, account_id: int, nm_id: int
    ) -> PortalProductQualityRead:
        snapshot = await self.latest_snapshot(
            session, account_id=account_id, nm_id=nm_id
        )
        if snapshot is None:
            card_exists = (
                await session.execute(
                    select(WBProductCard.id).where(
                        WBProductCard.account_id == account_id,
                        WBProductCard.nm_id == nm_id,
                    )
                )
            ).first()
            if card_exists:
                return PortalProductQualityRead(
                    status="not_analyzed",
                    module="card_quality",
                    source="card_quality",
                    mode="local",
                    nm_id=nm_id,
                    message="card quality snapshot is not analyzed yet",
                    analysis_available=True,
                    analyze_endpoint=f"/api/v1/portal/card-quality/products/{nm_id}/analyze",
                )
            return PortalProductQualityRead(
                status="empty",
                module="card_quality",
                source="card_quality",
                mode="local",
                nm_id=nm_id,
                message="product card is not present in finance product cards",
            )
        issues = await self.open_issues_for_product(
            session, account_id=account_id, nm_id=nm_id
        )
        issue_payloads = [self._issue_payload(issue) for issue in issues]
        by_category: dict[str, int] = {}
        for issue in issues:
            by_category[issue.category] = by_category.get(issue.category, 0) + 1
        summary = dict(snapshot.summary_json or {})
        product_dna_audit = (
            summary.get("product_dna_audit")
            if isinstance(summary.get("product_dna_audit"), dict)
            else {}
        )
        warnings: list[str] = []
        if summary.get("product_dna_enabled") is False:
            warnings.append(
                str(product_dna_audit.get("reason") or "product_dna_disabled")
            )
        non_info_issues = [
            issue for issue in issue_payloads if issue.get("severity") != "info"
        ]
        status = "ok" if snapshot.status == "clean" else snapshot.status
        card_payload = self._snapshot_card_payload(snapshot)
        return PortalProductQualityRead(
            status=status,  # type: ignore[arg-type]
            module="card_quality",
            source="card_quality",
            mode="local",
            nm_id=nm_id,
            card_id=int(snapshot.source_card_id or 0) or None,
            score=snapshot.score,
            updated_at=snapshot.analyzed_at,
            analyzed_at=snapshot.analyzed_at,
            source_revision=snapshot.source_revision,
            score_breakdown=summary,
            category_scores=summary.get("category_scores") or {},
            critical_issue_count=int(summary.get("critical_count") or 0),
            warning_issue_count=len(non_info_issues)
            - int(summary.get("critical_count") or 0),
            issues_by_category=by_category,
            title_issues=[
                item for item in issue_payloads if item.get("category") == "title"
            ],
            description_issues=[
                item for item in issue_payloads if item.get("category") == "description"
            ],
            characteristics_issues=[
                item
                for item in issue_payloads
                if item.get("category") == "characteristics"
            ],
            photo_video_issues=[
                item
                for item in issue_payloads
                if item.get("category") in {"media", "photos", "photo", "video"}
            ],
            issues=issue_payloads,
            recommendations=[
                item["recommendation"]
                for item in issue_payloads
                if item.get("recommendation")
            ][:10],
            summary={
                **summary,
                "photos_count": snapshot.photos_count,
                "video_count": snapshot.video_count,
                "mode": "local",
                "card": card_payload,
            },
            next_recommended_action=self._next_recommended_action(
                issue_payloads, nm_id=nm_id
            ),
            warnings=warnings,
            message="card quality analyzed locally",
            raw={
                "snapshot_id": snapshot.id,
                "issue_count": len(issue_payloads),
                "source": "finance_product_cards",
                "product_dna_status": product_dna_audit.get("status")
                or product_dna_audit.get("trust_state"),
                "product_dna_disabled_reason": product_dna_audit.get("reason")
                if summary.get("product_dna_enabled") is False
                else None,
                "card": card_payload,
            },
        )

    def _snapshot_card_payload(self, snapshot: CardQualitySnapshot) -> dict[str, Any]:
        media = copy.deepcopy(snapshot.media_json or {})
        characteristics = copy.deepcopy(snapshot.characteristics_json or [])
        photos = media.get("photos") if isinstance(media, dict) else []
        videos = media.get("videos") if isinstance(media, dict) else []
        sizes = media.get("sizes") if isinstance(media, dict) else []
        return {
            "nm_id": snapshot.nm_id,
            "card_id": int(snapshot.source_card_id or 0) or None,
            "title": snapshot.title,
            "description": snapshot.description,
            "brand": snapshot.brand,
            "subject_name": snapshot.subject_name,
            "vendor_code": snapshot.vendor_code,
            "characteristics": characteristics
            if isinstance(characteristics, list)
            else [],
            "media": media if isinstance(media, dict) else {},
            "photos": photos if isinstance(photos, list) else [],
            "videos": videos if isinstance(videos, list) else [],
            "sizes": sizes if isinstance(sizes, list) else [],
            "photos_count": snapshot.photos_count,
            "video_count": snapshot.video_count,
            "source_revision": snapshot.source_revision,
            "source_updated_at": snapshot.source_updated_at.isoformat()
            if snapshot.source_updated_at
            else None,
            "primary_photo": self._snapshot_primary_photo(
                photos if isinstance(photos, list) else []
            ),
        }

    def _snapshot_primary_photo(self, photos: list[Any]) -> str | None:
        for photo in photos:
            if isinstance(photo, str) and photo.strip():
                return photo.strip()
            if not isinstance(photo, dict):
                continue
            for key in (
                "canonical_url",
                "big",
                "c516x688",
                "square",
                "tm",
                "url",
                "photo",
                "image",
            ):
                value = photo.get(key)
                if isinstance(value, str) and value.strip():
                    return value.strip()
            variants = photo.get("variants")
            if isinstance(variants, dict):
                for key in ("big", "c516x688", "square", "tm", "url"):
                    value = variants.get(key)
                    if isinstance(value, str) and value.strip():
                        return value.strip()
        return None

    async def quality_actions(
        self, session: AsyncSession, *, account_id: int, limit: int = 100
    ) -> list[PortalActionRead]:
        rows = list(
            (
                await session.execute(
                    select(CardQualityIssue).where(
                        CardQualityIssue.account_id == account_id,
                        CardQualityIssue.status.in_(tuple(ACTIVE_ISSUE_STATUSES)),
                        CardQualityIssue.resolved_at.is_(None),
                        CardQualityIssue.severity.in_(("critical", "high")),
                    )
                )
            ).scalars()
        )
        return [
            self._action_from_issue(account_id=account_id, issue=issue)
            for issue in self._sort_issues_source_order(rows)[:limit]
        ]

    async def latest_snapshot(
        self, session: AsyncSession, *, account_id: int, nm_id: int
    ) -> CardQualitySnapshot | None:
        return (
            await session.execute(
                select(CardQualitySnapshot)
                .where(
                    CardQualitySnapshot.account_id == account_id,
                    CardQualitySnapshot.nm_id == nm_id,
                )
                .order_by(
                    CardQualitySnapshot.analyzed_at.desc().nullslast(),
                    CardQualitySnapshot.id.desc(),
                )
                .limit(1)
            )
        ).scalar_one_or_none()

    async def open_issues_for_product(
        self, session: AsyncSession, *, account_id: int, nm_id: int
    ) -> list[CardQualityIssue]:
        rows = list(
            (
                await session.execute(
                    select(CardQualityIssue).where(
                        CardQualityIssue.account_id == account_id,
                        CardQualityIssue.nm_id == nm_id,
                        CardQualityIssue.status.in_(tuple(ACTIVE_ISSUE_STATUSES)),
                        CardQualityIssue.resolved_at.is_(None),
                    )
                )
            ).scalars()
        )
        return self._sort_issues_source_order(rows)

    async def list_product_cards(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        search: str | None = None,
        quality_status: str | None = None,
        score_filter: str | None = None,
        ai_filter: str | None = None,
        media_filter: str | None = None,
        sort_by: str = "quality_issues",
        sort_dir: str = "desc",
        limit: int = 100,
        offset: int = 0,
    ) -> CardQualityProductsPage:
        cache_key = (
            "product_cards",
            int(account_id),
            str(search or "").strip().lower(),
            str(quality_status or "").strip().lower(),
            str(score_filter or "").strip().lower(),
            str(ai_filter or "").strip().lower(),
            str(media_filter or "").strip().lower(),
            str(sort_by or "").strip().lower(),
            str(sort_dir or "").strip().lower(),
            int(limit),
            int(offset),
        )
        cached = self._product_cards_cache.get(cache_key)
        if cached is not None:
            return cached.model_copy(deep=True)

        cards = list(
            (
                await session.execute(
                    select(WBProductCard)
                    .where(WBProductCard.account_id == account_id)
                    .order_by(
                        WBProductCard.updated_at.desc().nullslast(),
                        WBProductCard.id.desc(),
                    )
                )
            ).scalars()
        )
        query = str(search or "").strip().lower()
        if query:
            cards = [
                card
                for card in cards
                if query
                in " ".join(
                    [
                        str(card.nm_id or ""),
                        str(card.vendor_code or ""),
                        str(card.title or ""),
                        str(card.brand or ""),
                        str(card.subject_name or ""),
                    ]
                ).lower()
            ]

        nm_ids = [int(card.nm_id) for card in cards if card.nm_id is not None]
        latest_snapshots: dict[int, CardQualitySnapshot] = {}
        issues_by_nm: dict[int, list[CardQualityIssue]] = {}
        if nm_ids:
            snapshot_rows = list(
                (
                    await session.execute(
                        select(CardQualitySnapshot)
                        .where(
                            CardQualitySnapshot.account_id == account_id,
                            CardQualitySnapshot.nm_id.in_(nm_ids),
                        )
                        .order_by(
                            CardQualitySnapshot.nm_id.asc(),
                            CardQualitySnapshot.analyzed_at.desc().nullslast(),
                            CardQualitySnapshot.id.desc(),
                        )
                    )
                ).scalars()
            )
            for snapshot in snapshot_rows:
                latest_snapshots.setdefault(int(snapshot.nm_id), snapshot)

            issue_rows = list(
                (
                    await session.execute(
                        select(CardQualityIssue).where(
                            CardQualityIssue.account_id == account_id,
                            CardQualityIssue.nm_id.in_(nm_ids),
                            CardQualityIssue.status.in_(tuple(ACTIVE_ISSUE_STATUSES)),
                            CardQualityIssue.severity != "info",
                            CardQualityIssue.resolved_at.is_(None),
                        )
                    )
                ).scalars()
            )
            for issue in issue_rows:
                if not self._should_surface_checker_issue(issue):
                    continue
                issues_by_nm.setdefault(int(issue.nm_id), []).append(issue)
            for nm_id, issues in list(issues_by_nm.items()):
                issues_by_nm[nm_id] = self._sort_issues_source_order(issues)

        items: list[CardQualityProductListItem] = []
        for card in cards:
            nm_id = int(card.nm_id)
            snapshot = latest_snapshots.get(nm_id)
            issues = issues_by_nm.get(nm_id, [])
            critical_count = sum(
                1
                for issue in issues
                if str(issue.severity or "").lower() in {"critical", "high"}
            )
            warning_count = sum(
                1
                for issue in issues
                if str(issue.severity or "").lower() not in {"critical", "high"}
            )
            ai_issue_count = sum(
                1 for issue in issues if str(issue.source or "").lower() == "ai"
            )
            no_solution_ai_count = sum(
                1
                for issue in issues
                if str(issue.source or "").lower() == "ai"
                and not self._issue_has_any_fix_path(issue)
            )
            top_issue = issues[0] if issues else None
            raw_status = str(
                snapshot.status if snapshot is not None else "not_analyzed"
            )
            status = "ok" if raw_status == "clean" else raw_status
            items.append(
                CardQualityProductListItem(
                    account_id=account_id,
                    nm_id=nm_id,
                    title=snapshot.title if snapshot is not None else card.title,
                    vendor_code=snapshot.vendor_code
                    if snapshot is not None
                    else card.vendor_code,
                    brand=snapshot.brand if snapshot is not None else card.brand,
                    subject_name=snapshot.subject_name
                    if snapshot is not None
                    else card.subject_name,
                    thumbnail_url=self._product_card_primary_photo(
                        card, snapshot=snapshot
                    ),
                    photos_count=int(
                        snapshot.photos_count
                        if snapshot is not None
                        else self._product_card_photo_count(card)
                    ),
                    video_count=int(
                        snapshot.video_count
                        if snapshot is not None
                        else (1 if card.video else 0)
                    ),
                    source_updated_at=(
                        snapshot.source_updated_at
                        if snapshot is not None
                        else (card.updated_at_wb or card.updated_at)
                    ),
                    updated_at=card.updated_at,
                    score=snapshot.score if snapshot is not None else None,
                    status=status,
                    analyzed_at=snapshot.analyzed_at if snapshot is not None else None,
                    source_revision=snapshot.source_revision
                    if snapshot is not None
                    else None,
                    issue_count=len(issues),
                    actionable_issue_count=len(issues),
                    critical_issue_count=critical_count,
                    warning_issue_count=warning_count,
                    ai_issue_count=ai_issue_count,
                    no_solution_ai_issue_count=no_solution_ai_count,
                    top_issue_title=top_issue.title if top_issue is not None else None,
                    top_issue_category=top_issue.category
                    if top_issue is not None
                    else None,
                    top_issue_severity=top_issue.severity
                    if top_issue is not None
                    else None,
                    top_issue_source=top_issue.source
                    if top_issue is not None
                    else None,
                    top_issue_recommended_fix=top_issue.recommended_fix
                    if top_issue is not None
                    else None,
                )
            )

        status_filter = str(quality_status or "").strip().lower()
        if status_filter and status_filter != "all":
            if status_filter == "issues":
                items = [item for item in items if item.issue_count > 0]
            elif status_filter in {"clean", "ok", "passed"}:
                items = [
                    item
                    for item in items
                    if item.status in {"ok", "clean"} and item.issue_count == 0
                ]
            elif status_filter in {"not_checked", "not_analyzed"}:
                items = [
                    item
                    for item in items
                    if item.analyzed_at is None or item.status == "not_analyzed"
                ]
            else:
                items = [
                    item
                    for item in items
                    if str(item.status or "").lower() == status_filter
                ]

        score_filter_key = str(score_filter or "").strip().lower()
        if score_filter_key and score_filter_key != "all":
            if score_filter_key in {"critical", "red"}:
                items = [
                    item for item in items if item.score is not None and item.score < 50
                ]
            elif score_filter_key in {"warning", "yellow"}:
                items = [
                    item
                    for item in items
                    if item.score is not None and 50 <= item.score < 75
                ]
            elif score_filter_key in {"good", "green"}:
                items = [
                    item
                    for item in items
                    if item.score is not None and item.score >= 75
                ]
            elif score_filter_key in {"none", "not_scored", "no_score"}:
                items = [item for item in items if item.score is None]

        ai_filter_key = str(ai_filter or "").strip().lower()
        if ai_filter_key and ai_filter_key != "all":
            if ai_filter_key == "has_ai":
                items = [item for item in items if item.ai_issue_count > 0]
            elif ai_filter_key == "no_ai":
                items = [item for item in items if item.ai_issue_count == 0]
            elif ai_filter_key == "hidden_no_fix":
                items = [item for item in items if item.no_solution_ai_issue_count > 0]

        media_filter_key = str(media_filter or "").strip().lower()
        if media_filter_key and media_filter_key != "all":
            if media_filter_key == "few_photos":
                items = [item for item in items if item.photos_count < 3]
            elif media_filter_key == "no_video":
                items = [item for item in items if item.video_count == 0]
            elif media_filter_key == "has_video":
                items = [item for item in items if item.video_count > 0]

        sort_key = str(sort_by or "quality_issues").strip().lower()
        descending = str(sort_dir or "desc").strip().lower() != "asc"

        def timestamp(value: datetime | None) -> float:
            return value.timestamp() if value is not None else 0.0

        if sort_key in {"quality_score", "score"}:
            items.sort(
                key=lambda item: (
                    item.score is not None,
                    item.score or -1,
                    -item.issue_count,
                ),
                reverse=descending,
            )
        elif sort_key in {"title", "name"}:
            items.sort(
                key=lambda item: str(item.title or item.vendor_code or "").lower(),
                reverse=descending,
            )
        elif sort_key == "status":
            status_rank = {
                "critical": 0,
                "warning": 1,
                "ok": 2,
                "clean": 2,
                "not_analyzed": 3,
                "empty": 4,
                "unavailable": 5,
            }
            items.sort(
                key=lambda item: status_rank.get(str(item.status or "").lower(), 9),
                reverse=descending,
            )
        elif sort_key in {"critical_issues", "critical"}:
            items.sort(
                key=lambda item: (item.critical_issue_count, item.issue_count),
                reverse=descending,
            )
        elif sort_key in {"ai_issues", "ai"}:
            items.sort(
                key=lambda item: (item.ai_issue_count, item.issue_count),
                reverse=descending,
            )
        elif sort_key in {"media", "photos"}:
            items.sort(
                key=lambda item: (item.photos_count, item.video_count),
                reverse=descending,
            )
        elif sort_key in {"analyzed_at", "checked_at"}:
            items.sort(key=lambda item: timestamp(item.analyzed_at), reverse=descending)
        elif sort_key in {"updated_at", "source_updated_at"}:
            items.sort(
                key=lambda item: timestamp(item.source_updated_at or item.updated_at),
                reverse=descending,
            )
        else:
            items.sort(
                key=lambda item: (
                    item.issue_count,
                    item.critical_issue_count,
                    -(item.score if item.score is not None else 101),
                    timestamp(item.analyzed_at),
                ),
                reverse=descending,
            )

        total = len(items)
        page_items = items[offset : offset + limit]
        analyzed = sum(1 for item in items if item.analyzed_at is not None)
        with_issues = sum(1 for item in items if item.issue_count > 0)
        clean = sum(
            1
            for item in items
            if item.analyzed_at is not None and item.issue_count == 0
        )
        critical_cards = sum(
            1
            for item in items
            if str(item.status or "").lower() == "critical"
            or item.critical_issue_count > 0
        )
        not_analyzed = total - analyzed
        avg_score_values = [item.score for item in items if item.score is not None]
        page = CardQualityProductsPage(
            total=total,
            limit=limit,
            offset=offset,
            items=page_items,
            summary={
                "total_cards": total,
                "analyzed_cards": analyzed,
                "cards_with_issues": with_issues,
                "critical_cards": critical_cards,
                "clean_cards": clean,
                "not_analyzed_cards": not_analyzed,
                "open_issue_count": sum(item.issue_count for item in items),
                "critical_issue_count": sum(
                    item.critical_issue_count for item in items
                ),
                "average_score": round(sum(avg_score_values) / len(avg_score_values), 1)
                if avg_score_values
                else None,
            },
        )
        self._product_cards_cache.set(cache_key, page)
        return page.model_copy(deep=True)

    def _should_surface_checker_issue(self, issue: CardQualityIssue) -> bool:
        if str(issue.status or "").lower() in {"done", "resolved", "ignored"}:
            return False
        if str(issue.source or "").lower() != "ai":
            return True
        return self._issue_has_any_fix_path(issue)

    def _issue_has_any_fix_path(self, issue: CardQualityIssue) -> bool:
        if str(issue.suggested_value or issue.ai_suggested_value or "").strip():
            return True
        if str(issue.recommended_fix or "").strip():
            return True
        if issue.alternatives_json or issue.ai_alternatives_json:
            return True
        for detail in issue.error_details_json or []:
            if not isinstance(detail, dict):
                continue
            if detail.get("fix_action") in {"clear", "swap", "compound"}:
                return True
            if (
                detail.get("type") == "compound"
                and isinstance(detail.get("fixes"), list)
                and detail["fixes"]
            ):
                return True
        return False

    def _product_card_photo_count(self, card: WBProductCard) -> int:
        photos = card.photos
        if isinstance(photos, list):
            return len(photos)
        if isinstance(photos, dict):
            nested = photos.get("photos") or photos.get("data") or photos.get("items")
            if isinstance(nested, list):
                return len(nested)
            return 1 if photos else 0
        return 0

    def _product_card_primary_photo(
        self, card: WBProductCard, *, snapshot: CardQualitySnapshot | None = None
    ) -> str | None:
        if snapshot is not None:
            media = snapshot.media_json if isinstance(snapshot.media_json, dict) else {}
            photos = media.get("photos")
            if isinstance(photos, list):
                found = self._snapshot_primary_photo(photos)
                if found:
                    return found
        raw_photos = card.photos
        candidates: list[Any] = raw_photos if isinstance(raw_photos, list) else []
        if isinstance(raw_photos, dict):
            nested = (
                raw_photos.get("photos")
                or raw_photos.get("data")
                or raw_photos.get("items")
            )
            candidates = nested if isinstance(nested, list) else [raw_photos]
        return self._snapshot_primary_photo(candidates)

    async def list_runs(
        self, session: AsyncSession, *, account_id: int, limit: int, offset: int
    ) -> CardQualityRunsPage:
        total = int(
            (
                await session.execute(
                    select(func.count())
                    .select_from(CardQualityAnalysisRun)
                    .where(CardQualityAnalysisRun.account_id == account_id)
                )
            ).scalar()
            or 0
        )
        rows = list(
            (
                await session.execute(
                    select(CardQualityAnalysisRun)
                    .where(CardQualityAnalysisRun.account_id == account_id)
                    .order_by(
                        CardQualityAnalysisRun.started_at.desc().nullslast(),
                        CardQualityAnalysisRun.id.desc(),
                    )
                    .limit(limit)
                    .offset(offset)
                )
            ).scalars()
        )
        return CardQualityRunsPage(
            total=total,
            limit=limit,
            offset=offset,
            items=[
                CardQualityRunRead.model_validate(row, from_attributes=True)
                for row in rows
            ],
        )

    async def get_run(
        self, session: AsyncSession, *, account_id: int, run_id: int
    ) -> CardQualityRunRead:
        run = await session.get(CardQualityAnalysisRun, run_id)
        if run is None or int(run.account_id) != int(account_id):
            raise ValueError("run_not_found")
        return CardQualityRunRead.model_validate(run, from_attributes=True)

    async def list_issues(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        category: str | None,
        status: str | None,
        include_info: bool = False,
        limit: int,
        offset: int,
    ) -> CardQualityIssuesPage:
        stmt = select(CardQualityIssue).where(CardQualityIssue.account_id == account_id)
        count_stmt = (
            select(func.count())
            .select_from(CardQualityIssue)
            .where(CardQualityIssue.account_id == account_id)
        )
        if not include_info:
            stmt = stmt.where(CardQualityIssue.severity != "info")
            count_stmt = count_stmt.where(CardQualityIssue.severity != "info")
        if category:
            stmt = stmt.where(CardQualityIssue.category == category)
            count_stmt = count_stmt.where(CardQualityIssue.category == category)
        if status:
            stmt = stmt.where(CardQualityIssue.status == status)
            count_stmt = count_stmt.where(CardQualityIssue.status == status)
        total = int((await session.execute(count_stmt)).scalar() or 0)
        rows_all = list((await session.execute(stmt)).scalars())
        rows = self._sort_issues_source_order(rows_all)[offset : offset + limit]
        for row in rows:
            self._ensure_issue_defaults(row)
        summary: dict[str, int] = {}
        for row in rows:
            summary[row.category] = summary.get(row.category, 0) + 1
        info_count = int(
            (
                await session.execute(
                    select(func.count())
                    .select_from(CardQualityIssue)
                    .where(
                        CardQualityIssue.account_id == account_id,
                        CardQualityIssue.severity == "info",
                    )
                )
            ).scalar()
            or 0
        )
        actionable_count = int(
            (
                await session.execute(
                    select(func.count())
                    .select_from(CardQualityIssue)
                    .where(
                        CardQualityIssue.account_id == account_id,
                        CardQualityIssue.severity != "info",
                        CardQualityIssue.status.in_(tuple(ACTIVE_ISSUE_STATUSES)),
                        CardQualityIssue.resolved_at.is_(None),
                    )
                )
            ).scalar()
            or 0
        )
        return CardQualityIssuesPage(
            total=total,
            limit=limit,
            offset=offset,
            items=[
                CardQualityIssueRead.model_validate(row, from_attributes=True)
                for row in rows
            ],
            summary={
                "by_category": summary,
                "actionable_open_issues": actionable_count,
                "informational_observations": info_count,
            },
        )

    async def list_issues_grouped(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        bucket: str = "actionable",
        limit: int = 200,
    ) -> CardQualityIssuesGrouped:
        rows = await self._queue_issue_rows(
            session, account_id=account_id, bucket=bucket
        )
        cards = await self._cards_by_nm_id(
            session, account_id=account_id, nm_ids=[int(row.nm_id) for row in rows]
        )
        grouped: dict[str, list[dict[str, Any]]] = {
            "critical": [],
            "warnings": [],
            "media": [],
            "postponed": [],
        }
        for issue in rows:
            payload = self._issue_with_card_payload(
                issue, card=cards.get(int(issue.nm_id)), pending_count=None
            )
            if issue.status == "postponed":
                grouped["postponed"].append(payload)
            elif self.issue_belongs_to_bucket(issue, "media"):
                grouped["media"].append(payload)
            elif str(issue.severity or "").lower() in {"critical", "high"}:
                grouped["critical"].append(payload)
            else:
                grouped["warnings"].append(payload)

        limited = {key: value[:limit] for key, value in grouped.items()}
        return CardQualityIssuesGrouped(
            bucket=bucket,
            critical=limited["critical"],
            warnings=limited["warnings"],
            media=limited["media"],
            postponed=limited["postponed"],
            totals={key: len(value) for key, value in grouped.items()},
            critical_count=len(grouped["critical"]),
            warnings_count=len(grouped["warnings"]),
            media_count=len(grouped["media"]),
            postponed_count=len(grouped["postponed"]),
        )

    async def get_next_issue(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        after_issue_id: int | None = None,
        nm_id: int | None = None,
        severity: str | None = None,
        bucket: str = "actionable",
    ) -> dict[str, Any] | None:
        rows = await self._queue_issue_rows(
            session,
            account_id=account_id,
            bucket=bucket,
            nm_id=nm_id,
            severity=severity,
        )
        if not rows:
            return None
        selected: CardQualityIssue | None = None
        if after_issue_id is None:
            selected = rows[0]
        else:
            ordered_ids = [int(row.id) for row in rows]
            try:
                next_index = ordered_ids.index(int(after_issue_id)) + 1
            except ValueError:
                next_index = 0
            if next_index < len(rows):
                selected = rows[next_index]
        if selected is None:
            return None
        card = (
            await self._cards_by_nm_id(
                session, account_id=account_id, nm_ids=[int(selected.nm_id)]
            )
        ).get(int(selected.nm_id))
        pending_count = sum(1 for row in rows if int(row.nm_id) == int(selected.nm_id))
        return self._issue_with_card_payload(
            selected, card=card, pending_count=pending_count
        )

    async def get_queue_progress(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        severity: str | None = None,
        bucket: str = "actionable",
    ) -> CardQualityQueueProgress:
        cache_key = (
            "queue_progress_v1",
            int(account_id),
            str(severity or "").strip().lower(),
            str(bucket or "actionable").strip().lower(),
        )
        cached = self._queue_progress_cache.get(cache_key)
        if cached is not None:
            return cached.model_copy(deep=True)

        count_rows = (
            await session.execute(
                select(CardQualityIssue.status, func.count(CardQualityIssue.id))
                .where(
                    *self._queue_issue_count_filters(
                        account_id=account_id,
                        bucket=bucket,
                        severity=severity,
                        include_done=True,
                    )
                )
                .group_by(CardQualityIssue.status)
            )
        ).all()
        counts = {
            str(status or "").strip().lower(): int(count or 0)
            for status, count in count_rows
        }
        pending = sum(
            counts.get(status, 0) for status in ("new", "in_progress", "blocked")
        )
        fixed = sum(counts.get(status, 0) for status in ("done", "resolved"))
        skipped = counts.get("ignored", 0)
        postponed = counts.get("postponed", 0)
        total = pending + fixed + skipped + postponed
        result = CardQualityQueueProgress(
            total=total,
            pending=pending,
            fixed=fixed,
            skipped=skipped,
            postponed=postponed,
            progress_percent=round(((fixed + skipped) / total) * 100, 1)
            if total
            else 0.0,
        )
        self._queue_progress_cache.set(cache_key, result)
        return result.model_copy(deep=True)

    async def fixed_file_status(
        self, session: AsyncSession, *, account_id: int
    ) -> CardQualityFixedFileStatus:
        total = int(
            (
                await session.execute(
                    select(func.count())
                    .select_from(CardQualityFixedFileEntry)
                    .where(CardQualityFixedFileEntry.account_id == account_id)
                )
            ).scalar()
            or 0
        )
        total_cards = int(
            (
                await session.execute(
                    select(
                        func.count(func.distinct(CardQualityFixedFileEntry.nm_id))
                    ).where(CardQualityFixedFileEntry.account_id == account_id)
                )
            ).scalar()
            or 0
        )
        total_brands = int(
            (
                await session.execute(
                    select(
                        func.count(func.distinct(CardQualityFixedFileEntry.brand))
                    ).where(
                        CardQualityFixedFileEntry.account_id == account_id,
                        CardQualityFixedFileEntry.brand.is_not(None),
                        CardQualityFixedFileEntry.brand != "",
                    )
                )
            ).scalar()
            or 0
        )
        total_subjects = int(
            (
                await session.execute(
                    select(
                        func.count(
                            func.distinct(CardQualityFixedFileEntry.subject_name)
                        )
                    ).where(
                        CardQualityFixedFileEntry.account_id == account_id,
                        CardQualityFixedFileEntry.subject_name.is_not(None),
                        CardQualityFixedFileEntry.subject_name != "",
                    )
                )
            ).scalar()
            or 0
        )
        total_characteristics = int(
            (
                await session.execute(
                    select(
                        func.count(func.distinct(CardQualityFixedFileEntry.char_name))
                    ).where(CardQualityFixedFileEntry.account_id == account_id)
                )
            ).scalar()
            or 0
        )
        last_updated_at = (
            await session.execute(
                select(func.max(CardQualityFixedFileEntry.updated_at)).where(
                    CardQualityFixedFileEntry.account_id == account_id
                )
            )
        ).scalar()
        return CardQualityFixedFileStatus(
            has_fixed_file=total > 0,
            total=total,
            total_cards=total_cards,
            total_brands=total_brands,
            total_subjects=total_subjects,
            total_characteristics=total_characteristics,
            last_updated_at=last_updated_at,
        )

    def _fixed_file_filters(
        self,
        *,
        account_id: int,
        nm_id: int | None = None,
        search: str | None = None,
        brand: str | None = None,
        subject_name: str | None = None,
        char_name: str | None = None,
    ) -> list[Any]:
        filters: list[Any] = [CardQualityFixedFileEntry.account_id == account_id]
        if nm_id is not None:
            filters.append(CardQualityFixedFileEntry.nm_id == nm_id)
        if brand and brand.strip():
            filters.append(CardQualityFixedFileEntry.brand.ilike(f"%{brand.strip()}%"))
        if subject_name and subject_name.strip():
            filters.append(
                CardQualityFixedFileEntry.subject_name.ilike(
                    f"%{subject_name.strip()}%"
                )
            )
        if char_name and char_name.strip():
            filters.append(
                CardQualityFixedFileEntry.char_name.ilike(f"%{char_name.strip()}%")
            )
        if search and search.strip():
            needle = search.strip()
            search_filters: list[Any] = [
                CardQualityFixedFileEntry.brand.ilike(f"%{needle}%"),
                CardQualityFixedFileEntry.subject_name.ilike(f"%{needle}%"),
                CardQualityFixedFileEntry.char_name.ilike(f"%{needle}%"),
                CardQualityFixedFileEntry.fixed_value.ilike(f"%{needle}%"),
                cast(CardQualityFixedFileEntry.nm_id, String).ilike(f"%{needle}%"),
            ]
            if needle.isdigit():
                search_filters.append(CardQualityFixedFileEntry.nm_id == int(needle))
            filters.append(or_(*search_filters))
        return filters

    def _fixed_file_order(self, sort_by: str | None, sort_dir: str | None) -> list[Any]:
        sort_map = {
            "nm_id": CardQualityFixedFileEntry.nm_id,
            "brand": CardQualityFixedFileEntry.brand,
            "subject_name": CardQualityFixedFileEntry.subject_name,
            "char_name": CardQualityFixedFileEntry.char_name,
            "fixed_value": CardQualityFixedFileEntry.fixed_value,
            "updated_at": CardQualityFixedFileEntry.updated_at,
            "created_at": CardQualityFixedFileEntry.created_at,
        }
        primary = sort_map.get(str(sort_by or "nm_id"), CardQualityFixedFileEntry.nm_id)
        ordered = (
            primary.desc()
            if str(sort_dir or "asc").lower() == "desc"
            else primary.asc()
        )
        return [
            ordered,
            CardQualityFixedFileEntry.nm_id.asc(),
            CardQualityFixedFileEntry.char_name.asc(),
        ]

    async def list_fixed_file_entries(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        nm_id: int | None = None,
        search: str | None = None,
        brand: str | None = None,
        subject_name: str | None = None,
        char_name: str | None = None,
        sort_by: str = "nm_id",
        sort_dir: str = "asc",
        limit: int = 100,
        offset: int = 0,
    ) -> CardQualityFixedFileEntriesPage:
        filters = self._fixed_file_filters(
            account_id=account_id,
            nm_id=nm_id,
            search=search,
            brand=brand,
            subject_name=subject_name,
            char_name=char_name,
        )
        stmt = select(CardQualityFixedFileEntry).where(*filters)
        count_stmt = (
            select(func.count()).select_from(CardQualityFixedFileEntry).where(*filters)
        )
        total = int((await session.execute(count_stmt)).scalar() or 0)
        rows = list(
            (
                await session.execute(
                    stmt.order_by(*self._fixed_file_order(sort_by, sort_dir))
                    .limit(limit)
                    .offset(offset)
                )
            ).scalars()
        )
        return CardQualityFixedFileEntriesPage(
            total=total,
            limit=limit,
            offset=offset,
            summary=await self.fixed_file_status(session, account_id=account_id),
            items=[
                CardQualityFixedFileEntryRead.model_validate(row, from_attributes=True)
                for row in rows
            ],
        )

    async def create_fixed_file_entry(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        payload: CardQualityFixedFileEntryMutation,
        created_by_user_id: int | None = None,
    ) -> CardQualityFixedFileEntryRead:
        if payload.nm_id is None or int(payload.nm_id) <= 0:
            raise ValueError("fixed_file_nm_id_required")
        char_name = str(payload.char_name or "").strip()
        fixed_value = str(payload.fixed_value or "").strip()
        if not char_name:
            raise ValueError("fixed_file_char_name_required")
        if not fixed_value:
            raise ValueError("fixed_file_value_required")
        row = CardQualityFixedFileEntry(
            account_id=account_id,
            nm_id=int(payload.nm_id),
            brand=str(payload.brand or "").strip() or None,
            subject_name=str(payload.subject_name or "").strip() or None,
            char_name=char_name,
            fixed_value=fixed_value,
            created_by_user_id=created_by_user_id,
        )
        session.add(row)
        try:
            await session.commit()
        except IntegrityError as exc:
            await session.rollback()
            raise ValueError("fixed_file_entry_duplicate") from exc
        await session.refresh(row)
        return CardQualityFixedFileEntryRead.model_validate(row, from_attributes=True)

    async def update_fixed_file_entry(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        entry_id: int,
        payload: CardQualityFixedFileEntryMutation,
        updated_by_user_id: int | None = None,
    ) -> CardQualityFixedFileEntryRead:
        row = (
            await session.execute(
                select(CardQualityFixedFileEntry).where(
                    CardQualityFixedFileEntry.account_id == account_id,
                    CardQualityFixedFileEntry.id == entry_id,
                )
            )
        ).scalar_one_or_none()
        if row is None:
            raise ValueError("fixed_file_entry_not_found")
        data = payload.model_dump(exclude_unset=True)
        if "nm_id" in data and data["nm_id"] is not None:
            nm_id = int(data["nm_id"])
            if nm_id <= 0:
                raise ValueError("fixed_file_nm_id_required")
            row.nm_id = nm_id
        if "brand" in data:
            row.brand = str(data.get("brand") or "").strip() or None
        if "subject_name" in data:
            row.subject_name = str(data.get("subject_name") or "").strip() or None
        if "char_name" in data:
            char_name = str(data.get("char_name") or "").strip()
            if not char_name:
                raise ValueError("fixed_file_char_name_required")
            row.char_name = char_name
        if "fixed_value" in data:
            fixed_value = str(data.get("fixed_value") or "").strip()
            if not fixed_value:
                raise ValueError("fixed_file_value_required")
            row.fixed_value = fixed_value
        if updated_by_user_id is not None:
            row.created_by_user_id = updated_by_user_id
        try:
            await session.commit()
        except IntegrityError as exc:
            await session.rollback()
            raise ValueError("fixed_file_entry_duplicate") from exc
        await session.refresh(row)
        return CardQualityFixedFileEntryRead.model_validate(row, from_attributes=True)

    async def delete_fixed_file_entry(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        entry_id: int,
    ) -> dict[str, Any]:
        result = await session.execute(
            delete(CardQualityFixedFileEntry).where(
                CardQualityFixedFileEntry.account_id == account_id,
                CardQualityFixedFileEntry.id == entry_id,
            )
        )
        deleted = int(result.rowcount or 0)
        await session.commit()
        if deleted <= 0:
            raise ValueError("fixed_file_entry_not_found")
        return {"status": "ok", "deleted": deleted}

    async def clear_fixed_file_entries(
        self, session: AsyncSession, *, account_id: int
    ) -> dict[str, Any]:
        result = await session.execute(
            delete(CardQualityFixedFileEntry).where(
                CardQualityFixedFileEntry.account_id == account_id
            )
        )
        deleted = int(result.rowcount or 0)
        await session.commit()
        return {"status": "ok", "deleted": deleted}

    async def export_fixed_file_entries(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        nm_id: int | None = None,
        search: str | None = None,
        brand: str | None = None,
        subject_name: str | None = None,
        char_name: str | None = None,
        sort_by: str = "nm_id",
        sort_dir: str = "asc",
    ) -> bytes:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except Exception as exc:  # pragma: no cover - dependency guard
            raise ValueError("openpyxl_unavailable") from exc
        filters = self._fixed_file_filters(
            account_id=account_id,
            nm_id=nm_id,
            search=search,
            brand=brand,
            subject_name=subject_name,
            char_name=char_name,
        )
        rows = list(
            (
                await session.execute(
                    select(CardQualityFixedFileEntry)
                    .where(*filters)
                    .order_by(*self._fixed_file_order(sort_by, sort_dir))
                )
            ).scalars()
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Fixed file"
        headers = [
            "nmID",
            "brand",
            "subjectName",
            "Характеристика",
            "Эталонное значение",
        ]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="08786F")
        sheet.freeze_panes = "A2"
        for row in rows:
            sheet.append(
                [
                    row.nm_id,
                    row.brand or "",
                    row.subject_name or "",
                    row.char_name,
                    row.fixed_value,
                ]
            )
        widths = [14, 22, 24, 32, 56]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    async def upload_fixed_file(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        content: bytes,
        filename: str | None,
        replace_all: bool = False,
        created_by_user_id: int | None = None,
    ) -> CardQualityFixedFileUploadResponse:
        if not filename or not filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
            raise ValueError("fixed_file_requires_excel")
        entries = self._parse_fixed_file_excel(content)
        if not entries:
            raise ValueError("fixed_file_empty")
        if replace_all:
            await session.execute(
                delete(CardQualityFixedFileEntry).where(
                    CardQualityFixedFileEntry.account_id == account_id
                )
            )
        deduped_entries = {
            (int(entry["nm_id"]), str(entry["char_name"]).strip().lower()): entry
            for entry in entries
            if entry.get("nm_id") is not None
            and str(entry.get("char_name") or "").strip()
        }
        entries = list(deduped_entries.values())
        existing_total = int(
            (
                await session.execute(
                    select(func.count())
                    .select_from(CardQualityFixedFileEntry)
                    .where(CardQualityFixedFileEntry.account_id == account_id)
                )
            ).scalar()
            or 0
        )
        if existing_total == 0:
            for entry in entries:
                session.add(
                    CardQualityFixedFileEntry(
                        account_id=account_id,
                        nm_id=entry["nm_id"],
                        brand=entry.get("brand"),
                        subject_name=entry.get("subject_name"),
                        char_name=entry["char_name"],
                        fixed_value=entry["fixed_value"],
                        created_by_user_id=created_by_user_id,
                    )
                )
            await session.commit()
            return CardQualityFixedFileUploadResponse(
                upserted=len(entries),
                total=len(entries),
                message=f"Загружено {len(entries)} эталонных значений",
            )

        existing_by_key: dict[tuple[int, str], CardQualityFixedFileEntry] = {}
        nm_ids = sorted({int(entry["nm_id"]) for entry in entries})
        chunk_size = 1000
        for start in range(0, len(nm_ids), chunk_size):
            chunk = nm_ids[start : start + chunk_size]
            rows = (
                await session.execute(
                    select(CardQualityFixedFileEntry).where(
                        CardQualityFixedFileEntry.account_id == account_id,
                        CardQualityFixedFileEntry.nm_id.in_(chunk),
                    )
                )
            ).scalars()
            for row in rows:
                existing_by_key[
                    (int(row.nm_id), str(row.char_name).strip().lower())
                ] = row
        upserted = 0
        for entry in entries:
            existing = existing_by_key.get(
                (int(entry["nm_id"]), str(entry["char_name"]).strip().lower())
            )
            if existing is None:
                session.add(
                    CardQualityFixedFileEntry(
                        account_id=account_id,
                        nm_id=entry["nm_id"],
                        brand=entry.get("brand"),
                        subject_name=entry.get("subject_name"),
                        char_name=entry["char_name"],
                        fixed_value=entry["fixed_value"],
                        created_by_user_id=created_by_user_id,
                    )
                )
            else:
                existing.brand = entry.get("brand") or existing.brand
                existing.subject_name = (
                    entry.get("subject_name") or existing.subject_name
                )
                existing.fixed_value = entry["fixed_value"]
                if created_by_user_id is not None:
                    existing.created_by_user_id = created_by_user_id
            upserted += 1
        await session.commit()
        total_after = int(
            (
                await session.execute(
                    select(func.count())
                    .select_from(CardQualityFixedFileEntry)
                    .where(CardQualityFixedFileEntry.account_id == account_id)
                )
            ).scalar()
            or 0
        )
        return CardQualityFixedFileUploadResponse(
            upserted=upserted,
            total=total_after,
            message=f"Загружено {upserted} эталонных значений",
        )

    async def preview_issue_apply(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        issue_id: int,
        fixed_value: str | None = None,
    ) -> CardQualityIssueApplyPreview:
        issue = await self._issue_for_mutation(
            session, account_id=account_id, issue_id=issue_id
        )
        value = self._effective_fixed_value(
            issue, fixed_value=fixed_value, require_manual_for_human_check=False
        )
        blocked_reason = self._direct_apply_blocked_reason(issue, value)
        return CardQualityIssueApplyPreview(
            issue_id=int(issue.id),
            nm_id=int(issue.nm_id),
            field_path=issue.field_name,
            current_value=issue.current_value_json,
            fixed_value=value,
            diff={
                "field_path": issue.field_name,
                "before": issue.current_value_json,
                "after": value,
                "source": issue.source,
                "requires_human_check": bool(issue.requires_human_check),
            },
            can_apply_to_wb=blocked_reason is None,
            requires_confirm=True,
            blocked_reason=blocked_reason,
            wb_write_status="blocked" if blocked_reason else "preview_ready",
            audit={
                "default_writes_to_wb": False,
                "confirm_required": True,
                "preview_diff_required": True,
                "content_token_permission_required": True,
                "audit_event_required": True,
                "follow_up_status_required": True,
                "source_behavior": "preview_diff_confirm_audit",
            },
        )

    async def fix_issue(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        issue_id: int,
        fixed_value: str | None,
        changed_by_user_id: int | None,
        apply_to_wb: bool = False,
        confirm: bool = False,
        reason: str | None = None,
    ) -> CardQualityIssueFixResponse:
        issue = await self._issue_for_mutation(
            session, account_id=account_id, issue_id=issue_id
        )
        value = self._effective_fixed_value(
            issue, fixed_value=fixed_value, require_manual_for_human_check=True
        )
        preview = await self.preview_issue_apply(
            session,
            account_id=account_id,
            issue_id=issue_id,
            fixed_value=value,
        )
        if apply_to_wb:
            if not confirm:
                return CardQualityIssueFixResponse(
                    status="confirmation_required",
                    issue=CardQualityIssueRead.model_validate(
                        issue, from_attributes=True
                    ),
                    preview=preview,
                    wb_write_status="confirmation_required",
                    message="WB apply requires explicit confirm=true after reviewing the preview diff.",
                )
            if not preview.can_apply_to_wb:
                return CardQualityIssueFixResponse(
                    status="blocked",
                    issue=CardQualityIssueRead.model_validate(
                        issue, from_attributes=True
                    ),
                    preview=preview,
                    wb_write_status="blocked",
                    message=preview.blocked_reason,
                )
            try:
                apply_result = await self._apply_issue_fix_to_wb(
                    session,
                    account_id=account_id,
                    issue=issue,
                    fixed_value=value,
                    actor_id=changed_by_user_id,
                )
            except ValueError as exc:
                detail = str(exc)
                blocked = detail.startswith("content_token_permission_required")
                return CardQualityIssueFixResponse(
                    status="blocked" if blocked else "wb_submit_failed",
                    issue=CardQualityIssueRead.model_validate(
                        issue, from_attributes=True
                    ),
                    preview=preview,
                    wb_write_status="blocked" if blocked else "failed",
                    message=detail,
                )
            updated = await self.update_issue_status(
                session,
                account_id=account_id,
                issue_id=issue_id,
                status="in_progress",
                changed_by_user_id=changed_by_user_id,
                reason=reason or "wb_submit_attempted_waiting_validation",
                fixed_value=value,
            )
            return CardQualityIssueFixResponse(
                status="submitted_to_wb",
                issue=updated,
                preview=preview,
                apply_result=apply_result,
                wb_write_status="submitted_waiting_validation",
                message="Fix submitted to WB after explicit confirmation. Waiting for WB validation/listing check.",
            )

        updated = await self.update_issue_status(
            session,
            account_id=account_id,
            issue_id=issue_id,
            status="done",
            changed_by_user_id=changed_by_user_id,
            reason=reason or "fixed_locally_no_wb_apply",
            fixed_value=value,
        )
        return CardQualityIssueFixResponse(
            status="fixed_local",
            issue=updated,
            preview=preview,
            wb_write_status="not_requested",
            message="Issue marked fixed locally. WB was not changed.",
        )

    async def _issue_for_mutation(
        self, session: AsyncSession, *, account_id: int, issue_id: int
    ) -> CardQualityIssue:
        issue = await session.get(CardQualityIssue, issue_id)
        if issue is None or int(issue.account_id) != int(account_id):
            raise ValueError("issue_not_found")
        self._ensure_issue_defaults(issue)
        return issue

    def _effective_fixed_value(
        self,
        issue: CardQualityIssue,
        *,
        fixed_value: str | None,
        require_manual_for_human_check: bool,
    ) -> str:
        value = str(fixed_value if fixed_value is not None else "").strip()
        if not value and not bool(issue.requires_human_check):
            value = str(issue.suggested_value or issue.ai_suggested_value or "").strip()
        if not value:
            if require_manual_for_human_check and bool(issue.requires_human_check):
                raise ValueError("human_check_issue_requires_manual_review")
            raise ValueError("fixed_value_required")
        self._validate_manual_fixed_value(issue, value)
        return value

    def _validate_manual_fixed_value(
        self, issue: CardQualityIssue, fixed_value: str
    ) -> None:
        if issue.source == "fixed_file":
            expected = str(
                issue.suggested_value or issue.ai_suggested_value or ""
            ).strip()
            if expected and fixed_value != expected:
                raise ValueError("fixed_file_requires_exact_suggested_value")
        if _field_key(issue.field_name) == "title" or issue.category == "title":
            card = {
                "title": issue.current_value_json,
                "subjectName": "",
                "characteristics": [],
            }
            keep_current, info = should_keep_current_title_as_safer(
                str(issue.current_value_json or ""), fixed_value, card
            )
            if keep_current:
                raise ValueError(
                    f"title_business_guard:{info.get('reason') or 'unsafe_candidate'}"
                )
        if (
            _field_key(issue.field_name) == "description"
            or issue.category == "description"
        ):
            valid, reason_text = validate_description_facts(
                fixed_value,
                {"description": issue.current_value_json},
                allow_visual_facts=False,
            )
            if not valid:
                raise ValueError(f"description_factual_guard:{reason_text}")
        if issue.allowed_values_json:
            valid, _, reason_text = self._validate_allowed_value(
                fixed_value,
                issue.allowed_values_json or [],
                issue.error_details_json or [],
                field_name=issue.field_name,
            )
            if not valid:
                raise ValueError(reason_text or "value_not_allowed")

    def _direct_apply_blocked_reason(
        self, issue: CardQualityIssue, fixed_value: str
    ) -> str | None:
        if not fixed_value:
            return "fixed_value_required"
        capabilities = checker_action_capabilities(issue)
        semantic_block = capabilities.get("apply_wb_disabled_reason")
        if semantic_block and semantic_block != "fixed_value_required":
            return str(semantic_block)
        canonical = _canonical_apply_field_path(issue.field_name, issue.category)
        if canonical in {"title", "description"} or canonical.startswith(
            "characteristics."
        ):
            if str(issue.category or "").lower() in {
                "media",
                "photos",
                "photo",
                "video",
            }:
                return "media_issues_require_dedicated_media_flow"
            return None
        return f"unsupported_wb_apply_field:{issue.field_name or issue.category or issue.issue_code}"

    async def accept_issue_local(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        issue_id: int,
        fixed_value: str | None,
        changed_by_user_id: int | None,
        reason: str | None = None,
    ) -> CardQualityIssueFixResponse:
        return await self.fix_issue(
            session,
            account_id=account_id,
            issue_id=issue_id,
            fixed_value=fixed_value,
            changed_by_user_id=changed_by_user_id,
            apply_to_wb=False,
            confirm=False,
            reason=reason or "accepted_locally_no_wb_write",
        )

    async def mark_issue_fixed(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        issue_id: int,
        fixed_value: str | None,
        changed_by_user_id: int | None,
        reason: str | None = None,
    ) -> CardQualityIssueRead:
        issue = await self._issue_for_mutation(
            session, account_id=account_id, issue_id=issue_id
        )
        capabilities = checker_action_capabilities(issue)
        if not capabilities["can_mark_fixed"]:
            raise ValueError(
                str(capabilities["mark_fixed_disabled_reason"] or "mark_fixed_disabled")
            )
        return await self.update_issue_status(
            session,
            account_id=account_id,
            issue_id=issue_id,
            status="done",
            changed_by_user_id=changed_by_user_id,
            reason=reason or "marked_fixed_no_wb_write",
            fixed_value=fixed_value,
        )

    async def save_issue_draft(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        issue_id: int,
        fixed_value: str | None,
        changed_by_user_id: int | None,
        reason: str | None = None,
    ) -> CardQualityIssueRead:
        issue = await self._issue_for_mutation(
            session, account_id=account_id, issue_id=issue_id
        )
        capabilities = checker_action_capabilities(issue)
        if not capabilities["can_save_draft"]:
            raise ValueError(
                str(capabilities["save_draft_disabled_reason"] or "save_draft_disabled")
            )
        value = self._effective_fixed_value(
            issue, fixed_value=fixed_value, require_manual_for_human_check=False
        )
        return await self.update_issue_status(
            session,
            account_id=account_id,
            issue_id=issue_id,
            status="in_progress",
            changed_by_user_id=changed_by_user_id,
            reason=reason or "draft_saved_no_wb_write",
            fixed_value=value,
        )

    async def recheck_issue(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        issue_id: int,
        requested_by_user_id: int | None = None,
    ) -> CardQualityIssueRead:
        issue = await self._issue_for_mutation(
            session, account_id=account_id, issue_id=issue_id
        )
        capabilities = checker_action_capabilities(issue)
        if not capabilities["can_recheck"]:
            raise ValueError(
                str(capabilities["recheck_disabled_reason"] or "recheck_disabled")
            )
        nm_id = int(issue.nm_id)
        await self.analyze_product(
            session,
            account_id=account_id,
            nm_id=nm_id,
            force=True,
            requested_by_user_id=requested_by_user_id,
        )
        refreshed = await session.get(CardQualityIssue, issue_id)
        if refreshed is None or int(refreshed.account_id) != int(account_id):
            raise ValueError("issue_not_found")
        self._ensure_issue_defaults(refreshed)
        return CardQualityIssueRead.model_validate(refreshed, from_attributes=True)

    async def recheck_product(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        nm_id: int,
        requested_by_user_id: int | None = None,
    ) -> CardQualityProductRecheckResponse:
        previous_snapshot = await self.latest_snapshot(
            session, account_id=account_id, nm_id=nm_id
        )
        previous_issues = await self._all_issues_for_product(
            session, account_id=account_id, nm_id=nm_id
        )
        previous_open_ids = self._active_issue_ids(previous_issues)
        previous_status_by_id = {
            int(issue.id): str(issue.status or "")
            for issue in previous_issues
            if issue.id is not None
        }
        previous_score = (
            int(previous_snapshot.score)
            if previous_snapshot is not None and previous_snapshot.score is not None
            else None
        )
        previous_open_count = len(previous_open_ids)

        try:
            product_quality = await self.analyze_product(
                session,
                account_id=account_id,
                nm_id=nm_id,
                force=True,
                requested_by_user_id=requested_by_user_id,
            )
            status: str = "completed"
            message = product_quality.message
        except Exception as exc:
            latest_run = await self._latest_product_run(
                session, account_id=account_id, nm_id=nm_id
            )
            return CardQualityProductRecheckResponse(
                run_id=int(latest_run.id)
                if latest_run is not None and latest_run.id is not None
                else None,
                nm_id=nm_id,
                status="failed",
                previous_score=previous_score,
                previous_open_issue_count=previous_open_count,
                result_status="not_enough_data",
                message=f"Card quality re-check failed: {exc.__class__.__name__}",
            )

        latest_run = await self._latest_product_run(
            session, account_id=account_id, nm_id=nm_id
        )
        new_snapshot = await self.latest_snapshot(
            session, account_id=account_id, nm_id=nm_id
        )
        new_issues = await self._all_issues_for_product(
            session, account_id=account_id, nm_id=nm_id
        )
        new_open_ids = self._active_issue_ids(new_issues)
        new_score = (
            int(new_snapshot.score)
            if new_snapshot is not None and new_snapshot.score is not None
            else None
        )
        new_open_count = len(new_open_ids) if new_snapshot is not None else None
        resolved_issue_ids = sorted(previous_open_ids - new_open_ids)
        reopened_issue_ids = sorted(
            int(issue.id)
            for issue in new_issues
            if issue.id is not None
            and int(issue.id) in previous_status_by_id
            and previous_status_by_id[int(issue.id)] in {"done", "ignored", "resolved"}
            and self._issue_is_active(issue)
        )
        result_status = self._product_recheck_result_status(
            previous_score=previous_score,
            new_score=new_score,
            previous_open_issue_count=previous_open_count,
            new_open_issue_count=new_open_count,
        )
        action_center_updates = self._checker_recheck_action_updates(
            resolved_issue_ids=resolved_issue_ids,
            reopened_issue_ids=reopened_issue_ids,
            nm_id=nm_id,
        )
        result_event_id = await self._record_product_recheck_event(
            session,
            account_id=account_id,
            nm_id=nm_id,
            run_id=int(latest_run.id)
            if latest_run is not None and latest_run.id is not None
            else None,
            previous_score=previous_score,
            new_score=new_score,
            previous_open_issue_count=previous_open_count,
            new_open_issue_count=new_open_count,
            resolved_issue_ids=resolved_issue_ids,
            reopened_issue_ids=reopened_issue_ids,
            result_status=result_status,
            created_by=requested_by_user_id,
            message=message,
        )
        if hasattr(session, "commit"):
            await session.commit()
        return CardQualityProductRecheckResponse(
            run_id=int(latest_run.id)
            if latest_run is not None and latest_run.id is not None
            else None,
            nm_id=nm_id,
            status=status,  # type: ignore[arg-type]
            previous_score=previous_score,
            new_score=new_score,
            previous_open_issue_count=previous_open_count,
            new_open_issue_count=new_open_count,
            resolved_issue_ids=resolved_issue_ids,
            reopened_issue_ids=reopened_issue_ids,
            result_status=result_status,  # type: ignore[arg-type]
            result_event_id=result_event_id,
            action_center_updates=action_center_updates,
            message=message or "Card quality re-check completed.",
        )

    async def _all_issues_for_product(
        self, session: AsyncSession, *, account_id: int, nm_id: int
    ) -> list[CardQualityIssue]:
        rows = list(
            (
                await session.execute(
                    select(CardQualityIssue).where(
                        CardQualityIssue.account_id == account_id,
                        CardQualityIssue.nm_id == nm_id,
                    )
                )
            ).scalars()
        )
        for row in rows:
            self._ensure_issue_defaults(row)
        return rows

    async def _latest_product_run(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        nm_id: int,
    ) -> CardQualityAnalysisRun | None:
        return (
            await session.execute(
                select(CardQualityAnalysisRun)
                .where(
                    CardQualityAnalysisRun.account_id == account_id,
                    CardQualityAnalysisRun.run_type == "single_product",
                    CardQualityAnalysisRun.last_processed_key == str(nm_id),
                )
                .order_by(
                    CardQualityAnalysisRun.started_at.desc().nullslast(),
                    CardQualityAnalysisRun.id.desc(),
                )
                .limit(1)
            )
        ).scalar_one_or_none()

    def _issue_is_active(self, issue: CardQualityIssue) -> bool:
        return (
            str(issue.status or "") in ACTIVE_ISSUE_STATUSES
            and issue.resolved_at is None
        )

    def _active_issue_ids(self, issues: list[CardQualityIssue]) -> set[int]:
        return {
            int(issue.id)
            for issue in issues
            if issue.id is not None and self._issue_is_active(issue)
        }

    def _product_recheck_result_status(
        self,
        *,
        previous_score: int | None,
        new_score: int | None,
        previous_open_issue_count: int,
        new_open_issue_count: int | None,
    ) -> str:
        if new_score is None or new_open_issue_count is None:
            return "pending_data"
        if previous_score is None:
            return "not_enough_data"
        score_delta = int(new_score) - int(previous_score)
        open_delta = int(new_open_issue_count) - int(previous_open_issue_count)
        if score_delta > 0 or open_delta < 0:
            return "improved"
        if score_delta < 0 or open_delta > 0:
            return "worse"
        return "neutral"

    def _checker_recheck_action_updates(
        self,
        *,
        resolved_issue_ids: list[int],
        reopened_issue_ids: list[int],
        nm_id: int,
    ) -> list[dict[str, Any]]:
        updates: list[dict[str, Any]] = []
        for issue_id in resolved_issue_ids:
            updates.append(
                {
                    "source_module": "checker",
                    "source": "card_quality_issues",
                    "source_id": str(issue_id),
                    "issue_id": issue_id,
                    "nm_id": nm_id,
                    "status": "done",
                    "result_badge": "resolved_after_recheck",
                }
            )
        for issue_id in reopened_issue_ids:
            updates.append(
                {
                    "source_module": "checker",
                    "source": "card_quality_issues",
                    "source_id": str(issue_id),
                    "issue_id": issue_id,
                    "nm_id": nm_id,
                    "status": "reopened",
                    "result_badge": "still_open_after_recheck",
                }
            )
        return updates

    async def _record_product_recheck_event(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        nm_id: int,
        run_id: int | None,
        previous_score: int | None,
        new_score: int | None,
        previous_open_issue_count: int,
        new_open_issue_count: int | None,
        resolved_issue_ids: list[int],
        reopened_issue_ids: list[int],
        result_status: str,
        created_by: int | None,
        message: str | None,
    ) -> int | None:
        before_snapshot = {
            "nm_id": nm_id,
            "quality_score": previous_score,
            "open_issue_count": previous_open_issue_count,
            "source_module": "checker",
        }
        after_snapshot = (
            {
                "nm_id": nm_id,
                "quality_score": new_score,
                "open_issue_count": new_open_issue_count,
                "source_module": "checker",
            }
            if new_score is not None and new_open_issue_count is not None
            else {}
        )
        comparison = ResultTrackingService().compare(
            before_snapshot, after_snapshot, problem_code="card_quality_issue"
        )
        if result_status in {"improved", "worse", "neutral"}:
            comparison["outcome"] = result_status
        event = ResultEvent(
            account_id=account_id,
            problem_code="card_quality_issue",
            source_module="checker",
            source_id=f"card_quality:{nm_id}",
            external_id=f"card_quality:{nm_id}",
            nm_id=nm_id,
            event_type="recheck_result",
            status=result_status,
            message=message or "Card quality product re-check completed.",
            payload_json={
                "source_module": "checker",
                "source": "card_quality",
                "run_id": run_id,
                "nm_id": nm_id,
                "before_snapshot": before_snapshot,
                "after_snapshot": after_snapshot,
                "comparison": comparison,
                "outcome": result_status
                if result_status in {"improved", "worse", "neutral"}
                else "not_enough_data",
                "resolved_issue_ids": resolved_issue_ids,
                "reopened_issue_ids": reopened_issue_ids,
                "saved_money_claimed": False,
                "created_by": created_by,
            },
        )
        session.add(event)
        await session.flush()
        return int(event.id) if event.id is not None else None

    async def _apply_issue_fix_to_wb(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        issue: CardQualityIssue,
        fixed_value: str,
        actor_id: int | None,
    ) -> dict[str, Any]:
        card = (
            await session.execute(
                select(WBProductCard).where(
                    WBProductCard.account_id == account_id,
                    WBProductCard.nm_id == issue.nm_id,
                )
            )
        ).scalar_one_or_none()
        if card is None:
            raise ValueError("product_card_not_found")
        try:
            await self.accounts.get_decrypted_token(
                session, account_id, WBAPICategory.CONTENT.value
            )
        except Exception as exc:
            detail = str(getattr(exc, "detail", None) or exc)
            raise ValueError(f"content_token_permission_required:{detail}") from exc
        payload, target_raw_data = await self._build_wb_update_payload(
            session,
            account_id=account_id,
            card=card,
            issue=issue,
            fixed_value=fixed_value,
        )
        response_payload = await self._submit_wb_card_update(
            session,
            account_id=account_id,
            payload=payload,
        )
        await self._apply_local_card_snapshot(
            session,
            account_id=account_id,
            card=card,
            target_raw_data=target_raw_data,
        )
        return {
            "submitted": True,
            "submitted_at": utcnow().isoformat(),
            "actor_id": actor_id,
            "external_status": "waiting_wb_validation",
            "follow_up_status": "waiting_wb_validation",
            "request_payload": payload,
            "target_field_paths": [issue.field_name],
            "wb_response": response_payload,
            "audit": {
                "event_type": "card_quality_wb_submit_attempted",
                "confirm_required": True,
                "confirmed": True,
                "preview_diff_required": True,
                "preview_diff_reviewed": True,
                "content_token_permission_checked": True,
                "listing_validation_required": True,
                "follow_up_status": "waiting_wb_validation",
                "default_writes_to_wb": False,
            },
        }

    async def _build_wb_update_payload(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        card: WBProductCard,
        issue: CardQualityIssue,
        fixed_value: str,
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        raw_data = copy.deepcopy(card.payload or {})
        raw_data["nmID"] = raw_data.get("nmID") or int(card.nm_id)
        raw_data["vendorCode"] = raw_data.get("vendorCode") or card.vendor_code
        if "kizMarked" not in raw_data and card.kiz_marked is not None:
            raw_data["kizMarked"] = bool(card.kiz_marked)
        if "title" not in raw_data and card.title is not None:
            raw_data["title"] = card.title
        if "description" not in raw_data and card.description is not None:
            raw_data["description"] = card.description
        if "brand" not in raw_data and card.brand is not None:
            raw_data["brand"] = card.brand
        if "subjectName" not in raw_data and card.subject_name is not None:
            raw_data["subjectName"] = card.subject_name
        if "dimensions" not in raw_data and isinstance(card.dimensions, dict):
            raw_data["dimensions"] = copy.deepcopy(card.dimensions)
        characteristics = raw_data.get("characteristics")
        if not isinstance(characteristics, list):
            rows = list(
                (
                    await session.execute(
                        select(WBProductCardCharacteristic)
                        .where(
                            WBProductCardCharacteristic.account_id == account_id,
                            WBProductCardCharacteristic.product_card_id == card.id,
                        )
                        .order_by(WBProductCardCharacteristic.id)
                    )
                ).scalars()
            )
            characteristics = [
                {"id": row.char_id, "name": row.name, "value": row.value}
                for row in rows
                if str(row.name or "").strip()
            ]
        raw_data["characteristics"] = copy.deepcopy(characteristics)
        if not isinstance(raw_data.get("sizes"), list):
            sizes = list(
                (
                    await session.execute(
                        select(WBProductCardSize)
                        .where(
                            WBProductCardSize.account_id == account_id,
                            WBProductCardSize.product_card_id == card.id,
                        )
                        .order_by(WBProductCardSize.id)
                    )
                ).scalars()
            )
            raw_data["sizes"] = [
                {
                    "chrtID": row.chrt_id,
                    "techSize": row.tech_size,
                    "skus": row.skus or [],
                }
                for row in sizes
            ]
        canonical = _canonical_apply_field_path(issue.field_name, issue.category)
        if canonical == "title":
            raw_data["title"] = fixed_value
        elif canonical == "description":
            raw_data["description"] = fixed_value
        elif canonical.startswith("characteristics."):
            _set_characteristic_value(
                raw_data["characteristics"],
                issue.field_name or canonical,
                fixed_value,
                issue.charc_id,
            )
        else:
            raise ValueError(
                f"Unsupported WB field for direct apply: {issue.field_name or issue.category}"
            )
        return _sanitize_wb_update_snapshot(raw_data), raw_data

    async def _submit_wb_card_update(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        payload: dict[str, Any],
    ) -> Any:
        url = "https://content-api.wildberries.ru/content/v2/cards/update"
        endpoint = "/content/v2/cards/update"
        try:
            data = await self.product_cards_sync._request_json(
                session,
                account_id=account_id,
                endpoint=endpoint,
                url=url,
                method="POST",
                json_body=[payload],
                api_category=WBAPICategory.CONTENT.value,
            )
        except WBAPIError as exc:
            raise ValueError(f"WB API error: {exc.response_text or exc}") from exc
        embedded_error = _embedded_wb_error_details(data)
        if embedded_error:
            raise ValueError(f"WB card update was rejected: {embedded_error}")
        return data if isinstance(data, dict) else {"raw": data}

    async def _apply_local_card_snapshot(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        card: WBProductCard,
        target_raw_data: dict[str, Any],
    ) -> None:
        card.payload = target_raw_data
        card.title = _as_text(target_raw_data.get("title")) or card.title
        card.description = (
            _as_text(target_raw_data.get("description")) or card.description
        )
        card.brand = _as_text(target_raw_data.get("brand")) or card.brand
        card.subject_name = (
            _as_text(target_raw_data.get("subjectName")) or card.subject_name
        )
        if "kizMarked" in target_raw_data:
            card.kiz_marked = bool(target_raw_data.get("kizMarked"))
        await session.flush()

    async def update_issue_status(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        issue_id: int,
        status: str,
        changed_by_user_id: int | None,
        reason: str | None,
        fixed_value: str | None = None,
        postponed_until: datetime | None = None,
    ) -> CardQualityIssueRead:
        issue = await session.get(CardQualityIssue, issue_id)
        if issue is None or int(issue.account_id) != int(account_id):
            raise ValueError("issue_not_found")
        old_status = issue.status
        normalized_status = self._canonical_issue_status(status)
        allowed = {
            "new": {"in_progress", "postponed", "done", "ignored", "blocked"},
            "in_progress": {"new", "postponed", "done", "ignored", "blocked"},
            "postponed": {"new", "in_progress", "done", "ignored", "blocked"},
            "done": {"new", "in_progress"},
            "ignored": {"new", "in_progress"},
            "blocked": {"new", "in_progress", "postponed", "done", "ignored"},
            "resolved": {"new", "in_progress"},
        }
        if normalized_status == old_status:
            if reason is not None:
                issue.status_reason = reason
            if fixed_value is not None:
                issue.fixed_value = fixed_value
            if postponed_until is not None:
                issue.postponed_until = postponed_until
            self._ensure_issue_defaults(issue)
            await session.commit()
            self.clear_runtime_caches()
            return CardQualityIssueRead.model_validate(issue, from_attributes=True)
        if normalized_status not in allowed.get(old_status, set()):
            raise ValueError("illegal_status_transition")
        if (
            normalized_status == "done"
            and bool(issue.requires_human_check)
            and str(reason or "").strip().lower()
            in {
                "accepted_recommendation_status_only",
                "applied_recommendation",
                "auto_apply",
                "quick_apply",
            }
        ):
            raise ValueError("human_check_issue_requires_manual_review")
        issue.status = normalized_status
        issue.status_reason = reason
        if fixed_value is not None:
            issue.fixed_value = fixed_value
        if normalized_status == "done":
            issue.fixed_value = (
                fixed_value
                if fixed_value is not None
                else issue.fixed_value
                or issue.suggested_value
                or issue.ai_suggested_value
            )
            issue.fixed_at = utcnow()
            issue.fixed_by_user_id = changed_by_user_id
            issue.resolved_at = issue.fixed_at
            issue.postponed_until = None
        elif normalized_status in {"ignored", "resolved"}:
            issue.resolved_at = utcnow()
            issue.postponed_until = None
        elif normalized_status == "postponed":
            issue.resolved_at = None
            issue.postponed_until = postponed_until
        elif normalized_status == "blocked":
            issue.resolved_at = None
            issue.postponed_until = postponed_until
        elif old_status in {"done", "ignored", "resolved"}:
            issue.resolved_at = None
            if normalized_status != "done":
                issue.fixed_at = None
                issue.fixed_by_user_id = None
        self._ensure_issue_defaults(issue)
        session.add(
            CardQualityIssueStatusHistory(
                account_id=account_id,
                issue_id=issue.id,
                old_status=old_status,
                new_status=normalized_status,
                changed_by_user_id=changed_by_user_id,
                reason=reason,
            )
        )
        await session.commit()
        self.clear_runtime_caches()
        return CardQualityIssueRead.model_validate(issue, from_attributes=True)

    def _canonical_issue_status(self, status: str | None) -> str:
        raw = str(status or "new").strip().lower()
        mapped = CHECKER_STATUS_TO_ACTION_STATUS.get(raw, raw)
        if mapped not in ACTION_STATUS_TO_CHECKER_STATUS:
            raise ValueError("illegal_status_transition")
        return ACTION_STATUS_TO_CHECKER_STATUS[mapped]

    def _ensure_issue_defaults(self, issue: CardQualityIssue) -> None:
        for attr in (
            "alternatives_json",
            "allowed_values_json",
            "error_details_json",
            "ai_alternatives_json",
            "ai_used_sources_json",
            "photo_evidence_json",
        ):
            if getattr(issue, attr, None) is None:
                setattr(issue, attr, [])
        if getattr(issue, "ai_evidence_json", None) is None:
            issue.ai_evidence_json = {}
        if getattr(issue, "requires_human_check", None) is None:
            issue.requires_human_check = False
        if getattr(issue, "score_impact", None) is None:
            issue.score_impact = SEVERITY_WEIGHTS.get(
                str(issue.severity or "").lower(), 0
            )
        if getattr(issue, "status_reason", None) is None:
            issue.status_reason = None

    async def _sync_issues(
        self,
        session: AsyncSession,
        *,
        snapshot: CardQualitySnapshot,
        normalized: NormalizedCard,
        issues: list[RuleIssue],
    ) -> tuple[int, int]:
        now = utcnow()
        fingerprints = {self.rules.fingerprint(normalized, issue) for issue in issues}
        existing_rows = list(
            (
                await session.execute(
                    select(CardQualityIssue).where(
                        CardQualityIssue.account_id == normalized.account_id,
                        CardQualityIssue.nm_id == normalized.nm_id,
                    )
                )
            ).scalars()
        )
        existing: dict[str, CardQualityIssue] = {}
        preserved_by_restore_key: dict[tuple[str, str, int], CardQualityIssue] = {}
        for issue in existing_rows:
            if issue.fingerprint not in existing or int(issue.id or 0) > int(
                existing[issue.fingerprint].id or 0
            ):
                existing[issue.fingerprint] = issue
            if issue.status in PRESERVED_USER_DECISION_STATUSES:
                key = self.rules.restore_key(
                    issue.issue_code, issue.field_name, issue.charc_id
                )
                current = preserved_by_restore_key.get(key)
                if current is None or int(issue.id or 0) > int(current.id or 0):
                    preserved_by_restore_key[key] = issue
        created = 0
        for rule_issue in issues:
            fingerprint = self.rules.fingerprint(normalized, rule_issue)
            row = existing.get(fingerprint)
            preserved = None
            if row is None:
                preserved = preserved_by_restore_key.get(
                    self.rules.restore_key(
                        rule_issue.issue_code,
                        rule_issue.field_name,
                        rule_issue.charc_id,
                    )
                )
            if row is None:
                row = CardQualityIssue(
                    account_id=normalized.account_id,
                    nm_id=normalized.nm_id,
                    snapshot_id=snapshot.id,
                    issue_code=rule_issue.issue_code,
                    category=rule_issue.category,
                    severity=rule_issue.severity,
                    title=rule_issue.title,
                    business_explanation=rule_issue.business_explanation,
                    recommended_fix=rule_issue.recommended_fix,
                    field_name=rule_issue.field_name,
                    current_value_json=rule_issue.current_value_json,
                    expected_value_json=rule_issue.expected_value_json,
                    suggested_value=rule_issue.suggested_value,
                    alternatives_json=rule_issue.alternatives or [],
                    charc_id=rule_issue.charc_id,
                    allowed_values_json=rule_issue.allowed_values or [],
                    error_details_json=rule_issue.error_details or [],
                    ai_suggested_value=rule_issue.ai_suggested_value,
                    ai_reason=rule_issue.ai_reason,
                    ai_alternatives_json=rule_issue.ai_alternatives or [],
                    ai_confidence=rule_issue.ai_confidence,
                    requires_human_check=rule_issue.requires_human_check,
                    ai_reason_short=rule_issue.ai_reason_short,
                    ai_reason_full=rule_issue.ai_reason_full,
                    ai_evidence_json=rule_issue.ai_evidence or {},
                    ai_used_sources_json=rule_issue.ai_used_sources or [],
                    photo_evidence_json=rule_issue.photo_evidence or [],
                    source=rule_issue.source,
                    score_impact=rule_issue.score_impact
                    or SEVERITY_WEIGHTS.get(rule_issue.severity, 0),
                    confidence=rule_issue.confidence,
                    status=preserved.status if preserved is not None else "new",
                    fingerprint=fingerprint,
                    first_seen_at=now,
                    last_seen_at=now,
                )
                if preserved is not None:
                    row.status_reason = preserved.status_reason
                    row.postponed_until = preserved.postponed_until
                    row.fixed_value = preserved.fixed_value
                    row.fixed_at = preserved.fixed_at
                    row.fixed_by_user_id = preserved.fixed_by_user_id
                    if row.status in {"done", "ignored"}:
                        row.resolved_at = now
                session.add(row)
                created += 1
            else:
                row.snapshot_id = snapshot.id
                row.issue_code = rule_issue.issue_code
                row.category = rule_issue.category
                row.severity = rule_issue.severity
                row.title = rule_issue.title
                row.business_explanation = rule_issue.business_explanation
                row.recommended_fix = rule_issue.recommended_fix
                row.field_name = rule_issue.field_name
                row.current_value_json = rule_issue.current_value_json
                row.expected_value_json = rule_issue.expected_value_json
                row.suggested_value = rule_issue.suggested_value
                row.alternatives_json = rule_issue.alternatives or []
                row.charc_id = rule_issue.charc_id
                row.allowed_values_json = rule_issue.allowed_values or []
                row.error_details_json = rule_issue.error_details or []
                row.ai_suggested_value = rule_issue.ai_suggested_value
                row.ai_reason = rule_issue.ai_reason
                row.ai_alternatives_json = rule_issue.ai_alternatives or []
                row.ai_confidence = rule_issue.ai_confidence
                row.requires_human_check = rule_issue.requires_human_check
                row.ai_reason_short = rule_issue.ai_reason_short
                row.ai_reason_full = rule_issue.ai_reason_full
                row.ai_evidence_json = rule_issue.ai_evidence or {}
                row.ai_used_sources_json = rule_issue.ai_used_sources or []
                row.photo_evidence_json = rule_issue.photo_evidence or []
                row.source = rule_issue.source
                row.score_impact = rule_issue.score_impact or SEVERITY_WEIGHTS.get(
                    rule_issue.severity, 0
                )
                row.confidence = rule_issue.confidence
                row.last_seen_at = now
                if row.status == "resolved":
                    old_status = row.status
                    row.status = "new"
                    row.resolved_at = None
                    session.add(
                        CardQualityIssueStatusHistory(
                            account_id=normalized.account_id,
                            issue_id=row.id,
                            old_status=old_status,
                            new_status="new",
                            reason="auto_reopen_reappeared",
                        )
                    )
                elif row.status in PRESERVED_USER_DECISION_STATUSES:
                    row.resolved_at = (
                        row.resolved_at or now
                        if row.status in {"done", "ignored"}
                        else row.resolved_at
                    )
        resolved = 0
        for fingerprint, issue in existing.items():
            if (
                fingerprint not in fingerprints
                and issue.status in ACTIVE_ISSUE_STATUSES
            ):
                old_status = issue.status
                issue.status = "resolved"
                issue.resolved_at = now
                issue.last_seen_at = now
                session.add(
                    CardQualityIssueStatusHistory(
                        account_id=normalized.account_id,
                        issue_id=issue.id,
                        old_status=old_status,
                        new_status="resolved",
                        reason="auto_resolve_absent_from_latest_analysis",
                    )
                )
                resolved += 1
        return created, resolved

    async def _upsert_registry(self, session: AsyncSession, *, account_id: int) -> None:
        snapshots = int(
            (
                await session.execute(
                    select(func.count(func.distinct(CardQualitySnapshot.nm_id))).where(
                        CardQualitySnapshot.account_id == account_id
                    )
                )
            ).scalar()
            or 0
        )
        eligible = int(
            (
                await session.execute(
                    select(func.count(func.distinct(WBProductCard.nm_id))).where(
                        WBProductCard.account_id == account_id
                    )
                )
            ).scalar()
            or 0
        )
        open_issues = int(
            (
                await session.execute(
                    select(func.count())
                    .select_from(CardQualityIssue)
                    .where(
                        CardQualityIssue.account_id == account_id,
                        CardQualityIssue.resolved_at.is_(None),
                        CardQualityIssue.status.in_(tuple(ACTIVE_ISSUE_STATUSES)),
                        CardQualityIssue.severity != "info",
                    )
                )
            ).scalar()
            or 0
        )
        info_count = int(
            (
                await session.execute(
                    select(func.count())
                    .select_from(CardQualityIssue)
                    .where(
                        CardQualityIssue.account_id == account_id,
                        CardQualityIssue.severity == "info",
                    )
                )
            ).scalar()
            or 0
        )
        latest = (
            await session.execute(
                select(CardQualitySnapshot.analyzed_at)
                .where(CardQualitySnapshot.account_id == account_id)
                .order_by(CardQualitySnapshot.analyzed_at.desc().nullslast())
                .limit(1)
            )
        ).scalar()
        integration = (
            await session.execute(
                select(PortalIntegration).where(
                    PortalIntegration.account_id == account_id,
                    PortalIntegration.module == "checker",
                )
            )
        ).scalar_one_or_none()
        status = "ok" if open_issues else "empty"
        if integration is None:
            integration = PortalIntegration(
                account_id=account_id,
                module="checker",
                enabled=True,
                mode="local",
                status=status,
            )
            session.add(integration)
        integration.enabled = True
        integration.mode = "local"
        integration.status = status
        integration.last_sync_at = latest
        integration.last_success_at = latest
        integration.last_error_code = None
        integration.last_error_message = None
        integration.metadata_json = {
            "mode": "local",
            "eligible_products": eligible,
            "unique_products_analyzed": snapshots,
            "coverage_percent": round((snapshots / eligible) * 100, 2)
            if eligible
            else 0,
            "actionable_open_issues": open_issues,
            "informational_observations": info_count,
        }

    async def _record_module_run(
        self, session: AsyncSession, *, run: CardQualityAnalysisRun
    ) -> None:
        session.add(
            PortalModuleSyncRun(
                account_id=run.account_id,
                module="checker",
                run_type=run.run_type,
                status=run.status,
                started_at=run.started_at,
                finished_at=run.finished_at,
                rows_received=run.cards_total,
                rows_processed=run.cards_processed,
                rows_created=run.cards_analyzed,
                rows_updated=run.issues_resolved,
                rows_skipped=run.cards_skipped_unchanged,
                rows_failed=run.cards_failed,
                error_summary=run.error_summary,
            )
        )

    def _issue_payload(self, issue: CardQualityIssue) -> dict[str, Any]:
        estimated_opportunity_score = float(
            SEVERITY_WEIGHTS.get(str(issue.severity or "").lower(), 0)
        )
        contract = checker_contract_fields(issue)
        capabilities = checker_action_capabilities(issue)
        return {
            "id": issue.id,
            "issue_id": issue.id,
            "nm_id": issue.nm_id,
            "problem_instance_id": None,
            "action_id": f"card_quality:{issue.id}",
            "source_order": self._source_issue_order(issue),
            "code": issue.issue_code,
            "issue_code": issue.issue_code,
            "type": issue.category,
            "category": issue.category,
            "severity": issue.severity,
            "title": issue.title,
            "description": issue.business_explanation,
            "recommendation": issue.recommended_fix,
            "field_path": issue.field_name,
            "field_name": issue.field_name,
            "current_value": issue.current_value_json,
            "current_value_json": issue.current_value_json,
            "expected_value": issue.expected_value_json,
            "expected_value_json": issue.expected_value_json,
            "suggested_value": issue.suggested_value,
            "alternatives": issue.alternatives_json or [],
            "charc_id": issue.charc_id,
            "allowed_values": issue.allowed_values_json or [],
            "error_details": issue.error_details_json or [],
            "ai_suggested_value": issue.ai_suggested_value,
            "ai_reason": issue.ai_reason,
            "ai_alternatives": issue.ai_alternatives_json or [],
            "ai_confidence": issue.ai_confidence,
            "requires_human_check": bool(issue.requires_human_check),
            "suggestion_kind": self._suggestion_kind(issue),
            "has_confirmed_suggestion": self._has_confirmed_suggestion(issue),
            "ai_reason_short": issue.ai_reason_short,
            "ai_reason_full": issue.ai_reason_full,
            "ai_evidence": issue.ai_evidence_json or {},
            "ai_used_sources": issue.ai_used_sources_json or [],
            "photo_evidence": issue.photo_evidence_json or [],
            "source": issue.source,
            "score_impact": issue.score_impact,
            "confidence": issue.confidence,
            "status": issue.status,
            "fixed_value": issue.fixed_value,
            "fixed_at": issue.fixed_at,
            "fixed_by_user_id": issue.fixed_by_user_id,
            "postponed_until": issue.postponed_until,
            "status_reason": issue.status_reason,
            "fingerprint": issue.fingerprint,
            "impact_type": contract["impact_type"],
            "trust_state": contract["trust_state"],
            "issue_group": contract["issue_group"],
            "can_fix_locally": contract["can_fix_locally"],
            "can_apply_to_wb": contract["can_apply_to_wb"],
            "apply_disabled_reason": contract["apply_disabled_reason"],
            "can_accept_local": capabilities["can_accept_local"],
            "accept_local_disabled_reason": capabilities[
                "accept_local_disabled_reason"
            ],
            "can_mark_fixed": capabilities["can_mark_fixed"],
            "mark_fixed_disabled_reason": capabilities["mark_fixed_disabled_reason"],
            "can_save_draft": capabilities["can_save_draft"],
            "save_draft_disabled_reason": capabilities["save_draft_disabled_reason"],
            "can_preview_wb": capabilities["can_preview_wb"],
            "preview_wb_disabled_reason": capabilities["preview_wb_disabled_reason"],
            "apply_wb_disabled_reason": capabilities["apply_wb_disabled_reason"],
            "can_recheck": capabilities["can_recheck"],
            "recheck_disabled_reason": capabilities["recheck_disabled_reason"],
            "score_band": contract["score_band"],
            "opportunity_score": contract["opportunity_score"],
            "expected_opportunity_count": contract["expected_opportunity_count"],
            "recheck_available": contract["recheck_available"],
            "result_status": contract["result_status"],
            "missing_data": contract["missing_data"],
            "estimated_opportunity_score": contract["opportunity_score"]
            if contract["opportunity_score"] is not None
            else estimated_opportunity_score,
            "impact_kind": "estimated_opportunity"
            if contract["impact_type"] == "opportunity"
            else contract["impact_type"],
            "impact_note": (
                "Оценочная возможность по правилам качества карточки, не подтверждённый финансовый убыток."
                if contract["impact_type"] == "opportunity"
                else "Блокер данных Checker: сначала восстановите источник, затем повторите проверку."
                if contract["impact_type"] == "data_blocker"
                else "Системная проверка Checker, не подтверждённый финансовый убыток."
                if contract["impact_type"] == "system_warning"
                else "Подтверждённый убыток только при явных финансовых after-data."
            ),
        }

    def _suggestion_kind(self, issue: CardQualityIssue) -> str:
        has_value = bool(
            str(issue.ai_suggested_value or issue.suggested_value or "").strip()
        )
        has_candidates = bool(issue.ai_alternatives_json or issue.alternatives_json)
        category = str(issue.category or "").lower()
        field_name = str(issue.field_name or "").lower()
        code = str(issue.issue_code or "").lower()
        if not has_value and not has_candidates:
            return "no_safe_fix"
        if (
            category in {"media", "photo", "photos", "video"}
            or field_name.startswith(("photos", "videos"))
            or code
            in {
                "media_no_images",
                "media_too_few_images",
                "media_no_video_info",
                "no_photos",
                "few_photos",
                "add_more_photos",
                "no_video",
            }
        ):
            return "no_safe_fix"
        if field_name in {"title", "description"} or category in {
            "title",
            "description",
        }:
            return "draft_text" if issue.requires_human_check else "exact_fix"
        return "candidate" if issue.requires_human_check else "exact_fix"

    def _has_confirmed_suggestion(self, issue: CardQualityIssue) -> bool:
        return (
            bool(str(issue.ai_suggested_value or issue.suggested_value or "").strip())
            and not bool(issue.requires_human_check)
            and self._suggestion_kind(issue) == "exact_fix"
        )

    def _next_recommended_action(
        self, issues: list[dict[str, Any]], *, nm_id: int
    ) -> dict[str, Any]:
        actionable = [
            issue
            for issue in issues
            if issue.get("severity") in {"critical", "high", "medium"}
        ]
        if not actionable:
            return {}
        issue = actionable[0]
        return {
            "action_type": "CARD_QUALITY_FIX",
            "category": issue.get("category"),
            "title": issue.get("title"),
            "deep_link": f"/products/{nm_id}?tab=quality",
            "can_execute": False,
        }

    def _action_from_issue(
        self, *, account_id: int, issue: CardQualityIssue
    ) -> PortalActionRead:
        priority = "P1" if issue.severity == "critical" else "P2"
        payload = self._issue_payload(issue)
        payload["deep_link"] = f"/products/{issue.nm_id}?tab=quality"
        bridge = build_checker_problem_bridge(
            issue, account_id=account_id, nm_id=issue.nm_id, issue_id=issue.id
        )
        payload.update(bridge.payload)
        guided_fix = self.guided_fixes.map(
            source_module="checker",
            action_type="CARD_QUALITY_FIX",
            nm_id=issue.nm_id,
            target_id=str(issue.id),
        )
        return PortalActionRead(
            id=f"card_quality:{issue.id}",
            source="card_quality_issues",
            source_module="checker",
            source_id=str(issue.id),
            account_id=account_id,
            action_type="CARD_QUALITY_FIX",
            detector_code=issue.issue_code,
            title=issue.title,
            priority=priority,
            severity=self._action_severity(issue.severity),
            status=self._action_status(issue.status),
            reason=issue.business_explanation or "",
            next_step=issue.recommended_fix
            or "Открыть Product 360 и исправить карточку",
            priority_score=payload["estimated_opportunity_score"],
            confidence=self._action_confidence(issue.confidence),
            nm_id=issue.nm_id,
            linked_entity={"nm_id": issue.nm_id, "issue_id": issue.id},
            payload=payload,
            evidence_ledger=bridge.evidence_ledger,
            money_trust=bridge.money_trust,
            trust_state=bridge.trust_state,
            impact_type=bridge.impact_type,
            allowed_actions=payload["allowed_actions"],
            recheck_rule=payload["recheck_rule_human"],
            can_execute=False,
            can_update_status=True,
            can_update=True,
            guided_fix=guided_fix,
        )

    def _action_status(self, status: str | None) -> str:
        raw = str(status or "new").lower()
        if raw == "resolved":
            return "done"
        if raw in {"new", "in_progress", "done", "postponed", "ignored", "blocked"}:
            return raw
        return "new"

    def _action_severity(self, severity: str | None) -> str:
        raw = str(severity or "medium").lower()
        return raw if raw in {"critical", "high", "medium", "low"} else "low"

    def _action_confidence(self, confidence: float | None) -> str:
        if confidence is None:
            return "medium"
        if confidence >= 0.8:
            return "high"
        if confidence >= 0.5:
            return "medium"
        return "low"
