from __future__ import annotations

import csv
import io
from copy import deepcopy
from datetime import date, timedelta
from decimal import Decimal
from typing import Any
from collections import defaultdict

from fastapi import HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import load_only

from app.core.cache import TTLMemoryCache
from app.core.pagination import Page
from app.core.time import utcnow
from app.models.marts import MartSKUDaily
from app.models.control_tower import UserBusinessSetting
from app.models.product_cards import CoreSKU
from app.models.manual_costs import ManualCost, ManualCostUpload
from app.schemas.manual_costs import ManualCostInlineSaveRow, ManualCostUpdateRequest
from app.repositories.manual_costs import (
    ManualCostRepository,
    ManualCostTemplateRepository,
    ManualCostUploadRepository,
)

EXPECTED_COLUMNS = {
    "vendorCode": "vendor_code",
    "vendor_code": "vendor_code",
    "nmId": "nm_id",
    "nm_id": "nm_id",
    "barcode": "barcode",
    "techSize": "tech_size",
    "tech_size": "tech_size",
    "costPrice": "cost_price",
    "cost_price": "cost_price",
    "unitCost": "unit_cost",
    "unit_cost": "unit_cost",
    "sellerOtherExpense": "seller_other_expense",
    "seller_other_expense": "seller_other_expense",
    "packagingCost": "packaging_cost",
    "packaging_cost": "packaging_cost",
    "inboundLogisticsCost": "inbound_logistics_cost",
    "inbound_logistics_cost": "inbound_logistics_cost",
    "supplier": "supplier",
    "currency": "currency",
    "validFrom": "valid_from",
    "valid_from": "valid_from",
    "validTo": "valid_to",
    "valid_to": "valid_to",
    "comment": "comment",
}

TEMPLATE_COLUMNS = [
    "vendorCode",
    "nmId",
    "barcode",
    "techSize",
    "productTitle",
    "current_cost_price",
    "current_seller_other_expense",
    "cost_price",
    "seller_other_expense",
    "supplier",
    "valid_from",
    "comment",
]

PLACEHOLDER_SUPPLIER = "AUTO_TEMPLATE"
TRUSTED_OPERATOR_SUPPLIER = "OPERATOR_TRUSTED_COST"
LEGACY_COST_INPUT_KEYS = (
    "packagingCost",
    "packaging_cost",
    "inboundLogisticsCost",
    "inbound_logistics_cost",
)
SELLER_OTHER_EXPENSE_INPUT_KEYS = (
    "sellerOtherExpense",
    "seller_other_expense",
)


class ManualCostService:
    def __init__(self) -> None:
        self.costs = ManualCostRepository()
        self.uploads = ManualCostUploadRepository()
        self.template_rows = ManualCostTemplateRepository()
        self._missing_costs_cache: TTLMemoryCache[dict[str, Any]] = TTLMemoryCache(
            default_ttl_seconds=30
        )

    def clear_runtime_caches(self) -> None:
        self._missing_costs_cache.clear()

    async def parse_file(self, file: UploadFile) -> list[dict[str, Any]]:
        content = await file.read()
        if file.filename and file.filename.lower().endswith(".csv"):
            return self._parse_csv(content)
        if file.filename and file.filename.lower().endswith(".xlsx"):
            return self._parse_xlsx(content)
        raise HTTPException(
            status_code=400, detail="Only CSV and XLSX files are supported"
        )

    def _normalize_row(self, row: dict[str, Any]) -> dict[str, Any]:
        normalized: dict[str, Any] = {}
        for input_key, target_key in EXPECTED_COLUMNS.items():
            if target_key not in normalized or normalized[target_key] in (None, ""):
                normalized[target_key] = row.get(input_key)
        legacy_cost_fields = sorted(
            {
                EXPECTED_COLUMNS[input_key]
                for input_key in LEGACY_COST_INPUT_KEYS
                if input_key in row
            }
        )
        explicit_seller_other_expense = next(
            (
                row.get(input_key)
                for input_key in SELLER_OTHER_EXPENSE_INPUT_KEYS
                if input_key in row and row.get(input_key) not in (None, "")
            ),
            None,
        )
        normalized["_used_legacy_cost_fields"] = bool(legacy_cost_fields)
        normalized["_legacy_cost_fields"] = legacy_cost_fields
        normalized["_seller_other_expense_explicit"] = (
            explicit_seller_other_expense is not None
        )
        normalized["cost_price"] = normalized.get("cost_price") or normalized.get(
            "unit_cost"
        )
        if not normalized["vendor_code"] or normalized["cost_price"] in (None, ""):
            raise ValueError("vendorCode and costPrice/unitCost are required")
        normalized["cost_price"] = Decimal(str(normalized["cost_price"]))
        normalized["unit_cost"] = normalized["cost_price"]
        normalized["packaging_cost"] = Decimal(
            str(normalized.get("packaging_cost") or 0)
        )
        normalized["inbound_logistics_cost"] = Decimal(
            str(normalized.get("inbound_logistics_cost") or 0)
        )
        if explicit_seller_other_expense not in (None, ""):
            normalized["seller_other_expense"] = Decimal(
                str(explicit_seller_other_expense)
            )
        else:
            normalized["seller_other_expense"] = (
                normalized["packaging_cost"] + normalized["inbound_logistics_cost"]
            )
        normalized["currency"] = normalized["currency"] or "RUB"
        normalized["valid_from"] = (
            self._parse_date(normalized["valid_from"]) or utcnow().date()
        )
        normalized["valid_to"] = self._parse_date(normalized["valid_to"])
        normalized["nm_id"] = (
            int(normalized["nm_id"]) if normalized["nm_id"] not in (None, "") else None
        )
        return normalized

    def _parse_date(self, value: Any) -> date | None:
        if value in (None, ""):
            return None
        if isinstance(value, date):
            return value
        return date.fromisoformat(str(value))

    def _parse_csv(self, content: bytes) -> list[dict[str, Any]]:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [self._normalize_row(row) for row in reader]

    def _parse_xlsx(self, content: bytes) -> list[dict[str, Any]]:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(cell) if cell is not None else "" for cell in rows[0]]
        data_rows: list[dict[str, Any]] = []
        for raw_row in rows[1:]:
            row_dict = {header[idx]: value for idx, value in enumerate(raw_row)}
            data_rows.append(self._normalize_row(row_dict))
        return data_rows

    @staticmethod
    def _is_placeholder_row(row: dict[str, Any]) -> bool:
        return str(row.get("supplier") or "").strip().upper() == PLACEHOLDER_SUPPLIER

    @staticmethod
    def _is_trusted_operator_row(row: dict[str, Any]) -> bool:
        return (
            str(row.get("supplier") or "").strip().upper() == TRUSTED_OPERATOR_SUPPLIER
        )

    @classmethod
    def _derive_cost_metadata(cls, row: dict[str, Any]) -> tuple[bool, bool, str]:
        if cls._is_placeholder_row(row):
            return True, False, "placeholder_auto_template"
        if cls._is_trusted_operator_row(row):
            return False, False, "operator_trusted_manual"
        return False, True, "operator_trusted_manual"

    @staticmethod
    def _apply_supplier_confirmation_flags(
        *,
        cost: ManualCost,
        is_supplier_confirmed: bool,
        user_id: int | None = None,
    ) -> None:
        if is_supplier_confirmed:
            cost.is_supplier_confirmed = True
            cost.cost_source = "supplier_confirmed"
            cost.is_placeholder = False
            cost.is_business_trusted = True
            cost.supplier_confirmed_at = cost.supplier_confirmed_at or utcnow()
            if user_id is not None:
                cost.supplier_confirmed_by_user_id = user_id
        else:
            cost.is_supplier_confirmed = False
            if str(cost.cost_source or "").strip().lower() == "supplier_confirmed":
                cost.cost_source = (
                    "operator_trusted_manual"
                    if cost.is_business_trusted
                    else "manual_untrusted"
                )
            cost.supplier_confirmed_at = None
            cost.supplier_confirmed_by_user_id = None

    @staticmethod
    def _jsonable_row(row: dict[str, Any]) -> dict[str, Any]:
        result: dict[str, Any] = {}
        for key, value in row.items():
            if isinstance(value, Decimal):
                result[key] = str(value)
            elif isinstance(value, date):
                result[key] = value.isoformat()
            else:
                result[key] = value
        return result

    def _row_from_json(self, row: dict[str, Any]) -> dict[str, Any]:
        result = dict(row)
        for key in (
            "unit_cost",
            "cost_price",
            "seller_other_expense",
            "packaging_cost",
            "inbound_logistics_cost",
        ):
            result[key] = Decimal(str(result.get(key) or 0))
        for key in ("valid_from", "valid_to"):
            result[key] = self._parse_date(result.get(key))
        if result.get("nm_id") not in (None, ""):
            result["nm_id"] = int(result["nm_id"])
        if result.get("sku_id") not in (None, ""):
            result["sku_id"] = int(result["sku_id"])
        return result

    async def _seller_other_expense_required(
        self,
        session: AsyncSession,
        *,
        account_id: int,
    ) -> bool:
        settings_json = (
            await session.execute(
                select(UserBusinessSetting.settings_json).where(
                    UserBusinessSetting.account_id == account_id
                )
            )
        ).scalar_one_or_none()
        if not isinstance(settings_json, dict):
            return False
        return bool(settings_json.get("require_seller_other_expense"))

    async def _commit_valid_rows(
        self,
        session: AsyncSession,
        *,
        upload: ManualCostUpload,
        rows: list[dict[str, Any]],
        user_id: int | None,
    ) -> int:
        committed = 0
        for stored_row in rows:
            row = self._row_from_json(stored_row)
            effective_from = row["valid_from"] or utcnow().date()
            is_placeholder, is_business_trusted, cost_source = (
                self._derive_cost_metadata(row)
            )
            if cost_source == "operator_trusted_manual":
                await self.costs.delete_placeholder_rows_for_sku(
                    session,
                    account_id=upload.account_id,
                    sku_id=row["sku_id"],
                )
            await self.costs.close_overlapping_rows(
                session,
                account_id=upload.account_id,
                sku_id=row["sku_id"],
                effective_from=effective_from,
            )
            session.add(
                ManualCost(
                    account_id=upload.account_id,
                    upload_id=upload.id,
                    uploaded_by_user_id=user_id or upload.created_by_user_id,
                    sku_id=row["sku_id"],
                    vendor_code=row["vendor_code"],
                    nm_id=row.get("nm_id"),
                    barcode=row.get("barcode"),
                    tech_size=row.get("tech_size"),
                    unit_cost=row["unit_cost"],
                    cost_price=row["cost_price"],
                    seller_other_expense=row["seller_other_expense"],
                    packaging_cost=row["packaging_cost"],
                    inbound_logistics_cost=row["inbound_logistics_cost"],
                    supplier=row.get("supplier"),
                    currency=row.get("currency") or "RUB",
                    valid_from=effective_from,
                    valid_to=row.get("valid_to"),
                    source_file_name=upload.filename,
                    uploaded_at=upload.imported_at or utcnow(),
                    match_rule=row.get("match_rule"),
                    cost_source=row.get("cost_source") or cost_source,
                    is_ambiguous=bool(row.get("is_ambiguous")),
                    is_placeholder=is_placeholder,
                    is_business_trusted=is_business_trusted,
                    is_supplier_confirmed=str(row.get("cost_source") or cost_source)
                    .strip()
                    .lower()
                    == "supplier_confirmed",
                    supplier_confirmed_at=(
                        utcnow()
                        if str(row.get("cost_source") or cost_source).strip().lower()
                        == "supplier_confirmed"
                        else None
                    ),
                    supplier_confirmed_by_user_id=(
                        (user_id or upload.created_by_user_id)
                        if str(row.get("cost_source") or cost_source).strip().lower()
                        == "supplier_confirmed"
                        else None
                    ),
                    comment=row.get("comment"),
                )
            )
            committed += 1
        await session.flush()
        return committed

    @staticmethod
    def _build_sku_index(
        sku_rows: list[CoreSKU],
    ) -> dict[str, dict[Any, list[CoreSKU]]]:
        index: dict[str, dict[Any, list[CoreSKU]]] = {
            "vendor_barcode_size": defaultdict(list),
            "nm_barcode": defaultdict(list),
            "barcode": defaultdict(list),
            "nm_size": defaultdict(list),
            "vendor_size": defaultdict(list),
            "vendor": defaultdict(list),
        }
        for sku in sku_rows:
            index["vendor_barcode_size"][
                (sku.vendor_code, sku.barcode, sku.tech_size)
            ].append(sku)
            index["nm_barcode"][(sku.nm_id, sku.barcode)].append(sku)
            index["barcode"][sku.barcode].append(sku)
            index["nm_size"][(sku.nm_id, sku.tech_size)].append(sku)
            index["vendor_size"][(sku.vendor_code, sku.tech_size)].append(sku)
            index["vendor"][sku.vendor_code].append(sku)
        return index

    @staticmethod
    def _resolve_sku_candidates(
        sku_rows: list[CoreSKU],
        *,
        vendor_code: str,
        nm_id: int | None,
        barcode: str | None,
        tech_size: str | None,
    ) -> tuple[list[CoreSKU], str | None]:
        return ManualCostService._resolve_sku_candidates_from_index(
            ManualCostService._build_sku_index(sku_rows),
            vendor_code=vendor_code,
            nm_id=nm_id,
            barcode=barcode,
            tech_size=tech_size,
        )

    @staticmethod
    def _resolve_sku_candidates_from_index(
        sku_index: dict[str, dict[Any, list[CoreSKU]]],
        *,
        vendor_code: str,
        nm_id: int | None,
        barcode: str | None,
        tech_size: str | None,
    ) -> tuple[list[CoreSKU], str | None]:
        rules: list[tuple[str, list[CoreSKU]]] = [
            (
                "vendor_code+barcode+tech_size",
                list(
                    sku_index["vendor_barcode_size"].get(
                        (vendor_code, barcode, tech_size), []
                    )
                ),
            ),
            (
                "nm_id+barcode",
                list(sku_index["nm_barcode"].get((nm_id, barcode), [])),
            ),
            (
                "barcode",
                list(sku_index["barcode"].get(barcode, [])),
            ),
            (
                "nm_id+tech_size",
                list(sku_index["nm_size"].get((nm_id, tech_size), [])),
            ),
            (
                "vendor_code+tech_size",
                list(sku_index["vendor_size"].get((vendor_code, tech_size), [])),
            ),
            (
                "vendor_code",
                list(sku_index["vendor"].get(vendor_code, [])),
            ),
        ]
        for rule, matches in rules:
            if matches:
                return matches, rule
        return [], None

    async def import_costs(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        created_by_user_id: int | None,
        file: UploadFile,
        commit_rows: bool = True,
    ) -> tuple[ManualCostUpload, list[dict[str, Any]]]:
        parsed_rows = await self.parse_file(file)
        require_seller_other_expense = await self._seller_other_expense_required(
            session,
            account_id=account_id,
        )
        upload = ManualCostUpload(
            account_id=account_id,
            created_by_user_id=created_by_user_id,
            filename=file.filename or "costs.csv",
            content_type=file.content_type,
            rows_total=len(parsed_rows),
            rows_valid=len(parsed_rows),
            rows_invalid=0,
            status="processed",
            imported_at=utcnow(),
            summary={"matchingPriority": ["vendorCode", "nmId", "barcode"]},
        )
        session.add(upload)
        await session.flush()
        sku_rows = list(
            (
                await session.execute(
                    select(CoreSKU)
                    .options(
                        load_only(
                            CoreSKU.id,
                            CoreSKU.nm_id,
                            CoreSKU.vendor_code,
                            CoreSKU.barcode,
                            CoreSKU.tech_size,
                            CoreSKU.is_active,
                        )
                    )
                    .where(
                        CoreSKU.account_id == account_id,
                        CoreSKU.is_active.is_(True),
                    )
                )
            ).scalars()
        )
        sku_index = self._build_sku_index(sku_rows)
        preview: list[dict[str, Any]] = []
        valid_rows: list[dict[str, Any]] = []
        invalid_rows = 0
        legacy_field_mapped_rows = 0
        legacy_field_names: set[str] = set()
        seller_other_expense_missing_rows = 0
        for row in parsed_rows:
            matches, match_rule = self._resolve_sku_candidates_from_index(
                sku_index,
                vendor_code=row["vendor_code"],
                nm_id=row["nm_id"],
                barcode=row["barcode"],
                tech_size=row["tech_size"],
            )
            warning_codes: list[str] = []
            warning_messages: list[str] = []
            if bool(row.get("_used_legacy_cost_fields")):
                legacy_field_mapped_rows += 1
                legacy_field_names.update(
                    str(value) for value in (row.get("_legacy_cost_fields") or [])
                )
                warning_codes.append("manual_cost_old_fields_used")
                warning_messages.append(
                    "Deprecated packaging/inbound logistics fields were mapped into seller_other_expense."
                )
            seller_other_expense_missing = (
                require_seller_other_expense
                and not bool(row.get("_seller_other_expense_explicit"))
                and not bool(row.get("_used_legacy_cost_fields"))
            )
            if seller_other_expense_missing:
                seller_other_expense_missing_rows += 1
                warning_codes.append("seller_other_expense_missing")
                warning_messages.append(
                    "seller_other_expense is required by account config but was not provided in this row."
                )
            preview_row = {
                **row,
                "resolved_sku_id": matches[0].id if len(matches) == 1 else None,
                "match_rule": match_rule,
                "is_ambiguous": len(matches) > 1,
                "invalid_reason": None,
                "warning_codes": warning_codes,
                "warning_messages": warning_messages,
                "used_legacy_cost_fields": bool(row.get("_used_legacy_cost_fields")),
                "legacy_cost_field_names": list(row.get("_legacy_cost_fields") or []),
                "seller_other_expense_missing": seller_other_expense_missing,
            }
            if not matches:
                preview_row["invalid_reason"] = "sku_not_found"
                invalid_rows += 1
            elif len(matches) > 1:
                preview_row["invalid_reason"] = "ambiguous_sku_match"
                invalid_rows += 1
            else:
                valid_rows.append(
                    {
                        **row,
                        "sku_id": matches[0].id,
                        "match_rule": match_rule,
                        "is_ambiguous": False,
                        "warning_codes": warning_codes,
                        "used_legacy_cost_fields": bool(
                            row.get("_used_legacy_cost_fields")
                        ),
                        "legacy_cost_field_names": list(
                            row.get("_legacy_cost_fields") or []
                        ),
                        "seller_other_expense_missing": seller_other_expense_missing,
                    }
                )
            if len(preview) < 20:
                preview.append(preview_row)
        upload.rows_valid = len(valid_rows)
        upload.rows_invalid = invalid_rows
        upload.status = "processed_with_errors" if invalid_rows else "processed"
        serialized_valid_rows = [self._jsonable_row(row) for row in valid_rows]
        upload.summary = {
            "matchingPriority": [
                "vendorCode+barcode+techSize",
                "nmId+barcode",
                "barcode",
                "nmId+techSize",
                "vendorCode+techSize",
                "vendorCode",
            ],
            "commitRows": commit_rows,
            "rowsValid": len(valid_rows),
            "rowsInvalid": invalid_rows,
            "rowsCommitted": 0,
            "legacyFieldMappedRows": legacy_field_mapped_rows,
            "legacyFieldsUsed": legacy_field_mapped_rows > 0,
            "legacyFieldNames": sorted(legacy_field_names),
            "sellerOtherExpenseRequiredByConfig": require_seller_other_expense,
            "sellerOtherExpenseMissingRows": seller_other_expense_missing_rows,
            "warningCodes": [
                code
                for code, condition in (
                    ("manual_cost_old_fields_used", legacy_field_mapped_rows > 0),
                    (
                        "seller_other_expense_missing",
                        seller_other_expense_missing_rows > 0,
                    ),
                )
                if condition
            ],
            "previewRows": [self._jsonable_row(row) for row in preview],
            "validRows": serialized_valid_rows,
        }
        if commit_rows:
            committed = await self._commit_valid_rows(
                session,
                upload=upload,
                rows=serialized_valid_rows,
                user_id=created_by_user_id,
            )
            upload.summary = {**upload.summary, "rowsCommitted": committed}
        else:
            upload.status = "validated_with_errors" if invalid_rows else "validated"
        await session.flush()
        return upload, preview

    async def list_uploads(
        self, session: AsyncSession, *, account_id: int | None = None
    ) -> list[ManualCostUpload]:
        stmt = select(ManualCostUpload).order_by(ManualCostUpload.created_at.desc())
        if account_id is not None:
            stmt = stmt.where(ManualCostUpload.account_id == account_id)
        return list((await session.execute(stmt)).scalars())

    async def list_costs(
        self,
        session: AsyncSession,
        *,
        account_id: int | None,
        limit: int = 50,
        offset: int = 0,
    ):
        return await self.costs.list_filtered(
            session,
            account_id=account_id,
            limit=limit,
            offset=offset,
        )

    async def get_upload_preview(
        self,
        session: AsyncSession,
        *,
        upload_id: int,
    ) -> tuple[ManualCostUpload, list[dict[str, Any]]]:
        upload = await session.get(ManualCostUpload, upload_id)
        if upload is None:
            raise HTTPException(status_code=404, detail="Cost upload not found")
        return upload, list((upload.summary or {}).get("previewRows") or [])

    async def confirm_upload(
        self,
        session: AsyncSession,
        *,
        upload_id: int,
        user_id: int | None,
    ) -> tuple[ManualCostUpload, int]:
        upload = await session.get(ManualCostUpload, upload_id)
        if upload is None:
            raise HTTPException(status_code=404, detail="Cost upload not found")
        summary = dict(upload.summary or {})
        if int(summary.get("rowsCommitted") or 0) > 0:
            return upload, 0
        valid_rows = list(summary.get("validRows") or [])
        if not valid_rows:
            upload.status = "processed_with_errors"
            upload.error_text = upload.error_text or "No valid rows to commit"
            await session.flush()
            return upload, 0
        committed = await self._commit_valid_rows(
            session, upload=upload, rows=valid_rows, user_id=user_id
        )
        upload.status = (
            "processed" if upload.rows_invalid == 0 else "processed_with_errors"
        )
        upload.imported_at = upload.imported_at or utcnow()
        upload.summary = {**summary, "commitRows": True, "rowsCommitted": committed}
        await session.flush()
        return upload, committed

    async def update_cost(
        self,
        session: AsyncSession,
        *,
        cost_id: int,
        payload: ManualCostUpdateRequest,
        user_id: int | None = None,
    ) -> ManualCost:
        cost = await session.get(ManualCost, cost_id)
        if cost is None:
            raise HTTPException(
                status_code=404, detail="Строка себестоимости не найдена"
            )
        update_data = payload.model_dump(exclude_unset=True)
        if not update_data:
            return cost
        if "cost_price" in update_data and update_data["cost_price"] is not None:
            cost.cost_price = update_data["cost_price"]
            cost.unit_cost = update_data["cost_price"]
        seller_other_expense_explicit = (
            "seller_other_expense" in update_data
            and update_data["seller_other_expense"] is not None
        )
        if seller_other_expense_explicit:
            cost.seller_other_expense = update_data["seller_other_expense"]
            if (
                "packaging_cost" not in update_data
                and "inbound_logistics_cost" not in update_data
            ):
                cost.packaging_cost = Decimal("0")
                cost.inbound_logistics_cost = Decimal("0")
        for field in (
            "packaging_cost",
            "inbound_logistics_cost",
            "supplier",
            "currency",
            "valid_from",
            "valid_to",
            "comment",
            "cost_source",
            "is_placeholder",
            "is_business_trusted",
        ):
            if field in update_data:
                setattr(cost, field, update_data[field])
        if not seller_other_expense_explicit and (
            "packaging_cost" in update_data or "inbound_logistics_cost" in update_data
        ):
            cost.seller_other_expense = Decimal(
                str(cost.packaging_cost or 0)
            ) + Decimal(str(cost.inbound_logistics_cost or 0))
        if (
            "is_supplier_confirmed" in update_data
            and update_data["is_supplier_confirmed"] is not None
        ):
            self._apply_supplier_confirmation_flags(
                cost=cost,
                is_supplier_confirmed=bool(update_data["is_supplier_confirmed"]),
                user_id=user_id,
            )
        supplier = (cost.supplier or "").strip().upper()
        if supplier == PLACEHOLDER_SUPPLIER:
            cost.is_placeholder = True
            cost.is_business_trusted = False
            cost.is_supplier_confirmed = False
            cost.cost_source = "placeholder_auto_template"
        elif supplier == TRUSTED_OPERATOR_SUPPLIER:
            cost.is_placeholder = False
            cost.is_business_trusted = True
            if not cost.is_supplier_confirmed:
                cost.cost_source = cost.cost_source or "operator_trusted_manual"
        elif cost.is_placeholder:
            cost.is_business_trusted = False
            cost.is_supplier_confirmed = False
            cost.cost_source = cost.cost_source or "placeholder_auto_template"
        else:
            if cost.is_supplier_confirmed:
                self._apply_supplier_confirmation_flags(
                    cost=cost,
                    is_supplier_confirmed=True,
                    user_id=user_id,
                )
            else:
                cost.cost_source = cost.cost_source or "operator_trusted_manual"
                cost.is_business_trusted = bool(cost.is_business_trusted)
        await session.flush()
        return cost

    async def save_inline_costs(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        rows: list[ManualCostInlineSaveRow],
        user_id: int | None = None,
    ) -> list[ManualCost]:
        saved: list[ManualCost] = []
        today = utcnow().date()
        default_effective_from = today - timedelta(days=30)
        for row in rows:
            if row.cost_id is not None:
                cost = await session.get(ManualCost, row.cost_id)
                if cost is None or int(cost.account_id) != int(account_id):
                    raise HTTPException(
                        status_code=404, detail="Строка себестоимости не найдена"
                    )
                saved.append(
                    await self.update_cost(
                        session,
                        cost_id=int(row.cost_id),
                        payload=self._inline_update_payload(row),
                        user_id=user_id,
                    )
                )
                continue

            if row.sku_id is None:
                raise HTTPException(
                    status_code=400, detail="Передайте cost_id или sku_id"
                )

            sku = await session.get(CoreSKU, int(row.sku_id))
            if (
                sku is None
                or int(sku.account_id) != int(account_id)
                or not bool(sku.is_active)
            ):
                raise HTTPException(status_code=404, detail="Активный SKU не найден")

            effective_from = row.valid_from or default_effective_from
            current = await self._active_real_cost_for_sku(
                session,
                account_id=account_id,
                sku_id=int(sku.id),
                active_on=effective_from,
            )
            if current is not None:
                saved.append(
                    await self.update_cost(
                        session,
                        cost_id=int(current.id),
                        payload=self._inline_update_payload(row),
                        user_id=user_id,
                    )
                )
                continue

            await self.costs.delete_placeholder_rows_for_sku(
                session,
                account_id=account_id,
                sku_id=int(sku.id),
            )
            await self.costs.close_overlapping_rows(
                session,
                account_id=account_id,
                sku_id=int(sku.id),
                effective_from=effective_from,
            )

            cost_price = Decimal(str(row.cost_price))
            seller_other_expense = Decimal(str(row.seller_other_expense or 0))
            supplier_confirmed = bool(row.is_supplier_confirmed)
            created = ManualCost(
                account_id=account_id,
                upload_id=None,
                uploaded_by_user_id=user_id,
                sku_id=int(sku.id),
                vendor_code=sku.vendor_code or f"SKU-{sku.id}",
                nm_id=sku.nm_id,
                barcode=sku.barcode,
                tech_size=sku.tech_size,
                unit_cost=cost_price,
                cost_price=cost_price,
                seller_other_expense=seller_other_expense,
                packaging_cost=Decimal("0"),
                inbound_logistics_cost=Decimal("0"),
                supplier=row.supplier or TRUSTED_OPERATOR_SUPPLIER,
                currency=row.currency or "RUB",
                valid_from=effective_from,
                valid_to=None,
                source_file_name="inline-platform",
                uploaded_at=utcnow(),
                match_rule="sku_id",
                cost_source="supplier_confirmed"
                if supplier_confirmed
                else "operator_trusted_manual",
                is_ambiguous=False,
                is_placeholder=False,
                is_business_trusted=True,
                is_supplier_confirmed=supplier_confirmed,
                supplier_confirmed_at=utcnow() if supplier_confirmed else None,
                supplier_confirmed_by_user_id=user_id if supplier_confirmed else None,
                comment=row.comment or "Заполнено вручную в платформе",
            )
            session.add(created)
            await session.flush()
            saved.append(created)
        await session.flush()
        return saved

    def _inline_update_payload(
        self, row: ManualCostInlineSaveRow
    ) -> ManualCostUpdateRequest:
        payload: dict[str, Any] = {
            "cost_price": row.cost_price,
            "supplier": row.supplier or TRUSTED_OPERATOR_SUPPLIER,
            "currency": row.currency or "RUB",
            "cost_source": "operator_trusted_manual",
            "is_placeholder": False,
            "is_business_trusted": True,
        }
        if row.seller_other_expense is not None:
            payload["seller_other_expense"] = row.seller_other_expense
        if row.valid_from is not None:
            payload["valid_from"] = row.valid_from
        if row.comment is not None:
            payload["comment"] = row.comment
        if row.is_supplier_confirmed is not None:
            payload["is_supplier_confirmed"] = row.is_supplier_confirmed
        return ManualCostUpdateRequest(**payload)

    async def _active_real_cost_for_sku(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        sku_id: int,
        active_on: date,
    ) -> ManualCost | None:
        return (
            (
                await session.execute(
                    select(ManualCost)
                    .where(
                        ManualCost.account_id == account_id,
                        ManualCost.sku_id == sku_id,
                        ManualCost.is_placeholder.is_not(True),
                        or_(
                            ManualCost.supplier.is_(None),
                            ManualCost.supplier != PLACEHOLDER_SUPPLIER,
                        ),
                        or_(
                            ManualCost.valid_from.is_(None),
                            ManualCost.valid_from <= active_on,
                        ),
                        or_(
                            ManualCost.valid_to.is_(None),
                            ManualCost.valid_to >= active_on,
                        ),
                    )
                    .order_by(
                        ManualCost.valid_from.desc().nullslast(), ManualCost.id.desc()
                    )
                    .limit(1)
                )
            )
            .scalars()
            .first()
        )

    async def mark_supplier_confirmed(
        self,
        session: AsyncSession,
        *,
        cost_id: int,
        user_id: int | None,
        comment: str | None = None,
    ) -> ManualCost:
        cost = await session.get(ManualCost, cost_id)
        if cost is None:
            raise HTTPException(
                status_code=404, detail="Строка себестоимости не найдена"
            )
        self._apply_supplier_confirmation_flags(
            cost=cost,
            is_supplier_confirmed=True,
            user_id=user_id,
        )
        if comment:
            cost.comment = comment if not cost.comment else f"{cost.comment}\n{comment}"
        await session.flush()
        return cost

    async def relink_costs(
        self,
        session: AsyncSession,
        *,
        account_id: int,
    ) -> dict[str, int]:
        costs = list(
            (
                await session.execute(
                    select(ManualCost).where(ManualCost.account_id == account_id)
                )
            ).scalars()
        )
        sku_rows = list(
            (
                await session.execute(
                    select(CoreSKU)
                    .options(
                        load_only(
                            CoreSKU.id,
                            CoreSKU.nm_id,
                            CoreSKU.vendor_code,
                            CoreSKU.barcode,
                            CoreSKU.tech_size,
                            CoreSKU.is_active,
                        )
                    )
                    .where(
                        CoreSKU.account_id == account_id, CoreSKU.is_active.is_(True)
                    )
                )
            ).scalars()
        )
        sku_index = self._build_sku_index(sku_rows)
        active_by_id = {sku.id: sku for sku in sku_rows}
        checked = relinked = ambiguous = unresolved = 0
        for cost in costs:
            checked += 1
            if (
                cost.sku_id is not None
                and cost.sku_id in active_by_id
                and not cost.is_ambiguous
            ):
                continue
            matches, match_rule = self._resolve_sku_candidates_from_index(
                sku_index,
                vendor_code=cost.vendor_code,
                nm_id=cost.nm_id,
                barcode=cost.barcode,
                tech_size=cost.tech_size,
            )
            if len(matches) == 1:
                cost.sku_id = matches[0].id
                cost.match_rule = match_rule
                cost.is_ambiguous = False
                relinked += 1
            elif len(matches) > 1:
                cost.sku_id = None
                cost.match_rule = match_rule
                cost.is_ambiguous = True
                ambiguous += 1
            else:
                cost.sku_id = None
                cost.is_ambiguous = False
                unresolved += 1
        await session.flush()
        return {
            "checked_count": checked,
            "relinked_count": relinked,
            "ambiguous_count": ambiguous,
            "unresolved_count": unresolved,
        }

    async def list_unresolved_costs(
        self,
        session: AsyncSession,
        *,
        account_id: int | None,
        limit: int = 50,
        offset: int = 0,
    ) -> list[ManualCost]:
        page = await self.costs.list_unresolved_page(
            session,
            account_id=account_id,
            limit=limit,
            offset=offset,
        )
        return list(page.items)

    async def list_unresolved_costs_page(
        self,
        session: AsyncSession,
        *,
        account_id: int | None,
        limit: int = 50,
        offset: int = 0,
    ) -> Page[ManualCost]:
        return await self.costs.list_unresolved_page(
            session,
            account_id=account_id,
            limit=limit,
            offset=offset,
        )

    async def list_unresolved_costs_for_product(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        nm_id: int | None = None,
        sku_id: int | None = None,
        vendor_code: str | None = None,
        barcode: str | None = None,
        limit: int = 20,
    ) -> list[ManualCost]:
        return await self.costs.list_unresolved_for_product(
            session,
            account_id=account_id,
            nm_id=nm_id,
            sku_id=sku_id,
            vendor_code=vendor_code,
            barcode=barcode,
            limit=limit,
        )

    async def build_template_csv(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        mode: str = "all",
        date_from: date | None = None,
        date_to: date | None = None,
    ) -> str:
        rows = await self._template_rows(
            session,
            account_id=account_id,
            mode=mode,
            date_from=date_from,
            date_to=date_to,
        )
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=TEMPLATE_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow(self._template_dict(row))
        return output.getvalue()

    async def build_template_xlsx(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        mode: str = "all",
        date_from: date | None = None,
        date_to: date | None = None,
    ) -> bytes:
        rows = await self._template_rows(
            session,
            account_id=account_id,
            mode=mode,
            date_from=date_from,
            date_to=date_to,
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Себестоимость"
        sheet.append(TEMPLATE_COLUMNS)
        sheet.freeze_panes = "A2"
        for row in rows:
            data = self._template_dict(row)
            sheet.append([data[column] for column in TEMPLATE_COLUMNS])
        instruction = workbook.create_sheet("Инструкция")
        instruction.append(["Поле", "Как заполнить"])
        instruction.append(
            ["cost_price", "Укажите закупочную себестоимость за единицу."]
        )
        instruction.append(
            [
                "seller_other_expense",
                "Укажите прочие расходы продавца на единицу, если они есть.",
            ]
        )
        instruction.append(["valid_from", "Дата начала действия в формате YYYY-MM-DD."])
        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _template_dict(row: dict[str, Any]) -> dict[str, Any]:
        return {
            "vendorCode": row.get("vendor_code") or "",
            "nmId": row.get("nm_id") or "",
            "barcode": row.get("barcode") or "",
            "techSize": row.get("tech_size") or "",
            "productTitle": row.get("product_title") or "",
            "current_cost_price": row.get("current_cost_price") or "",
            "current_seller_other_expense": row.get("current_seller_other_expense")
            or "",
            "cost_price": "",
            "seller_other_expense": "",
            "supplier": "",
            "valid_from": "",
            "comment": row.get("product_title") or "",
        }

    async def _template_rows(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        mode: str,
        date_from: date | None,
        date_to: date | None,
    ) -> list[dict[str, Any]]:
        normalized_mode = str(mode or "all").lower()
        if normalized_mode not in {"all", "missing"}:
            raise HTTPException(status_code=400, detail="mode must be all or missing")
        if normalized_mode == "missing":
            result = await self.list_missing_costs(
                session,
                account_id=account_id,
                limit=10000,
                offset=0,
                date_from=date_from,
                date_to=date_to,
                only_revenue=False,
            )
            return [
                {
                    "sku_id": item["sku_id"],
                    "nm_id": item.get("nm_id"),
                    "vendor_code": item.get("vendor_code"),
                    "barcode": item.get("barcode"),
                    "tech_size": item.get("tech_size"),
                    "product_title": item.get("product_title"),
                    "current_cost_price": "",
                    "current_seller_other_expense": "",
                }
                for item in result["items"]
            ]

        active_rows = await self.template_rows.list_template_rows(
            session, account_id=account_id
        )
        current_costs = await self._current_costs_by_sku(
            session, account_id=account_id, active_on=date_to or utcnow().date()
        )
        rows: list[dict[str, Any]] = []
        for sku in active_rows:
            current = current_costs.get(int(sku.id))
            rows.append(
                {
                    "sku_id": sku.id,
                    "nm_id": sku.nm_id,
                    "vendor_code": sku.vendor_code,
                    "barcode": sku.barcode,
                    "tech_size": sku.tech_size,
                    "product_title": sku.title,
                    "current_cost_price": getattr(current, "cost_price", "")
                    if current is not None
                    else "",
                    "current_seller_other_expense": getattr(
                        current, "seller_other_expense", ""
                    )
                    if current is not None
                    else "",
                }
            )
        return rows

    async def _current_costs_by_sku(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        active_on: date,
    ) -> dict[int, ManualCost]:
        rows = list(
            (
                await session.execute(
                    select(ManualCost)
                    .where(
                        ManualCost.account_id == account_id,
                        ManualCost.sku_id.is_not(None),
                        ManualCost.is_placeholder.is_not(True),
                        or_(
                            ManualCost.supplier.is_(None),
                            ManualCost.supplier != PLACEHOLDER_SUPPLIER,
                        ),
                        or_(
                            ManualCost.valid_from.is_(None),
                            ManualCost.valid_from <= active_on,
                        ),
                        or_(
                            ManualCost.valid_to.is_(None),
                            ManualCost.valid_to >= active_on,
                        ),
                    )
                    .order_by(
                        ManualCost.sku_id.asc(),
                        ManualCost.valid_from.desc().nullslast(),
                        ManualCost.id.desc(),
                    )
                )
            ).scalars()
        )
        result: dict[int, ManualCost] = {}
        for row in rows:
            if row.sku_id is not None:
                result.setdefault(int(row.sku_id), row)
        return result

    async def list_missing_costs(
        self,
        session: AsyncSession,
        *,
        account_id: int,
        limit: int = 50,
        offset: int = 0,
        date_from: date | None = None,
        date_to: date | None = None,
        only_revenue: bool = True,
    ) -> dict[str, Any]:
        today = utcnow().date()
        actual_to = date_to or today
        actual_from = date_from or actual_to.replace(day=1)
        active_on = actual_to
        cache_key = (
            "missing_costs",
            int(account_id),
            actual_from.isoformat(),
            actual_to.isoformat(),
            bool(only_revenue),
            int(limit),
            int(offset),
        )
        cached = self._missing_costs_cache.get(cache_key)
        if cached is not None:
            return deepcopy(cached)
        costs_at_period_start = await self._current_costs_by_sku(
            session, account_id=account_id, active_on=actual_from
        )
        costs_at_period_end = await self._current_costs_by_sku(
            session, account_id=account_id, active_on=active_on
        )
        missing_mart_rows = list(
            (
                await session.execute(
                    select(
                        MartSKUDaily.sku_id,
                        func.coalesce(func.sum(MartSKUDaily.final_revenue), 0),
                    )
                    .where(
                        MartSKUDaily.account_id == account_id,
                        MartSKUDaily.stat_date >= actual_from,
                        MartSKUDaily.stat_date <= actual_to,
                        MartSKUDaily.sku_id.is_not(None),
                        or_(MartSKUDaily.finance_rows > 0, MartSKUDaily.sale_rows > 0),
                        MartSKUDaily.has_manual_cost.is_(False),
                    )
                    .group_by(MartSKUDaily.sku_id)
                )
            ).all()
        )
        missing_revenue_by_sku = {
            int(sku_id): float(revenue or 0)
            for sku_id, revenue in missing_mart_rows
            if sku_id is not None
        }
        period_missing_sku_ids = set(missing_revenue_by_sku)
        revenue_rows = list(
            (
                await session.execute(
                    select(
                        MartSKUDaily.sku_id,
                        func.coalesce(func.sum(MartSKUDaily.final_revenue), 0),
                    )
                    .where(
                        MartSKUDaily.account_id == account_id,
                        MartSKUDaily.stat_date >= actual_from,
                        MartSKUDaily.stat_date <= actual_to,
                        MartSKUDaily.sku_id.is_not(None),
                    )
                    .group_by(MartSKUDaily.sku_id)
                )
            ).all()
        )
        revenue_by_sku = {
            int(sku_id): float(revenue or 0)
            for sku_id, revenue in revenue_rows
            if sku_id is not None
        }
        sku_rows = list(
            (
                await session.execute(
                    select(CoreSKU)
                    .options(
                        load_only(
                            CoreSKU.id,
                            CoreSKU.nm_id,
                            CoreSKU.vendor_code,
                            CoreSKU.barcode,
                            CoreSKU.tech_size,
                            CoreSKU.title,
                            CoreSKU.is_active,
                        )
                    )
                    .where(
                        CoreSKU.account_id == account_id, CoreSKU.is_active.is_(True)
                    )
                    .order_by(
                        CoreSKU.vendor_code.asc().nullslast(),
                        CoreSKU.tech_size.asc().nullslast(),
                        CoreSKU.id.asc(),
                    )
                )
            ).scalars()
        )
        missing_items: list[dict[str, Any]] = []
        for sku in sku_rows:
            sku_id = int(sku.id)
            affected_revenue = float(missing_revenue_by_sku.get(sku_id, 0.0))
            has_period_gap = sku_id in period_missing_sku_ids
            has_catalog_gap = (
                sku_id not in costs_at_period_start or sku_id not in costs_at_period_end
            )
            if not has_period_gap and not has_catalog_gap:
                continue
            if only_revenue and affected_revenue <= 0:
                continue
            missing_items.append(
                {
                    "sku_id": sku_id,
                    "nm_id": sku.nm_id,
                    "vendor_code": sku.vendor_code,
                    "barcode": sku.barcode,
                    "tech_size": sku.tech_size,
                    "product_title": sku.title,
                    "affected_revenue": affected_revenue,
                    "recommended_action": "Заполнить себестоимость",
                }
            )
        total_revenue = float(sum(revenue_by_sku.values()))
        affected_revenue = float(
            sum(item["affected_revenue"] for item in missing_items)
        )
        coverage_percent = None
        if total_revenue > 0:
            coverage_percent = round(
                (total_revenue - affected_revenue) / total_revenue * 100, 4
            )
        result = {
            "total": len(missing_items),
            "limit": limit,
            "offset": offset,
            "summary": {
                "missing_sku_count": len(missing_items),
                "affected_revenue": affected_revenue,
                "revenue_cost_coverage_percent": coverage_percent,
            },
            "items": missing_items[offset : offset + limit],
        }
        self._missing_costs_cache.set(cache_key, deepcopy(result))
        return result
