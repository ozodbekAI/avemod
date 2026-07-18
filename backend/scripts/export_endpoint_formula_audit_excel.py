from __future__ import annotations

import argparse
import json
import re
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_BUNDLE = Path("/home/ozodbek/AVEMOD_PROJECTS/Finance/exports/live_backend_full_audit_20260522_201627")
EXPORT_ROOT = Path("/home/ozodbek/AVEMOD_PROJECTS/Finance/exports")


SOURCE_GROUPS: list[tuple[str, str, str, str]] = [
    ("/api/v1/money", "money-management", "mart_sku_daily, mart_stock_daily, mart_finance_reconciliation, mart_account_expense_daily, action_recommendations, alert_events, user_business_settings, data_quality_issues, manual_costs, core_sku, wb_ad_stats_daily", "Primary business money layer over marts, costs, actions, ads, stocks and DQ gates."),
    ("/api/v1/dashboard/data-health", "quality-and-health", "data_quality_issues, wb_sync_runs, wb_sync_cursors, mart_sku_daily, manual_costs, wb_ad_cluster_stats", "Health and trust gate over sync, marts and manual cost state."),
    ("/api/v1/dashboard/article-audit", "business-marts", "mart_sku_daily, mart_stock_daily, mart_finance_reconciliation, mart_account_expense_daily, mart_reconciliation_daily, manual_costs, data_quality_issues, wb_sales, wb_orders, wb_realization_report_rows, wb_ad_stats_daily", "Article/nm_id audit built from marts plus raw operational, finance, stock and ads data."),
    ("/api/v1/dashboard/owner", "control-tower", "mart_sku_daily, mart_stock_daily, action_recommendations, alert_events, user_business_settings, formula_audit_runs, data_quality_issues", "Owner dashboard aggregates control rows and action recommendations."),
    ("/api/v1/dashboard/sku-profitability", "business-marts", "mart_sku_daily, manual_costs, core_sku", "SKU profitability built from daily mart rows plus cost truth flags."),
    ("/api/v1/skus", "control-tower", "mart_sku_daily, mart_stock_daily, action_recommendations, data_quality_issues, core_sku, wb_prices, manual_costs", "Control rows for SKU-level monitoring."),
    ("/api/v1/actions", "control-tower", "action_recommendations, action_recommendation_history", "Persisted recommendations and state transitions."),
    ("/api/v1/inventory/purchase-plan", "control-tower", "mart_sku_daily, mart_stock_daily, core_sku, user_business_settings", "Purchase status and reorder math on top of control rows."),
    ("/api/v1/pricing/safety", "price", "wb_prices, wb_price_sizes, core_sku, mart_sku_daily, manual_costs", "Price safety calculations from price snapshots, unit cost and economics."),
    ("/api/v1/pricing/simulate", "price", "wb_prices, wb_price_sizes, core_sku, mart_sku_daily, manual_costs", "What-if simulation over price safety baseline."),
    ("/api/v1/ads/efficiency", "ads", "wb_ad_campaigns, wb_ad_stats_daily, wb_ad_cluster_stats, mart_sku_daily", "Ads efficiency read model from ad source spend plus control rows."),
    ("/api/v1/core-sku", "identity", "wb_product_cards, wb_product_card_sizes, core_sku, data_quality_issues, manual_costs", "Resolved SKU identity enriched from product cards and cost metadata."),
    ("/api/v1/products", "identity", "wb_product_cards, wb_product_card_sizes", "Raw product card reads."),
    ("/api/v1/prices", "price", "wb_prices, wb_price_sizes, core_sku", "Raw/latest price snapshots."),
    ("/api/v1/orders", "operations", "wb_orders", "Operational orders feed from WB statistics API."),
    ("/api/v1/sales", "operations", "wb_sales", "Operational sales and returns feed from WB statistics API."),
    ("/api/v1/stocks/snapshots", "stock", "wb_stock_snapshots, wb_stock_snapshot_rows", "Raw stock snapshot tasks and rows."),
    ("/api/v1/balance", "finance", "wb_balance_snapshots", "Latest WB balance snapshot."),
    ("/api/v1/finance/reports", "finance", "wb_realization_reports", "Finance report headers."),
    ("/api/v1/finance/report-rows", "finance", "wb_realization_report_rows", "Detailed finance rows from WB realization reports."),
    ("/api/v1/costs", "manual-costs", "manual_cost_uploads, manual_costs, core_sku", "Operator-uploaded cost layer."),
    ("/api/v1/marts", "business-marts", "mart_sku_daily, mart_stock_daily, mart_finance_reconciliation, mart_account_expense_daily, mart_reconciliation_daily", "Derived business marts."),
    ("/api/v1/dq", "quality-and-health", "data_quality_issues", "Data quality issues and investigator payloads."),
    ("/api/v1/sync", "sync-control", "wb_sync_runs, wb_sync_cursors, raw_wb_api_responses", "Sync history, cursors and scheduler state."),
    ("/api/v1/tariffs", "tariffs", "wb_tariff_commissions, wb_tariff_boxes, wb_tariff_pallets, wb_tariff_returns, wb_tariff_acceptance", "Reference tariff coefficients from WB APIs."),
    ("/api/v1/analytics", "analytics", "wb_card_funnel_daily, wb_region_sales_daily, wb_hidden_products", "Card funnel and regional sales analytics."),
    ("/api/v1/documents", "documents", "wb_document_categories, wb_documents", "WB document metadata."),
    ("/api/v1/supplies", "supplies", "wb_supplies, wb_supply_goods, wb_supply_packages, wb_supply_warehouses, wb_supply_acceptance_options", "Inbound supplies and acceptance data."),
    ("/api/v1/accounts", "account-config", "wb_accounts, wb_api_tokens", "Seller account and token registry."),
    ("/api/v1/meta", "metadata", "", "Internal metadata and enum labels."),
    ("/health", "service-health", "", "Fast health check endpoint."),
]


SERVICE_MAP: dict[str, str] = {
    "/api/v1/money/summary": "MoneyManagementService.summary",
    "/api/v1/money/cards": "MoneyManagementService.cards",
    "/api/v1/money/cards/{sku_id}": "MoneyManagementService.card_detail",
    "/api/v1/money/articles": "MoneyManagementService.articles",
    "/api/v1/money/articles/{nm_id}": "MoneyManagementService.article_detail",
    "/api/v1/money/actions/today": "MoneyManagementService.today_actions",
    "/api/v1/money/data-blockers": "MoneyManagementService.data_blockers",
    "/api/v1/money/filters": "MoneyManagementService.filters",
    "/api/v1/dashboard/data-health": "DashboardService.data_health",
    "/api/v1/dashboard/article-audit": "DashboardService.article_audit",
    "/api/v1/dashboard/sku-profitability": "DashboardService.sku_profitability_page",
    "/api/v1/dashboard/owner": "ControlTowerService.owner_dashboard",
    "/api/v1/skus": "ControlTowerService.list_control_skus",
    "/api/v1/skus/{sku_id}": "ControlTowerService.get_control_sku_detail",
    "/api/v1/actions": "ControlTowerService.list_actions",
    "/api/v1/inventory/purchase-plan": "ControlTowerService.list_purchase_plan",
    "/api/v1/pricing/safety": "ControlTowerService.list_price_safety",
    "/api/v1/pricing/simulate": "ControlTowerService.simulate_price",
    "/api/v1/ads/efficiency": "ControlTowerService.list_ads_efficiency",
    "/api/v1/core-sku": "CoreSKUService.list_core_skus",
    "/api/v1/core-sku/{sku_id}": "CoreSKUService.get_core_sku",
    "/api/v1/finance/report-rows": "FinanceService.list_report_rows",
    "/api/v1/balance": "FinanceService.latest_balance",
}


FORMULA_RULES: list[tuple[str, str, str, str, str]] = [
    ("/api/v1/money/summary", r"^kpis\.revenue$", "derived", "sum(state.profit_rows.realized_revenue)", "Source rows: mart/profit rows"),
    ("/api/v1/money/summary", r"^kpis\.for_pay$", "derived", "sum(state.profit_rows.for_pay)", "Source rows: mart/profit rows"),
    ("/api/v1/money/summary", r"^kpis\.finance_confirmed_revenue$", "derived", "sum(item.realized_revenue for item in state.profit_rows if finance_rows > 0)", "Uses finance_rows gate"),
    ("/api/v1/money/summary", r"^kpis\.supplier_cost_confirmed_revenue$", "derived", "state.health.revenue_with_real_cost", "Real supplier-confirmed cost coverage numerator"),
    ("/api/v1/money/summary", r"^kpis\.supplier_cost_confirmed_revenue_percent$", "ratio", "supplier_cost_confirmed_revenue / data_health revenue denominator * 100", "Exact denominator is revenue_with_cost + revenue_without_cost from data-health"),
    ("/api/v1/money/summary", r"^kpis\.net_profit_after_ads$|^kpis\.profit_after_source_ads$", "derived", "profit_before_ads - max(ads_source_spend, mart_ads_allocated_spend)", "profit_before_ads = sum(estimated_profit) + mart_ads_allocated_spend"),
    ("/api/v1/money/summary", r"^kpis\.profit_after_allocated_ads$", "derived", "profit_before_ads - ads_allocated_spend", "ads_allocated_spend = max(sum(mart ad_spend), sum(source ads by nm))"),
    ("/api/v1/money/summary", r"^kpis\.margin_percent$", "ratio", "profit_after_source_ads / revenue * 100", "Business margin after source ads"),
    ("/api/v1/money/summary", r"^kpis\.roi_percent$|^kpis\.roi_on_cogs_percent$", "ratio", "profit_after_source_ads / cogs * 100", "COGS from summed estimated_cogs"),
    ("/api/v1/money/summary", r"^kpis\.stock_roi_percent$", "ratio", "profit_after_source_ads / stock_value * 100", "Uses stock_value fallback from stock qty * unit cost"),
    ("/api/v1/money/summary", r"^kpis\.roas_percent$", "ratio", "revenue / ads_source_spend * 100", "Displayed as percent-like ROAS multiplier * 100"),
    ("/api/v1/money/summary", r"^kpis\.wb_expenses_total$|^kpis\.direct_wb_expenses$", "derived", "max(0, commission + acquiring_fee + logistics + paid_acceptance + storage + penalties + deductions - additional_payments)", "Direct WB expenses only"),
    ("/api/v1/money/summary", r"^kpis\.account_level_expenses$|^kpis\.unallocated_expenses$", "passthrough", "account_level_expense_total", "From mart_account_expense_daily / account-level pool"),
    ("/api/v1/money/summary", r"^kpis\.stock_value$", "derived", "sum(stock_qty * unit_cost) over control rows", "Unit cost from core_sku.total_unit_cost or estimated_cogs/net_units fallback"),
    ("/api/v1/money/summary", r"^kpis\.overstock_value$", "derived", "sum(stock_value where sku_status in {'LIQUIDATE','LIQUIDATE_FIRST'})", "Overstock subset of stock_value"),
    ("/api/v1/money/summary", r"^kpis\.in_transit_value$", "derived", "sum((in_way_to_client + in_way_from_client) * unit_cost)", "Transit valued at unit cost"),
    ("/api/v1/money/summary", r"^kpis\.ad_spend$|^kpis\.ads_source_spend$", "derived", "state.ads_source_total", "Raw/source ads spend from wb_ad_stats_daily"),
    ("/api/v1/money/summary", r"^kpis\.ads_allocated_spend$", "derived", "max(sum(mart ad_spend), sum(ads_source_by_nm.values()))", "Allocated spend visible to money layer"),
    ("/api/v1/money/summary", r"^kpis\.ads_unallocated_spend$", "derived", "max(0, ads_source_spend - ads_allocated_spend)", "Unmapped source ads spend"),
    ("/api/v1/money/summary", r"^quality\.ads_allocation_percent$", "ratio", "ads_allocated_spend / ads_source_spend * 100", "0 if no source ads"),
    ("/api/v1/money/summary", r"^revenue_sources\.finance_confirmed_revenue$", "derived", "sum(item.realized_revenue for profit rows with finance_rows > 0)", "Finance-confirmed by row presence"),
    (r"/api/v1/money/cards/\{sku_id\}|/api/v1/money/articles/\{nm_id\}", r"^money\.wb_expenses\.direct$|^money\.wb_expenses_total$", "derived", "commission + acquiring_fee + logistics + paid_acceptance + storage + penalties + deductions - additional_payments", "Direct WB expense block"),
    (r"/api/v1/money/cards/\{sku_id\}|/api/v1/money/articles/\{nm_id\}", r"^money\.ads\.source_spend$", "derived", "article/source ads spend allocated to current variant/card", "nm_id source ads with revenue/units/even fallback"),
    (r"/api/v1/money/cards/\{sku_id\}|/api/v1/money/articles/\{nm_id\}", r"^money\.ads\.unallocated_spend$", "derived", "max(0, source_spend - allocated_spend)", "Unmapped ad part"),
    (r"/api/v1/money/cards/\{sku_id\}|/api/v1/money/articles/\{nm_id\}", r"^money\.ads\.drr_percent_source$", "ratio", "source_spend / revenue * 100", "Source DRR"),
    (r"/api/v1/money/cards/\{sku_id\}|/api/v1/money/articles/\{nm_id\}", r"^money\.cogs\.unit_cost$", "ratio", "estimated_cogs / net_units", "If net_units > 0"),
    (r"/api/v1/money/cards/\{sku_id\}|/api/v1/money/articles/\{nm_id\}", r"^money\.profit\.before_ads$|^profit_variants\.before_ads$", "derived", "profit_after_allocated_ads + ads_allocated", "Variant/card profit before ads"),
    (r"/api/v1/money/cards/\{sku_id\}|/api/v1/money/articles/\{nm_id\}", r"^money\.profit\.after_allocated_ads$|^profit_variants\.after_allocated_ads$", "derived", "profit_before_ads - ads.allocated_spend", "Profit after allocated ads"),
    (r"/api/v1/money/cards/\{sku_id\}|/api/v1/money/articles/\{nm_id\}", r"^money\.profit\.after_source_ads$|^profit_variants\.after_source_ads$", "derived", "profit_before_ads - max(ads.source_spend, ads.allocated_spend)", "Profit after source ads"),
    (r"/api/v1/money/cards/\{sku_id\}|/api/v1/money/articles/\{nm_id\}", r"^money\.profit\.with_allocated_overhead$|^profit_variants\.with_allocated_overhead$", "derived", "profit_after_source_ads - wb_expenses.allocated_overhead", "Includes allocated account-level overhead"),
    (r"/api/v1/money/cards/\{sku_id\}|/api/v1/money/articles/\{nm_id\}", r"^money\.profit\.margin_after_ads_percent$", "ratio", "profit.after_ads / money.revenue * 100", "Card margin"),
    (r"/api/v1/money/cards/\{sku_id\}|/api/v1/money/articles/\{nm_id\}", r"^money\.profit\.roi_after_ads_percent$|^money\.profit\.roi_on_cogs_percent$", "ratio", "profit.after_ads / cogs.estimated_cogs * 100", "ROI on COGS"),
    (r"/api/v1/money/cards/\{sku_id\}|/api/v1/money/articles/\{nm_id\}", r"^money\.profit\.stock_roi_percent$", "ratio", "profit.after_ads / stock.stock_value * 100", "ROI on inventory value"),
    (r"/api/v1/money/cards/\{sku_id\}|/api/v1/money/articles/\{nm_id\}", r"^money\.profit\.roas_percent$", "ratio", "revenue / ads.source_spend * 100", "ROAS as percent-like multiplier"),
    (r"/api/v1/money/cards/\{sku_id\}|/api/v1/money/articles/\{nm_id\}", r"^stock\.stock_value$", "derived", "stock.quantity * cogs.unit_cost", "Inventory value fallback"),
    (r"/api/v1/money/cards/\{sku_id\}|/api/v1/money/articles/\{nm_id\}", r"^stock\.in_transit_value$", "derived", "stock.in_transit_qty * cogs.unit_cost", "Transit inventory value"),
    (r"/api/v1/money/cards/\{sku_id\}|/api/v1/money/articles/\{nm_id\}", r"^price\.safe_price_gap$|^price\.safe_price_gap_final$", "derived", "(current_discounted_price or current_price or average_sale_price) - break_even_price", "Negative means below break-even"),
    (r"/api/v1/money/cards/\{sku_id\}|/api/v1/money/articles/\{nm_id\}", r"^reconciliation\.difference_amount$", "derived", "finance_report_revenue_total - mart_revenue_total", "Finance vs mart revenue mismatch"),
    (r"/api/v1/money/cards/\{sku_id\}|/api/v1/money/articles/\{nm_id\}", r"^reconciliation\.difference_ratio_percent$", "ratio", "difference_amount / mart_revenue_total * 100", "Mismatch ratio"),
    ("/api/v1/dashboard/data-health", r"^active_sku_count$", "count", "count(core_sku where is_active=true)", "Active SKU universe"),
    ("/api/v1/dashboard/data-health", r"^active_sku_with_manual_cost_count$", "count", "count(active core_sku with exists manual_cost)", "Cost-linked active SKUs"),
    ("/api/v1/dashboard/data-health", r"^placeholder_manual_cost_count$", "count", "count(manual_costs where is_placeholder=true or supplier='AUTO_TEMPLATE')", "Placeholder costs only"),
    ("/api/v1/dashboard/data-health", r"^real_manual_cost_count$", "count", "count(manual_costs where cost_source='manual_upload' and not placeholder)", "Supplier-confirmed uploads"),
    ("/api/v1/dashboard/data-health", r"^trusted_manual_cost_count$", "count", "count(non-placeholder manual_costs accepted by current cost_trust_policy)", "Policy-aware trusted cost count"),
    ("/api/v1/dashboard/data-health", r"^revenue_rows_with_cost$", "count", "count(mart_sku_daily rows with final_revenue>0 and has_manual_cost=true)", "Rows covered by any cost"),
    ("/api/v1/dashboard/data-health", r"^revenue_rows_without_cost$", "count", "count(mart_sku_daily rows with final_revenue>0 and has_manual_cost=false)", "Rows lacking cost"),
    ("/api/v1/dashboard/data-health", r"^revenue_with_cost$", "derived", "sum(mart_sku_daily.final_revenue where has_manual_cost=true)", "Revenue covered by any cost"),
    ("/api/v1/dashboard/data-health", r"^revenue_without_cost$", "derived", "sum(mart_sku_daily.final_revenue where has_manual_cost=false)", "Revenue without any cost"),
    ("/api/v1/dashboard/data-health", r"^revenue_with_real_cost$", "derived", "sum(mart_sku_daily.final_revenue where has_real_manual_cost=true)", "Revenue with supplier-confirmed cost"),
    ("/api/v1/dashboard/data-health", r"^sku_cost_coverage_percent$", "ratio", "active_sku_with_manual_cost_count / active_sku_count * 100", "Any cost SKU coverage"),
    ("/api/v1/dashboard/data-health", r"^revenue_cost_coverage_percent$", "ratio", "revenue_with_cost / (revenue_with_cost + revenue_without_cost) * 100", "Any cost revenue coverage"),
    ("/api/v1/dashboard/data-health", r"^supplier_confirmed_revenue_coverage_percent$|^real_revenue_cost_coverage_percent$", "ratio", "revenue_with_real_cost / (revenue_with_cost + revenue_without_cost) * 100", "Supplier-confirmed coverage"),
    ("/api/v1/dashboard/data-health", r"^trusted_revenue_cost_coverage_percent$", "ratio", "trusted cost revenue / (revenue_with_cost + revenue_without_cost) * 100", "Policy-aware trusted coverage"),
    ("/api/v1/dashboard/owner", r"^revenue$", "derived", "sum(control_rows.revenue)", "Owner-level revenue"),
    ("/api/v1/dashboard/owner", r"^net_profit$", "derived", "sum(control_rows.net_profit)", "Owner-level profit"),
    ("/api/v1/dashboard/owner", r"^margin_percent$", "ratio", "net_profit / revenue * 100", "Owner margin"),
    ("/api/v1/dashboard/owner", r"^roi_percent$", "ratio", "net_profit / sum(profit_rows.estimated_cogs) * 100", "Owner ROI on COGS"),
    ("/api/v1/dashboard/owner", r"^ad_spend$", "derived", "sum(control_rows.ad_spend) with source ads fallback", "Owner ad spend"),
    ("/api/v1/dashboard/owner", r"^stock_value$", "derived", "sum(control_rows.stock_value)", "Owner inventory value"),
    ("/api/v1/dashboard/owner", r"^overstock_value$", "derived", "sum(control_rows.stock_value where sku_status='LIQUIDATE')", "Owner overstock value"),
    ("/api/v1/pricing/safety", r"^items\[\]\.break_even_price$", "derived", "fixed_unit_cost / (1 - variable_rate)", "fixed_unit_cost = total_unit_cost + per-unit variable expenses"),
    ("/api/v1/pricing/safety", r"^items\[\]\.target_margin_price$", "derived", "fixed_unit_cost / (1 - variable_rate - target_margin_rate)", "Computed when denominator > 0.05"),
    ("/api/v1/pricing/safety", r"^items\[\]\.safe_price_gap$", "derived", "reference_price - break_even_price", "reference_price = current_discounted_price or current_price or average_sale_price"),
    ("/api/v1/pricing/safety", r"^items\[\]\.estimated_margin_at_current_price$", "ratio", "(reference_price - break_even_price) / reference_price * 100", "Estimated margin at displayed price"),
    ("/api/v1/dashboard/article-audit", r"^finance\.estimated_profit_before_ads$", "derived", "realized_revenue + additional_payments - commission - acquiring_fee - logistics - paid_acceptance - storage - penalties - deductions - estimated_cogs", "Article pre-ads profit"),
    ("/api/v1/dashboard/article-audit", r"^reconciliation\.difference_amount$", "derived", "finance_report_revenue_total - mart_revenue_total", "Article finance mismatch"),
    ("/api/v1/dashboard/article-audit", r"^reconciliation\.difference_ratio_percent$|^reconciliation\.difference_ratio$", "ratio", "difference_amount / mart_revenue_total * 100", "Article mismatch ratio"),
    ("/api/v1/inventory/purchase-plan", r"^items\[\]\.recommended_qty$", "derived", "ceil(max(required_stock - available_stock - in_transit_qty, 0) / pack_multiple) * pack_multiple", "required_stock = sales_velocity_daily * (lead_time_days + safety_days)"),
    ("/api/v1/inventory/purchase-plan", r"^items\[\]\.required_cash$", "derived", "recommended_qty * unit_cost", "Unit cost from control row / core_sku"),
    ("/api/v1/ads/efficiency", r"^items\[\]\.drr_percent$", "ratio", "ad_spend / revenue * 100", "Ads DRR from control rows"),
]


def load_json(path: Path) -> Any:
    return json.loads(path.read_text())


def resolve_contract_path(method: str, path_template: str, contract_map: dict[tuple[str, str], Any]) -> str:
    candidates = [path_template]
    if not path_template.startswith("/api/v1") and path_template != "/health":
        candidates.append(f"/api/v1{path_template}")
    for candidate in candidates:
        if (method.upper(), candidate) in contract_map:
            return candidate
    return candidates[-1]


def lookup_source(full_path: str) -> tuple[str, str, str]:
    matches = [item for item in SOURCE_GROUPS if full_path.startswith(item[0])]
    if not matches:
        return "", "", ""
    prefix, layer, tables, note = sorted(matches, key=lambda item: len(item[0]), reverse=True)[0]
    _ = prefix
    return layer, tables, note


def lookup_service(full_path: str) -> str:
    if full_path in SERVICE_MAP:
        return SERVICE_MAP[full_path]
    matches = [value for key, value in SERVICE_MAP.items() if full_path.startswith(key.rstrip("{}nm_idsku_id"))]
    return matches[0] if matches else ""


def formula_for(full_path: str, normalized_path: str) -> tuple[str, str, str]:
    for endpoint_pattern, path_pattern, formula_type, formula_text, inputs_hint in FORMULA_RULES:
        if re.search(endpoint_pattern, full_path) and re.search(path_pattern, normalized_path):
            return formula_type, formula_text, inputs_hint
    return "passthrough", "Direct field from endpoint response / DB read model; no extra formula layer documented for this field.", ""


def flatten_shape(value: Any, path: str = "") -> dict[str, str]:
    result: dict[str, str] = {}
    if isinstance(value, dict):
        for key, item in value.items():
            child = f"{path}.{key}" if path else key
            result.update(flatten_shape(item, child))
    elif isinstance(value, list):
        child = f"{path}[]" if path else "[]"
        if value:
            result.update(flatten_shape(value[0], child))
    elif isinstance(value, str):
        result[path] = value
    return result


def flatten_numeric(value: Any, path: str = "", normalized_path: str = "") -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if isinstance(value, dict):
        for key, item in value.items():
            child_path = f"{path}.{key}" if path else key
            child_norm = f"{normalized_path}.{key}" if normalized_path else key
            rows.extend(flatten_numeric(item, child_path, child_norm))
    elif isinstance(value, list):
        child_norm = f"{normalized_path}[]" if normalized_path else "[]"
        for idx, item in enumerate(value):
            child_path = f"{path}[{idx}]" if path else f"[{idx}]"
            rows.extend(flatten_numeric(item, child_path, child_norm))
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        rows.append(
            {
                "json_path": path,
                "normalized_json_path": normalized_path,
                "value": value,
                "value_type": type(value).__name__,
            }
        )
    return rows


def get_body(response_file: Path) -> tuple[dict[str, Any], Any]:
    payload = load_json(response_file)
    response_meta = payload.get("response", {})
    return response_meta, response_meta.get("body")


def money_flow_amount(flow_items: list[dict[str, Any]], code: str) -> float:
    for item in flow_items:
        if item.get("code") == code:
            return float(item.get("amount") or 0.0)
    return 0.0


def append_header(sheet, headers: list[str]) -> None:
    sheet.append(headers)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"


def auto_width(sheet) -> None:
    for idx, column in enumerate(sheet.columns, start=1):
        max_len = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        sheet.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 60)


def build_manual_check_sheets(wb: Workbook, bundle: Path) -> None:
    responses = bundle / "responses"

    summary = get_body(responses / "get_money_summary.json")[1]
    health = get_body(responses / "get_dashboard_data_health.json")[1]
    owner = get_body(responses / "get_dashboard_owner.json")[1]
    card = get_body(responses / "get_money_cards_sku_id.json")[1]
    article = get_body(responses / "get_dashboard_article_audit.json")[1]
    pricing = get_body(responses / "get_pricing_safety.json")[1]["items"][0]

    sheet = wb.create_sheet("Check_Money_Summary")
    append_header(sheet, ["metric", "actual", "recomputed", "delta", "formula", "inputs"])
    incoming = summary["money_flow"]["incoming"]
    outgoing = summary["money_flow"]["outgoing"]
    cash_stock = summary["money_flow"]["cash_and_stock"]
    revenue = float(summary["kpis"]["revenue"])
    profit = float(summary["kpis"]["profit_after_source_ads"])
    stock_value = float(summary["kpis"]["stock_value"])
    ads_source = float(summary["kpis"]["ads_source_spend"])
    ads_alloc = float(summary["kpis"]["ads_allocated_spend"])
    sheet.append(["revenue", revenue, float(summary["revenue_sources"]["operational_revenue"]), revenue - float(summary["revenue_sources"]["operational_revenue"]), "summary.kpis.revenue == revenue_sources.operational_revenue", "Same source rows"])
    sheet.append(["margin_percent", float(summary["kpis"]["margin_percent"]), (profit / revenue * 100) if revenue else 0.0, float(summary["kpis"]["margin_percent"]) - ((profit / revenue * 100) if revenue else 0.0), "profit_after_source_ads / revenue * 100", f"profit={profit}; revenue={revenue}"])
    sheet.append(["stock_roi_percent", float(summary["kpis"]["stock_roi_percent"]), (profit / stock_value * 100) if stock_value else 0.0, float(summary["kpis"]["stock_roi_percent"]) - ((profit / stock_value * 100) if stock_value else 0.0), "profit_after_source_ads / stock_value * 100", f"profit={profit}; stock_value={stock_value}"])
    sheet.append(["roas_percent", float(summary["kpis"]["roas_percent"]), (revenue / ads_source * 100) if ads_source else 0.0, float(summary["kpis"]["roas_percent"]) - ((revenue / ads_source * 100) if ads_source else 0.0), "revenue / ads_source_spend * 100", f"revenue={revenue}; ads_source={ads_source}"])
    sheet.append(["ads_allocation_percent", float(summary["quality"]["ads_allocation_percent"]), (ads_alloc / ads_source * 100) if ads_source else 100.0, float(summary["quality"]["ads_allocation_percent"]) - ((ads_alloc / ads_source * 100) if ads_source else 100.0), "ads_allocated_spend / ads_source_spend * 100", f"ads_alloc={ads_alloc}; ads_source={ads_source}"])
    sheet.append(["cash_on_wb", float(summary["kpis"]["cash_on_wb"]), money_flow_amount(cash_stock, "available_for_withdraw"), float(summary["kpis"]["cash_on_wb"]) - money_flow_amount(cash_stock, "available_for_withdraw"), "cross-check against money_flow.cash_and_stock[available_for_withdraw]", "Uses same balance snapshot"])
    sheet.append(["stock_value", float(summary["kpis"]["stock_value"]), money_flow_amount(cash_stock, "stock_value"), float(summary["kpis"]["stock_value"]) - money_flow_amount(cash_stock, "stock_value"), "cross-check against money_flow.cash_and_stock[stock_value]", "Derived inventory value"])
    sheet.append(["direct_wb_expenses", float(summary["kpis"]["direct_wb_expenses"]), float(summary["kpis"]["wb_expenses_total"]), float(summary["kpis"]["direct_wb_expenses"]) - float(summary["kpis"]["wb_expenses_total"]), "kpis.direct_wb_expenses == kpis.wb_expenses_total", "Direct WB expense pool"])

    sheet = wb.create_sheet("Check_Data_Health")
    append_header(sheet, ["metric", "actual", "recomputed", "delta", "formula", "inputs"])
    active_sku = float(health["active_sku_count"])
    active_with_cost = float(health["active_sku_with_manual_cost_count"])
    rev_with_cost = float(health["revenue_with_cost"])
    rev_without_cost = float(health["revenue_without_cost"])
    rev_real_cost = float(health["revenue_with_real_cost"])
    denom = rev_with_cost + rev_without_cost
    sheet.append(["sku_cost_coverage_percent", float(health["sku_cost_coverage_percent"]), (active_with_cost / active_sku * 100) if active_sku else 0.0, float(health["sku_cost_coverage_percent"]) - ((active_with_cost / active_sku * 100) if active_sku else 0.0), "active_sku_with_manual_cost_count / active_sku_count * 100", f"{active_with_cost} / {active_sku}"])
    sheet.append(["revenue_cost_coverage_percent", float(health["revenue_cost_coverage_percent"]), (rev_with_cost / denom * 100) if denom else 0.0, float(health["revenue_cost_coverage_percent"]) - ((rev_with_cost / denom * 100) if denom else 0.0), "revenue_with_cost / (revenue_with_cost + revenue_without_cost) * 100", f"{rev_with_cost} / {denom}"])
    sheet.append(["supplier_confirmed_revenue_coverage_percent", float(health["supplier_confirmed_revenue_coverage_percent"]), (rev_real_cost / denom * 100) if denom else 0.0, float(health["supplier_confirmed_revenue_coverage_percent"]) - ((rev_real_cost / denom * 100) if denom else 0.0), "revenue_with_real_cost / (revenue_with_cost + revenue_without_cost) * 100", f"{rev_real_cost} / {denom}"])
    sheet.append(["trusted_revenue_cost_coverage_percent", float(health["trusted_revenue_cost_coverage_percent"]), float(health["trusted_revenue_cost_coverage_percent"]), 0.0, "Policy-aware trusted coverage", "Direct check from health payload"])

    sheet = wb.create_sheet("Check_Card_Detail")
    append_header(sheet, ["metric", "actual", "recomputed", "delta", "formula", "inputs"])
    card_money = card["money"]
    qty = float(card["stock"]["quantity"])
    unit_cost = float(card_money["cogs"]["unit_cost"])
    in_transit_qty = float(card["stock"]["in_transit_qty"])
    source_spend = float(card_money["ads"]["source_spend"])
    alloc_spend = float(card_money["ads"]["allocated_spend"])
    before_ads = float(card_money["profit"]["before_ads"])
    after_alloc = float(card_money["profit"]["after_allocated_ads"])
    after_source = float(card_money["profit"]["after_source_ads"])
    allocated_overhead = float(card_money["wb_expenses"]["allocated_overhead"])
    estimated_cogs = float(card_money["cogs"]["estimated_cogs"])
    revenue_card = float(card_money["revenue"])
    current_disc = float(card["price"]["current_discounted_price"])
    break_even = float(card["price"]["break_even_price"])
    rec_diff = float(card["reconciliation"]["difference_amount"])
    mart_rev = float(card["reconciliation"]["mart_revenue_total"])
    finance_rev = float(card["reconciliation"]["finance_report_revenue_total"])
    sheet.append(["stock.stock_value", float(card["stock"]["stock_value"]), qty * unit_cost, float(card["stock"]["stock_value"]) - (qty * unit_cost), "stock.quantity * cogs.unit_cost", f"{qty} * {unit_cost}"])
    sheet.append(["stock.in_transit_value", float(card["stock"]["in_transit_value"]), in_transit_qty * unit_cost, float(card["stock"]["in_transit_value"]) - (in_transit_qty * unit_cost), "stock.in_transit_qty * cogs.unit_cost", f"{in_transit_qty} * {unit_cost}"])
    sheet.append(["profit.after_allocated_ads", after_alloc, before_ads - alloc_spend, after_alloc - (before_ads - alloc_spend), "profit.before_ads - ads.allocated_spend", f"{before_ads} - {alloc_spend}"])
    sheet.append(["profit.after_source_ads", after_source, before_ads - max(source_spend, alloc_spend), after_source - (before_ads - max(source_spend, alloc_spend)), "profit.before_ads - max(ads.source_spend, ads.allocated_spend)", f"{before_ads} - max({source_spend}, {alloc_spend})"])
    sheet.append(["profit.with_allocated_overhead", float(card_money["profit"]["with_allocated_overhead"]), after_source - allocated_overhead, float(card_money["profit"]["with_allocated_overhead"]) - (after_source - allocated_overhead), "profit.after_source_ads - wb_expenses.allocated_overhead", f"{after_source} - {allocated_overhead}"])
    sheet.append(["profit.margin_after_ads_percent", float(card_money["profit"]["margin_after_ads_percent"]), (after_source / revenue_card * 100) if revenue_card else 0.0, float(card_money["profit"]["margin_after_ads_percent"]) - ((after_source / revenue_card * 100) if revenue_card else 0.0), "profit.after_ads / revenue * 100", f"{after_source} / {revenue_card}"])
    sheet.append(["profit.roi_on_cogs_percent", float(card_money["profit"]["roi_on_cogs_percent"]), (after_source / estimated_cogs * 100) if estimated_cogs else 0.0, float(card_money["profit"]["roi_on_cogs_percent"]) - ((after_source / estimated_cogs * 100) if estimated_cogs else 0.0), "profit.after_ads / estimated_cogs * 100", f"{after_source} / {estimated_cogs}"])
    sheet.append(["profit.stock_roi_percent", float(card_money["profit"]["stock_roi_percent"]), (after_source / float(card['stock']['stock_value']) * 100) if float(card["stock"]["stock_value"]) else 0.0, float(card_money["profit"]["stock_roi_percent"]) - ((after_source / float(card['stock']['stock_value']) * 100) if float(card["stock"]["stock_value"]) else 0.0), "profit.after_ads / stock.stock_value * 100", f"{after_source} / {float(card['stock']['stock_value'])}"])
    sheet.append(["profit.roas_percent", float(card_money["profit"]["roas_percent"]), (revenue_card / source_spend * 100) if source_spend else 0.0, float(card_money["profit"]["roas_percent"]) - ((revenue_card / source_spend * 100) if source_spend else 0.0), "revenue / ads.source_spend * 100", f"{revenue_card} / {source_spend}"])
    sheet.append(["ads.drr_percent_source", float(card_money["ads"]["drr_percent_source"]), (source_spend / revenue_card * 100) if revenue_card else 0.0, float(card_money["ads"]["drr_percent_source"]) - ((source_spend / revenue_card * 100) if revenue_card else 0.0), "ads.source_spend / revenue * 100", f"{source_spend} / {revenue_card}"])
    sheet.append(["price.safe_price_gap", float(card["price"]["safe_price_gap"]), current_disc - break_even, float(card["price"]["safe_price_gap"]) - (current_disc - break_even), "current_discounted_price - break_even_price", f"{current_disc} - {break_even}"])
    sheet.append(["reconciliation.difference_amount", rec_diff, finance_rev - mart_rev, rec_diff - (finance_rev - mart_rev), "finance_report_revenue_total - mart_revenue_total", f"{finance_rev} - {mart_rev}"])
    sheet.append(["reconciliation.difference_ratio_percent", float(card["reconciliation"]["difference_ratio_percent"]), ((finance_rev - mart_rev) / mart_rev * 100) if mart_rev else 0.0, float(card["reconciliation"]["difference_ratio_percent"]) - (((finance_rev - mart_rev) / mart_rev * 100) if mart_rev else 0.0), "difference_amount / mart_revenue_total * 100", f"{finance_rev - mart_rev} / {mart_rev}"])

    sheet = wb.create_sheet("Check_Price_Safety")
    append_header(sheet, ["metric", "actual", "recomputed", "delta", "formula", "inputs"])
    ref_price = float(pricing["current_discounted_price"] or pricing["current_price"] or pricing["average_sale_price"] or 0.0)
    sheet.append(["safe_price_gap", float(pricing["safe_price_gap"]), ref_price - float(pricing["break_even_price"]), float(pricing["safe_price_gap"]) - (ref_price - float(pricing["break_even_price"])), "reference_price - break_even_price", f"{ref_price} - {float(pricing['break_even_price'])}"])
    sheet.append(["estimated_margin_at_current_price", float(pricing["estimated_margin_at_current_price"]), ((ref_price - float(pricing["break_even_price"])) / ref_price * 100) if ref_price else 0.0, float(pricing["estimated_margin_at_current_price"]) - (((ref_price - float(pricing["break_even_price"])) / ref_price * 100) if ref_price else 0.0), "(reference_price - break_even_price) / reference_price * 100", f"({ref_price} - {float(pricing['break_even_price'])}) / {ref_price}"])

    sheet = wb.create_sheet("Check_Article_Audit")
    append_header(sheet, ["metric", "actual", "recomputed", "delta", "formula", "inputs"])
    finance = article["finance"]
    rec = article["reconciliation"]
    profit_before_ads = float(finance["realized_revenue"]) + float(finance["additional_payments"]) - float(finance["commission"]) - float(finance["acquiring_fee"]) - float(finance["logistics"]) - float(finance["paid_acceptance"]) - float(finance["storage"]) - float(finance["penalties"]) - float(finance["deductions"]) - float(finance["estimated_cogs"])
    sheet.append(["finance.estimated_profit_before_ads", float(finance["estimated_profit_before_ads"]), profit_before_ads, float(finance["estimated_profit_before_ads"]) - profit_before_ads, "realized_revenue + additional_payments - commission - acquiring_fee - logistics - paid_acceptance - storage - penalties - deductions - estimated_cogs", "Uses finance block numbers"])
    sheet.append(["reconciliation.difference_amount", float(rec["difference_amount"]), float(rec["finance_report_revenue_total"]) - float(rec["mart_revenue_total"]), float(rec["difference_amount"]) - (float(rec["finance_report_revenue_total"]) - float(rec["mart_revenue_total"])), "finance_report_revenue_total - mart_revenue_total", f"{float(rec['finance_report_revenue_total'])} - {float(rec['mart_revenue_total'])}"])
    sheet.append(["reconciliation.difference_ratio_percent", float(rec["difference_ratio_percent"]), ((float(rec["finance_report_revenue_total"]) - float(rec["mart_revenue_total"])) / float(rec["mart_revenue_total"]) * 100) if float(rec["mart_revenue_total"]) else 0.0, float(rec["difference_ratio_percent"]) - (((float(rec["finance_report_revenue_total"]) - float(rec["mart_revenue_total"])) / float(rec["mart_revenue_total"]) * 100) if float(rec["mart_revenue_total"]) else 0.0), "difference_amount / mart_revenue_total * 100", "Uses reconciliation block numbers"])

    sheet = wb.create_sheet("Check_Owner")
    append_header(sheet, ["metric", "actual", "recomputed", "delta", "formula", "inputs"])
    owner_rev = float(owner["revenue"])
    owner_profit = float(owner["net_profit"])
    sheet.append(["margin_percent", float(owner["margin_percent"]), (owner_profit / owner_rev * 100) if owner_rev else 0.0, float(owner["margin_percent"]) - ((owner_profit / owner_rev * 100) if owner_rev else 0.0), "net_profit / revenue * 100", f"{owner_profit} / {owner_rev}"])
    sheet.append(["blocked_data_sku_count", float(owner["blocked_data_sku_count"]), float(owner["action_summary"]["data_blocked_count"]), float(owner["blocked_data_sku_count"]) - float(owner["action_summary"]["data_blocked_count"]), "cross-check owner block count vs action_summary.data_blocked_count", "Same owner payload"])


def build_workbook(bundle: Path, output_xlsx: Path) -> tuple[Path, Path]:
    executed = load_json(bundle / "executed_live_index.json")
    contracts = load_json(bundle / "endpoint_contract_index.json")
    contract_map = {(item["method"].upper(), item["path"]): item for item in contracts}

    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws["A1"] = "Finance Endpoint Formula Audit"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Bundle: {bundle.name}"
    ws["A3"] = "This workbook contains: (1) endpoint index, (2) numeric field dump for every live executed endpoint sample, (3) manual calculator check sheets for key business endpoints."
    ws["A4"] = "Columns in Numeric_Dump: endpoint, json_path, real value, source tables, backend service/function, formula type, formula text, inputs hint."
    ws["A5"] = "Numbers come from the live audit bundle, not synthetic examples."
    ws["A6"] = "For direct/pass-through endpoints the formula column states that the field is read directly from DB/read-model without extra business math."
    ws.column_dimensions["A"].width = 140

    index = wb.create_sheet("Endpoint_Index")
    append_header(index, ["method", "path", "response_file", "status_code", "elapsed_ms", "query_string", "layer", "db_tables", "service_function", "notes"])

    numeric = wb.create_sheet("Numeric_Dump")
    append_header(
        numeric,
        [
            "method",
            "path",
            "response_file",
            "query_string",
            "json_path",
            "normalized_json_path",
            "value",
            "value_type",
            "expected_type",
            "layer",
            "db_tables",
            "service_function",
            "formula_type",
            "formula_text",
            "formula_inputs_hint",
        ],
    )

    for item in executed:
        method = item["method"].upper()
        path_template = item["path_template"]
        contract_path = resolve_contract_path(method, path_template, contract_map)
        response_file = bundle / "responses" / item["response_file"]
        response_meta, body = get_body(response_file)
        shape_map = flatten_shape(response_meta.get("shape"))
        layer, tables, notes = lookup_source(contract_path)
        service_fn = lookup_service(contract_path)
        index.append(
            [
                method,
                contract_path,
                item["response_file"],
                item["status_code"],
                item["elapsed_ms"],
                item.get("query_string") or "",
                layer,
                tables,
                service_fn,
                notes,
            ]
        )
        for row in flatten_numeric(body):
            formula_type, formula_text, inputs_hint = formula_for(contract_path, row["normalized_json_path"])
            numeric.append(
                [
                    method,
                    contract_path,
                    item["response_file"],
                    item.get("query_string") or "",
                    row["json_path"],
                    row["normalized_json_path"],
                    row["value"],
                    row["value_type"],
                    shape_map.get(row["normalized_json_path"], ""),
                    layer,
                    tables,
                    service_fn,
                    formula_type,
                    formula_text,
                    inputs_hint,
                ]
            )

    build_manual_check_sheets(wb, bundle)

    for sheet in wb.worksheets:
        auto_width(sheet)

    wb.save(output_xlsx)

    readme_path = output_xlsx.with_suffix(".md")
    readme_path.write_text(
        "\n".join(
            [
                "# Finance Endpoint Formula Audit Workbook",
                "",
                f"- Bundle source: `{bundle}`",
                f"- Workbook: `{output_xlsx.name}`",
                "",
                "## Sheets",
                "- `Endpoint_Index`: each executed live endpoint sample and its source layer",
                "- `Numeric_Dump`: flattened numeric values for every endpoint sample with source/formula notes",
                "- `Check_Money_Summary`: manual calculator checks for the summary endpoint",
                "- `Check_Data_Health`: manual checks for coverage and health math",
                "- `Check_Card_Detail`: manual checks for stock, ads, profit, price and reconciliation",
                "- `Check_Price_Safety`: manual checks for price safety sample",
                "- `Check_Article_Audit`: manual checks for article audit sample",
                "- `Check_Owner`: manual checks for owner summary sample",
            ]
        )
    )

    zip_path = output_xlsx.with_suffix(".zip")
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.write(output_xlsx, arcname=output_xlsx.name)
        zf.write(readme_path, arcname=readme_path.name)
    return output_xlsx, zip_path


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--bundle", type=Path, default=DEFAULT_BUNDLE)
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()

    bundle = args.bundle
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = args.output or EXPORT_ROOT / f"finance_endpoint_formula_audit_{timestamp}.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    xlsx_path, zip_path = build_workbook(bundle, output)
    print(json.dumps({"xlsx": str(xlsx_path), "zip": str(zip_path)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
